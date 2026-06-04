from __future__ import annotations

import json
import os
import re
import base64
import binascii
import csv
import secrets
import time
import uuid
import zipfile
from dataclasses import dataclass
from functools import lru_cache, wraps
from io import BytesIO, StringIO
from datetime import datetime, timezone
from hashlib import sha1, sha256
from threading import Lock
from typing import Any, Callable
from xml.etree import ElementTree
from xml.sax.saxutils import escape as xml_escape

from fastapi import Depends, FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import PlainTextResponse, StreamingResponse
import httpx
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from .database import (
    Base,
    SessionLocal,
    engine,
    finish_slow_query_capture,
    get_db,
    persist_slow_query_observations,
    start_slow_query_capture,
)
from .migrations.sprint55_migrate_ai_config import run_sprint55_migration
from .models import (
    Attribute,
    AttributeChange,
    AIAgentConfig,
    AuditLog,
    Brand,
    CapabilityAgentMapping,
    CapabilityModelMapping,
    CapabilityMapping,
    Category,
    CategoryAttribute,
    CategoryLibrary,
    FeaturePermission,
    LLMProviderConfig,
    Material,
    MaterialCodeChangeBatch,
    MaterialCodeChangeDetail,
    MaterialCodeMapping,
    MaterialCodeRuleVersion,
    MaterialCodeSerial,
    MaterialLibrary,
    MaterialLibraryAdminRole,
    MaterialLibraryCategoryLibrary,
    Model,
    ModelConfig,
    ProductName,
    ProductNameCodeSequence,
    Rule,
    RuleCategory,
    Role,
    RoleCodeSequence,
    RoleUser,
    SlowQueryLog,
    SystemConfig,
    TelemetryWebVital,
    TracerSpan,
    User,
    WorkflowApplication,
    WorkflowHistory,
)
from .schemas import (
    AiMaterialAddConfirmIn,
    AiMaterialAddPreviewIn,
    AuditLogListOut,
    AuditLogOut,
    AuthLoginIn,
    AuthUserOut,
    AttributeIn,
    AttributeOut,
    AttributeUpdate,
    BrandIn,
    BrandLogo,
    BrandOut,
    BrandUpdate,
    CapabilityMappingIn,
    CapabilityMappingOut,
    CapabilityMappingCreate,
    CapabilityMappingRead,
    CapabilityMappingUpdate,
    CategoryIn,
    CategoryAttributeCreate,
    CategoryAttributeRead,
    CategoryAttributeUpdate,
    CategoryLibraryIn,
    CategoryLibraryOut,
    CategoryLibraryUpdate,
    CategoryOut,
    CategoryPropertyList,
    CategoryRecognitionBatchRequest,
    CategoryRecognitionJob,
    CategoryRecognitionJobResult,
    CategoryRecognitionRequest,
    CategoryRecognitionResponse,
    CategoryUpdate,
    ChangeOut,
    GatewayInvokeIn,
    GovernanceImportIn,
    GovernancePreviewIn,
    BatchActionIn,
    MaterialCodeChangeBatchOut,
    MaterialCodeChangePreviewListOut,
    MaterialCodeChangeRowOut,
    MaterialCodeMappingListOut,
    MaterialCodeMappingOut,
    MaterialGovernanceImportIn,
    MaterialGovernancePreviewIn,
    MaterialCategoryMatchIn,
    MaterialCategoryMatchOut,
    MaterialMatchIn,
    ModelCreate,
    ModelRead,
    ModelTestResult,
    ModelUpdate,
    MaterialIn,
    MaterialCodeRuleVersionIn,
    MaterialCodeRuleVersionListOut,
    MaterialCodeRuleVersionOut,
    MaterialLibraryIn,
    MaterialLibraryOut,
    MaterialLibraryUpdate,
    MaterialOut,
    MaterialTransitionIn,
    MaterialUpdate,
    ManualStopPurchaseIn,
    ProductNameIn,
    ProductNameOut,
    ProductNameStatusUpdate,
    ProductNameUpdate,
    RecommendIn,
    RecodePreviewIn,
    PermissionEntry,
    PasswordResetOut,
    EvaluateRequest,
    EvaluateResponse,
    EvaluateResult,
    RoleIn,
    RoleOut,
    RolePermissionsIn,
    RolePermissionsOut,
    RoleUpdate,
    RoleUserBindingIn,
    RoleUserReplaceIn,
    RuleCategoryRead,
    RuleCreate,
    RuleListResponse,
    RuleRead,
    RuleToggle,
    RuleUpdate,
    ReasonOption,
    SystemConfigIn,
    SystemIcon,
    SystemConfigOut,
    SlowQueryLogOut,
    TraceDetailOut,
    TraceSummaryOut,
    UserIn,
    UserOut,
    UserSummaryOut,
    UserUpdate,
    WebVitalsTelemetryIn,
    WebVitalsTelemetryOut,
    WorkflowActionIn,
    WorkflowApplicationIn,
    WorkflowApplicationOut,
    WorkflowHistoryOut,
)


API_VERSION = "15.0.0"

app = FastAPI(title="AI Material Management Platform", version=API_VERSION)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

_metrics_lock = Lock()
_http_request_total: dict[tuple[str, str, str], int] = {}
_http_request_duration_seconds: dict[tuple[str, str, str], dict[str, float]] = {}


def _escape_prometheus_label(value: str) -> str:
    return value.replace("\\", "\\\\").replace("\n", "\\n").replace('"', '\\"')


def _request_route_label(request: Request) -> str:
    route = request.scope.get("route")
    return str(getattr(route, "path", request.url.path))


def _record_http_request_metric(method: str, route: str, status_code: int, duration_seconds: float) -> None:
    key = (method.upper(), route, str(status_code))
    with _metrics_lock:
        _http_request_total[key] = _http_request_total.get(key, 0) + 1
        duration = _http_request_duration_seconds.setdefault(key, {"count": 0.0, "sum": 0.0})
        duration["count"] += 1.0
        duration["sum"] += duration_seconds


def _metric_labels(method: str, route: str, status_code: str) -> str:
    return (
        f'method="{_escape_prometheus_label(method)}",'
        f'route="{_escape_prometheus_label(route)}",'
        f'status_code="{_escape_prometheus_label(status_code)}"'
    )


def render_prometheus_metrics() -> str:
    with _metrics_lock:
        request_total = dict(_http_request_total)
        request_durations = {key: dict(value) for key, value in _http_request_duration_seconds.items()}
    lines = [
        "# HELP http_requests_total Total HTTP requests handled by the FastAPI backend.",
        "# TYPE http_requests_total counter",
    ]
    for (method, route, status_code), count in sorted(request_total.items()):
        lines.append(f"http_requests_total{{{_metric_labels(method, route, status_code)}}} {count}")
    lines.extend(
        [
            "# HELP http_request_duration_seconds HTTP request duration in seconds.",
            "# TYPE http_request_duration_seconds summary",
        ]
    )
    for (method, route, status_code), duration in sorted(request_durations.items()):
        labels = _metric_labels(method, route, status_code)
        lines.append(f"http_request_duration_seconds_count{{{labels}}} {int(duration['count'])}")
        lines.append(f"http_request_duration_seconds_sum{{{labels}}} {duration['sum']:.9f}")
    return "\n".join(lines) + "\n"


@app.middleware("http")
async def observability_middleware(request: Request, call_next: Callable):
    started_at = time.perf_counter()
    token = start_slow_query_capture()
    status_code = 500
    try:
        response = await call_next(request)
        status_code = response.status_code
        return response
    finally:
        observations = finish_slow_query_capture(token)
        persist_slow_query_observations(observations)
        _record_http_request_metric(
            request.method,
            _request_route_label(request),
            status_code,
            time.perf_counter() - started_at,
        )


@app.middleware("http")
async def operational_audit_middleware(request: Request, call_next: Callable):
    response = await call_next(request)
    mutating = request.method in {"POST", "PUT", "PATCH", "DELETE"} and request.url.path.startswith("/api/v1/")
    excluded = request.url.path.startswith("/api/v1/audit-logs") or request.url.path.startswith("/api/v1/auth/")
    if mutating and not excluded and response.status_code < 400:
        db = SessionLocal()
        try:
            ensure_audit_log_schema()
            try:
                auth = current_auth(request, db)
            except Exception:
                auth = None
            source = "AI" if "/ai/" in request.url.path else "human"
            add_audit_log(
                db,
                auth,
                request.url.path.removeprefix("/api/v1/"),
                request.method.lower(),
                {},
                {"status_code": response.status_code, "path": request.url.path},
                source,
            )
            db.commit()
        finally:
            db.close()
    return response

SEED_PRODUCT = {
    "name": "Sprint 3 A4 彩色激光打印机",
    "unit": "台",
    "category": "办公设备 / 打印机",
}
SEED_LIBRARY = {
    "code": "MLIB-DEFAULT",
    "name": "Default Material Library",
    "description": "Default library for sprint verification materials",
}
SEED_CATEGORY_LIBRARY = {
    "code": "CLIB-DEFAULT",
    "name": "Default Category Library",
    "description": "Default category library for sprint verification categories",
}
SEED_CATEGORY = {
    "code": "CAT-PRINTER",
    "name": "办公设备 / 打印机",
    "description": "Default category bound to the seed product name",
}
MATERIAL_STATUSES = {"normal", "stop_purchase", "stop_use"}
MATERIAL_TRANSITIONS = {("normal", "stop_purchase"), ("stop_purchase", "stop_use")}
AI_CAPABILITIES = {
    "material_add",
    "material_match",
    "category_match",
    "category_recognition",
    "material_analysis",
    "attr_recommend",
    "material_governance",
}
MODEL_PROVIDERS = {"dashscope", "azure", "openai", "vllm", "ollama", "deepseek", "moonshot", "custom"}
DEFAULT_CAPABILITY_MAPPINGS = {
    "material_add",
    "category_recognition",
    "material_match",
    "attr_recommend",
    "material_governance",
}
APPROVAL_MODES = {"simple", "multi_node"}
APPLICATION_TYPES = {"new_category", "new_material_code", "stop_purchase", "stop_use"}
TERMINAL_WORKFLOW_STATUSES = {"approved", "rejected"}
USER_STATUSES = {"active", "disabled"}
ACCOUNT_OWNERSHIPS = {"HCM", "local"}
ROLE_CODE_PATTERN = re.compile(r"^ROLE_(\d{3,})$")
PRODUCT_NAME_STATUSES = {"active", "inactive"}
PRODUCT_NAME_CODE_PATTERN = re.compile(r"^PM(\d{8,})$")
SYSTEM_CONFIG_KEY = "system_configuration"
DEFAULT_SYSTEM_NAME = "AI Material Management Platform"

# Simple in-memory cache for model capability lookups ( TTL = 5 seconds )
_model_capability_cache: dict[str, tuple[float, CapabilityMapping | None]] = {}
_CAPABILITY_CACHE_TTL = 5.0
DEFAULT_SYSTEM_ICON = {
    "filename": "default-system-icon.svg",
    "content_type": "image/svg+xml",
    "data_url": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 64 64'%3E%3Crect width='64' height='64' rx='14' fill='%23205493'/%3E%3Cpath d='M18 42V22h9l5 10 5-10h9v20h-7V31l-5 11h-4l-5-11v11z' fill='white'/%3E%3C/svg%3E",
}
DEFAULT_STOP_PURCHASE_REASONS = ["供应商停产", "质量风险停采", "战略替代物料", "采购目录清理"]
DEFAULT_STOP_USE_REASONS = ["长期无库存且无业务需求", "安全合规风险", "技术标准淘汰", "资产归档完成"]
DEFAULT_PROVIDER = {
    "provider": "mock",
    "model": "mock-material-governance-v1",
    "endpoint": "local://deterministic",
    "capabilities": ["material_add", "material_match", "category_recognition"],
}
KNOWN_BRANDS = ["华为", "Huawei", "HUAWEI", "联想", "Lenovo", "惠普", "HP", "戴尔", "Dell", "治理测试品牌"]
HCM_SEED_USERS = [
    {
        "username": "hcm_zhangsan",
        "display_name": "张三",
        "hcm_id": "HCM-1001",
        "unit": "华东事业部",
        "department": "采购管理部",
        "team": "标准化一组",
        "email": "zhangsan@example.com",
    },
    {
        "username": "hcm_lisi",
        "display_name": "李四",
        "hcm_id": "HCM-1002",
        "unit": "华北事业部",
        "department": "资产管理部",
        "team": "物料治理组",
        "email": "lisi@example.com",
    },
]
DEFAULT_RULE_CATEGORIES = [
    {
        "slug": "unit_normalization",
        "display_name_zh": "单位标准化",
        "display_name_en": "Unit Normalization",
        "description_zh": "将物料单位统一为标准写法。",
        "description_en": "Normalizes material units to standard values.",
        "icon": "ruler",
        "sort_order": 10,
    },
    {
        "slug": "brand_alias",
        "display_name_zh": "品牌别名归一",
        "display_name_en": "Brand Alias Normalization",
        "description_zh": "将品牌别名统一为标准品牌名称。",
        "description_en": "Maps brand aliases to canonical brand names.",
        "icon": "badge",
        "sort_order": 20,
    },
    {
        "slug": "title_cleaning",
        "display_name_zh": "标题格式清洗",
        "display_name_en": "Title Cleaning",
        "description_zh": "清理物料标题中的多余空格和非法格式。",
        "description_en": "Cleans extra whitespace and invalid title formatting.",
        "icon": "text-cursor-input",
        "sort_order": 30,
    },
    {
        "slug": "enum_validation",
        "display_name_zh": "枚举值校验",
        "display_name_en": "Enum Validation",
        "description_zh": "校验属性值是否位于允许枚举范围内。",
        "description_en": "Validates attribute values against allowed enum sets.",
        "icon": "list-checks",
        "sort_order": 40,
    },
    {
        "slug": "required_field_check",
        "display_name_zh": "必填字段检查",
        "display_name_en": "Required Field Check",
        "description_zh": "检查物料必填属性是否已填写。",
        "description_en": "Checks that required material attributes are present.",
        "icon": "asterisk",
        "sort_order": 50,
    },
    {
        "slug": "blackwhite_list",
        "display_name_zh": "黑白名单过滤",
        "display_name_en": "Blacklist and Whitelist",
        "description_zh": "根据关键词黑名单或白名单过滤物料。",
        "description_en": "Filters materials with blacklist or whitelist keywords.",
        "icon": "shield-check",
        "sort_order": 60,
    },
]
DEFAULT_RULES = [
    {
        "category_slug": "unit_normalization",
        "name": "KG 转 kg",
        "description": "将 KG 统一为 kg。",
        "pattern": "KG",
        "value": "kg",
        "options": {"match": "exact_ignore_case"},
        "priority": 10,
        "enabled": True,
    },
    {
        "category_slug": "unit_normalization",
        "name": "公斤 转 kg",
        "description": "将中文公斤统一为 kg。",
        "pattern": "公斤",
        "value": "kg",
        "options": {"match": "exact"},
        "priority": 20,
        "enabled": True,
    },
    {
        "category_slug": "brand_alias",
        "name": "苹果 转 Apple",
        "description": "将苹果品牌别名统一为 Apple。",
        "pattern": "苹果",
        "value": "Apple",
        "options": {"match": "exact"},
        "priority": 10,
        "enabled": True,
    },
    {
        "category_slug": "brand_alias",
        "name": "APPLE 转 Apple",
        "description": "将全大写 APPLE 统一为 Apple。",
        "pattern": "APPLE",
        "value": "Apple",
        "options": {"match": "exact_ignore_case"},
        "priority": 20,
        "enabled": True,
    },
    {
        "category_slug": "title_cleaning",
        "name": "压缩标题空格",
        "description": "将连续空白压缩为单个空格。",
        "pattern": "\\s+",
        "value": " ",
        "options": {"strip": True},
        "priority": 10,
        "enabled": True,
    },
    {
        "category_slug": "title_cleaning",
        "name": "清理标题首尾标点",
        "description": "移除标题首尾多余标点。",
        "pattern": "^[,，。.;；\\s]+|[,，。.;；\\s]+$",
        "value": "",
        "options": {"strip": True},
        "priority": 20,
        "enabled": True,
    },
    {
        "category_slug": "enum_validation",
        "name": "颜色枚举校验",
        "description": "颜色必须为 red 或 blue。",
        "pattern": "color",
        "value": "red",
        "options": {"field": "color", "allowed": ["red", "blue"]},
        "priority": 10,
        "enabled": True,
    },
    {
        "category_slug": "enum_validation",
        "name": "状态枚举校验",
        "description": "状态必须为 normal 或 spare。",
        "pattern": "status",
        "value": "normal",
        "options": {"field": "status", "allowed": ["normal", "spare"]},
        "priority": 20,
        "enabled": True,
    },
    {
        "category_slug": "required_field_check",
        "name": "电压必填",
        "description": "电压属性必须填写。",
        "pattern": "voltage",
        "value": "",
        "options": {"field": "voltage"},
        "priority": 10,
        "enabled": True,
    },
    {
        "category_slug": "required_field_check",
        "name": "功率必填",
        "description": "功率属性必须填写。",
        "pattern": "power",
        "value": "",
        "options": {"field": "power"},
        "priority": 20,
        "enabled": True,
    },
    {
        "category_slug": "blackwhite_list",
        "name": "禁用关键词拦截",
        "description": "标题中不得包含禁用关键词。",
        "pattern": "禁用",
        "value": "",
        "options": {"mode": "blacklist", "keywords": ["禁用", "FORBIDDEN"]},
        "priority": 10,
        "enabled": True,
    },
    {
        "category_slug": "blackwhite_list",
        "name": "淘汰关键词拦截",
        "description": "标题中不得包含淘汰关键词。",
        "pattern": "淘汰",
        "value": "",
        "options": {"mode": "blacklist", "keywords": ["淘汰", "obsolete"]},
        "priority": 20,
        "enabled": True,
    },
]
PERMISSION_CATALOG = [
    {"module": "material_archives", "permission_type": "directory", "permission_key": "directory.material_archives", "label": "Material Archives Directory"},
    {"module": "attribute_management", "permission_type": "directory", "permission_key": "directory.attribute_management", "label": "Attribute Management Directory"},
    {"module": "material_library", "permission_type": "directory", "permission_key": "directory.material_library", "label": "Material Library Directory"},
    {"module": "workflow", "permission_type": "directory", "permission_key": "directory.workflow", "label": "Workflow Directory"},
    {"module": "system_admin", "permission_type": "directory", "permission_key": "directory.system_admin", "label": "System Admin Directory"},
    {"module": "category_library", "permission_type": "directory", "permission_key": "directory.category_library", "label": "Category Library Directory"},
    {"module": "category_management", "permission_type": "directory", "permission_key": "directory.category_management", "label": "Category Management Directory"},
    {"module": "brand_management", "permission_type": "directory", "permission_key": "directory.brand_management", "label": "Brand Management Directory"},
    {"module": "product_name_management", "permission_type": "directory", "permission_key": "directory.product_name_management", "label": "Product Name Directory"},
    {"module": "material_archives", "permission_type": "button", "permission_key": "button.material_archives.create", "label": "Material Archive Create"},
    {"module": "material_archives", "permission_type": "button", "permission_key": "button.material_archives.edit", "label": "Material Archive Edit"},
    {"module": "material_archives", "permission_type": "button", "permission_key": "button.material_archives.delete", "label": "Material Archive Delete"},
    {"module": "material_archives", "permission_type": "button", "permission_key": "button.material_archives.import", "label": "Material Archive Import"},
    {"module": "material_archives", "permission_type": "button", "permission_key": "button.material_archives.export", "label": "Material Archive Export"},
    {"module": "material_archives", "permission_type": "button", "permission_key": "button.material_archives.approval", "label": "Material Lifecycle Approval"},
    {"module": "material_library", "permission_type": "button", "permission_key": "button.material_library.create", "label": "Material Library Create"},
    {"module": "material_library", "permission_type": "button", "permission_key": "button.material_library.edit", "label": "Material Library Edit"},
    {"module": "material_library", "permission_type": "button", "permission_key": "button.material_library.delete", "label": "Material Library Delete"},
    {"module": "material_library", "permission_type": "button", "permission_key": "button.material_library.import", "label": "Material Library Import"},
    {"module": "material_library", "permission_type": "button", "permission_key": "button.material_library.export", "label": "Material Library Export"},
    {"module": "material_library", "permission_type": "button", "permission_key": "button.material_library.approval", "label": "Material Library Approval"},
    {"module": "attribute_management", "permission_type": "button", "permission_key": "button.attribute_management.create", "label": "Attribute Create"},
    {"module": "attribute_management", "permission_type": "button", "permission_key": "button.attribute_management.edit", "label": "Attribute Edit"},
    {"module": "attribute_management", "permission_type": "button", "permission_key": "button.attribute_management.delete", "label": "Attribute Delete"},
    {"module": "attribute_management", "permission_type": "button", "permission_key": "button.attribute_management.import", "label": "Attribute Import"},
    {"module": "attribute_management", "permission_type": "button", "permission_key": "button.attribute_management.export", "label": "Attribute Export"},
    {"module": "category_library", "permission_type": "button", "permission_key": "button.category_library.create", "label": "Category Library Create"},
    {"module": "category_library", "permission_type": "button", "permission_key": "button.category_library.edit", "label": "Category Library Edit"},
    {"module": "category_library", "permission_type": "button", "permission_key": "button.category_library.delete", "label": "Category Library Delete"},
    {"module": "category_management", "permission_type": "button", "permission_key": "button.category_management.create", "label": "Category Create"},
    {"module": "category_management", "permission_type": "button", "permission_key": "button.category_management.edit", "label": "Category Edit"},
    {"module": "category_management", "permission_type": "button", "permission_key": "button.category_management.delete", "label": "Category Delete"},
    {"module": "workflow", "permission_type": "button", "permission_key": "button.workflow.submit", "label": "Workflow Submit"},
    {"module": "workflow", "permission_type": "button", "permission_key": "button.workflow.approve", "label": "Workflow Approve"},
    {"module": "workflow", "permission_type": "button", "permission_key": "button.workflow.reject", "label": "Workflow Reject"},
    {"module": "system_admin", "permission_type": "button", "permission_key": "button.users.create", "label": "Create Local User"},
    {"module": "system_admin", "permission_type": "button", "permission_key": "button.users.edit", "label": "Edit Local User"},
    {"module": "system_admin", "permission_type": "button", "permission_key": "button.users.delete", "label": "Delete Local User"},
    {"module": "system_admin", "permission_type": "button", "permission_key": "button.users.reset_password", "label": "Reset Local User Password"},
    {"module": "system_admin", "permission_type": "button", "permission_key": "button.roles.create", "label": "Create Role"},
    {"module": "system_admin", "permission_type": "button", "permission_key": "button.roles.edit", "label": "Edit Role"},
    {"module": "system_admin", "permission_type": "button", "permission_key": "button.roles.delete", "label": "Delete Role"},
    {"module": "system_admin", "permission_type": "button", "permission_key": "button.roles.bind_users", "label": "Bind Role Users"},
    {"module": "system_admin", "permission_type": "button", "permission_key": "button.roles.configure_permissions", "label": "Configure Role Permissions"},
    {"module": "material_library", "permission_type": "api", "permission_key": "api.GET./api/v1/material-libraries", "label": "GET /api/v1/material-libraries"},
    {"module": "material_library", "permission_type": "api", "permission_key": "api.POST./api/v1/material-libraries", "label": "POST /api/v1/material-libraries"},
    {"module": "material_library", "permission_type": "api", "permission_key": "api.GET./api/v1/material-libraries/{library_id}", "label": "GET /api/v1/material-libraries/{library_id}"},
    {"module": "material_library", "permission_type": "api", "permission_key": "api.PUT./api/v1/material-libraries/{library_id}", "label": "PUT /api/v1/material-libraries/{library_id}"},
    {"module": "material_library", "permission_type": "api", "permission_key": "api.DELETE./api/v1/material-libraries/{library_id}", "label": "DELETE /api/v1/material-libraries/{library_id}"},
    {"module": "material_archives", "permission_type": "api", "permission_key": "api.GET./api/v1/materials", "label": "GET /api/v1/materials"},
    {"module": "material_archives", "permission_type": "api", "permission_key": "api.POST./api/v1/materials", "label": "POST /api/v1/materials"},
    {"module": "material_archives", "permission_type": "api", "permission_key": "api.GET./api/v1/materials/{material_id}", "label": "GET /api/v1/materials/{material_id}"},
    {"module": "material_archives", "permission_type": "api", "permission_key": "api.PUT./api/v1/materials/{material_id}", "label": "PUT /api/v1/materials/{material_id}"},
    {"module": "material_archives", "permission_type": "api", "permission_key": "api.DELETE./api/v1/materials/{material_id}", "label": "DELETE /api/v1/materials/{material_id}"},
    {"module": "material_archives", "permission_type": "api", "permission_key": "api.PATCH./api/v1/materials/{material_id}/stop-purchase", "label": "PATCH /api/v1/materials/{material_id}/stop-purchase"},
    {"module": "material_archives", "permission_type": "api", "permission_key": "api.POST./api/v1/materials/{material_id}/transition", "label": "POST /api/v1/materials/{material_id}/transition"},
    {"module": "material_archives", "permission_type": "api", "permission_key": "api.POST./api/v1/materials/governance/preview", "label": "POST /api/v1/materials/governance/preview"},
    {"module": "material_archives", "permission_type": "api", "permission_key": "api.POST./api/v1/materials/governance/import", "label": "POST /api/v1/materials/governance/import"},
    {"module": "material_archives", "permission_type": "api", "permission_key": "api.POST./api/v1/materials/ai-add/preview", "label": "POST /api/v1/materials/ai-add/preview"},
    {"module": "material_archives", "permission_type": "api", "permission_key": "api.POST./api/v1/materials/ai-add/confirm", "label": "POST /api/v1/materials/ai-add/confirm"},
    {"module": "material_archives", "permission_type": "api", "permission_key": "api.POST./api/v1/materials/match", "label": "POST /api/v1/materials/match"},
    {"module": "attribute_management", "permission_type": "api", "permission_key": "api.GET./api/v1/attributes", "label": "GET /api/v1/attributes"},
    {"module": "attribute_management", "permission_type": "api", "permission_key": "api.POST./api/v1/attributes", "label": "POST /api/v1/attributes"},
    {"module": "attribute_management", "permission_type": "api", "permission_key": "api.PUT./api/v1/attributes/{attribute_id}", "label": "PUT /api/v1/attributes/{attribute_id}"},
    {"module": "attribute_management", "permission_type": "api", "permission_key": "api.DELETE./api/v1/attributes/{attribute_id}", "label": "DELETE /api/v1/attributes/{attribute_id}"},
    {"module": "attribute_management", "permission_type": "api", "permission_key": "api.GET./api/v1/attributes/changes", "label": "GET /api/v1/attributes/changes"},
    {"module": "attribute_management", "permission_type": "api", "permission_key": "api.GET./api/v1/attributes/{attribute_id}/changes", "label": "GET /api/v1/attributes/{attribute_id}/changes"},
    {"module": "attribute_management", "permission_type": "api", "permission_key": "api.POST./api/v1/attributes/governance/preview", "label": "POST /api/v1/attributes/governance/preview"},
    {"module": "attribute_management", "permission_type": "api", "permission_key": "api.POST./api/v1/attributes/governance/import", "label": "POST /api/v1/attributes/governance/import"},
    {"module": "attribute_management", "permission_type": "api", "permission_key": "api.POST./api/v1/ai/attribute-recommend", "label": "POST /api/v1/ai/attribute-recommend"},
    {"module": "workflow", "permission_type": "api", "permission_key": "api.GET./api/v1/workflows/applications", "label": "GET /api/v1/workflows/applications"},
    {"module": "workflow", "permission_type": "api", "permission_key": "api.POST./api/v1/workflows/applications", "label": "POST /api/v1/workflows/applications"},
    {"module": "workflow", "permission_type": "api", "permission_key": "api.GET./api/v1/workflows/applications/{application_id}", "label": "GET /api/v1/workflows/applications/{application_id}"},
    {"module": "workflow", "permission_type": "api", "permission_key": "api.POST./api/v1/workflows/applications/{application_id}/approve", "label": "POST /api/v1/workflows/applications/{application_id}/approve"},
    {"module": "workflow", "permission_type": "api", "permission_key": "api.POST./api/v1/workflows/applications/{application_id}/reject", "label": "POST /api/v1/workflows/applications/{application_id}/reject"},
    {"module": "workflow", "permission_type": "api", "permission_key": "api.GET./api/v1/workflows/tasks", "label": "GET /api/v1/workflows/tasks"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.GET./api/v1/users", "label": "GET /api/v1/users"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.POST./api/v1/users", "label": "POST /api/v1/users"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.GET./api/v1/users/{user_id}", "label": "GET /api/v1/users/{user_id}"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.PUT./api/v1/users/{user_id}", "label": "PUT /api/v1/users/{user_id}"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.DELETE./api/v1/users/{user_id}", "label": "DELETE /api/v1/users/{user_id}"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.POST./api/v1/users/{user_id}/password-reset", "label": "POST /api/v1/users/{user_id}/password-reset"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.GET./api/v1/roles", "label": "GET /api/v1/roles"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.POST./api/v1/roles", "label": "POST /api/v1/roles"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.GET./api/v1/roles/{role_id}", "label": "GET /api/v1/roles/{role_id}"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.PUT./api/v1/roles/{role_id}", "label": "PUT /api/v1/roles/{role_id}"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.DELETE./api/v1/roles/{role_id}", "label": "DELETE /api/v1/roles/{role_id}"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.PATCH./api/v1/roles/{role_id}/enable", "label": "PATCH /api/v1/roles/{role_id}/enable"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.PATCH./api/v1/roles/{role_id}/disable", "label": "PATCH /api/v1/roles/{role_id}/disable"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.GET./api/v1/roles/{role_id}/users", "label": "GET /api/v1/roles/{role_id}/users"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.POST./api/v1/roles/{role_id}/users", "label": "POST /api/v1/roles/{role_id}/users"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.PUT./api/v1/roles/{role_id}/users", "label": "PUT /api/v1/roles/{role_id}/users"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.DELETE./api/v1/roles/{role_id}/users/{user_id}", "label": "DELETE /api/v1/roles/{role_id}/users/{user_id}"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.GET./api/v1/roles/{role_id}/permissions", "label": "GET /api/v1/roles/{role_id}/permissions"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.PUT./api/v1/roles/{role_id}/permissions", "label": "PUT /api/v1/roles/{role_id}/permissions"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.GET./api/v1/permissions/catalog", "label": "GET /api/v1/permissions/catalog"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.GET./api/v1/categories", "label": "GET /api/v1/categories"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.POST./api/v1/categories", "label": "POST /api/v1/categories"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.GET./api/v1/categories/template", "label": "GET /api/v1/categories/template"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.POST./api/v1/categories/bulk-import", "label": "POST /api/v1/categories/bulk-import"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.GET./api/v1/categories/{category_id}/attributes", "label": "GET /api/v1/categories/{category_id}/attributes"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.GET./api/v1/categories/{category_id}/attributes/own", "label": "GET /api/v1/categories/{category_id}/attributes/own"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.GET./api/v1/categories/{category_id}/properties", "label": "GET /api/v1/categories/{category_id}/properties"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.POST./api/v1/categories/{category_id}/attributes", "label": "POST /api/v1/categories/{category_id}/attributes"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.POST./api/v1/categories/{category_id}/attributes/batch", "label": "POST /api/v1/categories/{category_id}/attributes/batch"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.PUT./api/v1/categories/{category_id}/attributes/{attribute_id}", "label": "PUT /api/v1/categories/{category_id}/attributes/{attribute_id}"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.DELETE./api/v1/categories/{category_id}/attributes/{attribute_id}", "label": "DELETE /api/v1/categories/{category_id}/attributes/{attribute_id}"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.POST./api/v1/ai/category-recognition/recognize", "label": "POST /api/v1/ai/category-recognition/recognize"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.POST./api/v1/ai/category-recognition/recognize-async", "label": "POST /api/v1/ai/category-recognition/recognize-async"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.GET./api/v1/ai/category-recognition/jobs/{job_id}", "label": "GET /api/v1/ai/category-recognition/jobs/{job_id}"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.POST./api/v1/ai/category-recognition/batch", "label": "POST /api/v1/ai/category-recognition/batch"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.POST./api/v1/ai/material-category-match", "label": "POST /api/v1/ai/material-category-match"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.PUT./api/v1/categories/{category_id}", "label": "PUT /api/v1/categories/{category_id}"},
    {"module": "category_management", "permission_type": "api", "permission_key": "api.DELETE./api/v1/categories/{category_id}", "label": "DELETE /api/v1/categories/{category_id}"},
    {"module": "category_library", "permission_type": "api", "permission_key": "api.GET./api/v1/category-libraries", "label": "GET /api/v1/category-libraries"},
    {"module": "category_library", "permission_type": "api", "permission_key": "api.POST./api/v1/category-libraries", "label": "POST /api/v1/category-libraries"},
    {"module": "category_library", "permission_type": "api", "permission_key": "api.GET./api/v1/category-libraries/{library_id}", "label": "GET /api/v1/category-libraries/{library_id}"},
    {"module": "category_library", "permission_type": "api", "permission_key": "api.PUT./api/v1/category-libraries/{library_id}", "label": "PUT /api/v1/category-libraries/{library_id}"},
    {"module": "category_library", "permission_type": "api", "permission_key": "api.DELETE./api/v1/category-libraries/{library_id}", "label": "DELETE /api/v1/category-libraries/{library_id}"},
    {"module": "category_library", "permission_type": "api", "permission_key": "api.POST./api/v1/category-libraries/{library_id}/re-embed", "label": "POST /api/v1/category-libraries/{library_id}/re-embed"},
    {"module": "product_name_management", "permission_type": "api", "permission_key": "api.GET./api/v1/product-names", "label": "GET /api/v1/product-names"},
    {"module": "product_name_management", "permission_type": "api", "permission_key": "api.POST./api/v1/product-names", "label": "POST /api/v1/product-names"},
    {"module": "product_name_management", "permission_type": "api", "permission_key": "api.GET./api/v1/product-names/{product_name_id}", "label": "GET /api/v1/product-names/{product_name_id}"},
    {"module": "product_name_management", "permission_type": "api", "permission_key": "api.PUT./api/v1/product-names/{product_name_id}", "label": "PUT /api/v1/product-names/{product_name_id}"},
    {"module": "product_name_management", "permission_type": "api", "permission_key": "api.PATCH./api/v1/product-names/{product_name_id}/status", "label": "PATCH /api/v1/product-names/{product_name_id}/status"},
    {"module": "product_name_management", "permission_type": "api", "permission_key": "api.DELETE./api/v1/product-names/{product_name_id}", "label": "DELETE /api/v1/product-names/{product_name_id}"},
    {"module": "product_name_management", "permission_type": "button", "permission_key": "button.product_names.create", "label": "Product Name Create"},
    {"module": "product_name_management", "permission_type": "button", "permission_key": "button.product_names.edit", "label": "Product Name Edit"},
    {"module": "product_name_management", "permission_type": "button", "permission_key": "button.product_names.delete", "label": "Product Name Delete"},
    {"module": "brand_management", "permission_type": "api", "permission_key": "api.GET./api/v1/brands", "label": "GET /api/v1/brands"},
    {"module": "brand_management", "permission_type": "api", "permission_key": "api.POST./api/v1/brands", "label": "POST /api/v1/brands"},
    {"module": "brand_management", "permission_type": "api", "permission_key": "api.PUT./api/v1/brands/{brand_id}", "label": "PUT /api/v1/brands/{brand_id}"},
    {"module": "brand_management", "permission_type": "api", "permission_key": "api.DELETE./api/v1/brands/{brand_id}", "label": "DELETE /api/v1/brands/{brand_id}"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.GET./api/v1/system/config", "label": "GET /api/v1/system/config"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.PUT./api/v1/system/config", "label": "PUT /api/v1/system/config"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.GET./api/v1/audit-logs", "label": "GET /api/v1/audit-logs"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.GET./api/v1/audit-logs/{log_id}", "label": "GET /api/v1/audit-logs/{log_id}"},
    {"module": "system_admin", "permission_type": "api", "permission_key": "api.GET./api/v1/audit-logs/export", "label": "GET /api/v1/audit-logs/export"},
]


@app.on_event("startup")
def startup() -> None:
    Base.metadata.create_all(bind=engine)
    ensure_slow_query_schema()
    ensure_web_vitals_schema()
    ensure_audit_log_schema()
    ensure_material_code_rule_schema()
    ensure_category_schema()
    ensure_category_attribute_schema()
    ensure_material_library_association_schema()
    ensure_product_name_schema()
    db = next(get_db())
    try:
        ensure_product_name_code_sequence(db)
        ensure_role_code_sequence(db)
        ensure_seed_product(db)
        ensure_seed_material_context(db)
        ensure_provider_configs(db)
        ensure_system_config(db)
        ensure_hcm_seed_users(db)
        ensure_rule_engine_seed(db)
        db.commit()
    finally:
        db.close()


def ensure_slow_query_schema() -> None:
    with engine.begin() as connection:
        SlowQueryLog.__table__.create(bind=connection, checkfirst=True)


def ensure_web_vitals_schema() -> None:
    with engine.begin() as connection:
        TelemetryWebVital.__table__.create(bind=connection, checkfirst=True)


def ensure_audit_log_schema() -> None:
    required = {"id", "user", "resource", "action", "before_value", "after_value", "timestamp", "source"}
    with engine.begin() as connection:
        if engine.dialect.name == "sqlite":
            table_exists = connection.exec_driver_sql(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='audit_log'"
            ).fetchone()
            if table_exists:
                columns = {row[1] for row in connection.exec_driver_sql("PRAGMA table_info(audit_log)").fetchall()}
                if not required.issubset(columns):
                    legacy_name = f"audit_log_legacy_{int(time.time())}"
                    connection.exec_driver_sql(f"ALTER TABLE audit_log RENAME TO {legacy_name}")
                    legacy_indexes = connection.exec_driver_sql(
                        "SELECT name FROM sqlite_master WHERE type='index' AND name LIKE 'ix_audit_log_%'"
                    ).fetchall()
                    for index in legacy_indexes:
                        connection.exec_driver_sql(f"DROP INDEX IF EXISTS {index[0]}")
        AuditLog.__table__.create(bind=connection, checkfirst=True)


def ensure_material_code_rule_schema() -> None:
    if engine.dialect.name != "sqlite":
        Base.metadata.create_all(bind=engine)
        return
    table_columns = {
        "material_libraries": {
            "auto_code_enabled": "BOOLEAN DEFAULT 0 NOT NULL",
            "recode_enabled": "BOOLEAN DEFAULT 0 NOT NULL",
            "current_rule_version_id": "INTEGER",
        },
        "materials": {
            "original_code": "VARCHAR(64) DEFAULT '' NOT NULL",
            "previous_code": "VARCHAR(64) DEFAULT '' NOT NULL",
            "code_rule_version_id": "INTEGER",
            "code_change_count": "INTEGER DEFAULT 0 NOT NULL",
            "code_status": "VARCHAR(40) DEFAULT 'manual' NOT NULL",
        },
    }
    with engine.begin() as connection:
        for table_name, columns in table_columns.items():
            existing = {row[1] for row in connection.exec_driver_sql(f"PRAGMA table_info({table_name})").fetchall()}
            for column_name, ddl in columns.items():
                if column_name not in existing:
                    connection.exec_driver_sql(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {ddl}")
        Base.metadata.create_all(bind=connection)


def ensure_category_schema() -> None:
    with engine.begin() as connection:
        if engine.dialect.name == "sqlite":
            Base.metadata.create_all(bind=connection)
            existing = {row[1] for row in connection.exec_driver_sql("PRAGMA table_info(categories)").fetchall()}
            if "category_library_id" not in existing:
                connection.exec_driver_sql("ALTER TABLE categories ADD COLUMN category_library_id INTEGER")
            if "parent_category_id" not in existing:
                connection.exec_driver_sql("ALTER TABLE categories ADD COLUMN parent_category_id INTEGER")
            for index in connection.exec_driver_sql("PRAGMA index_list(categories)").fetchall():
                index_name = index[1]
                is_unique = bool(index[2])
                columns = [row[2] for row in connection.exec_driver_sql(f"PRAGMA index_info({index_name})").fetchall()]
                if is_unique and columns == ["name"]:
                    connection.exec_driver_sql(f"DROP INDEX {index_name}")
            return

        Base.metadata.create_all(bind=connection)
        connection.exec_driver_sql("ALTER TABLE categories DROP CONSTRAINT IF EXISTS categories_name_key")
        connection.exec_driver_sql("DROP INDEX IF EXISTS ix_categories_name")
        has_column = connection.exec_driver_sql(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_name = 'categories'
              AND column_name = 'category_library_id'
            """
        ).fetchone()
        if not has_column:
            connection.exec_driver_sql("ALTER TABLE categories ADD COLUMN category_library_id INTEGER")
        has_parent_column = connection.exec_driver_sql(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_name = 'categories'
              AND column_name = 'parent_category_id'
            """
        ).fetchone()
        if not has_parent_column:
            connection.exec_driver_sql("ALTER TABLE categories ADD COLUMN parent_category_id INTEGER")


def ensure_category_attribute_schema() -> None:
    with engine.begin() as connection:
        Base.metadata.create_all(bind=connection)
        CategoryAttribute.__table__.create(bind=connection, checkfirst=True)
        if engine.dialect.name == "sqlite":
            connection.exec_driver_sql(
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_category_attribute_name "
                "ON category_attributes (category_id, name)"
            )
        else:
            connection.exec_driver_sql(
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_category_attribute_name "
                "ON category_attributes (category_id, name)"
            )


def ensure_material_library_association_schema() -> None:
    with engine.begin() as connection:
        if engine.dialect.name == "sqlite":
            Base.metadata.create_all(bind=connection)
            MaterialLibraryAdminRole.__table__.create(bind=connection, checkfirst=True)
            MaterialLibraryCategoryLibrary.__table__.create(bind=connection, checkfirst=True)
            existing = {row[1] for row in connection.exec_driver_sql("PRAGMA table_info(material_libraries)").fetchall()}
            if "material_library_admin_id" not in existing:
                connection.exec_driver_sql("ALTER TABLE material_libraries ADD COLUMN material_library_admin_id INTEGER")
            if "category_library_id" not in existing:
                connection.exec_driver_sql("ALTER TABLE material_libraries ADD COLUMN category_library_id INTEGER")
            category_columns = {row[1] for row in connection.exec_driver_sql("PRAGMA table_info(category_libraries)").fetchall()}
            if "qdrant_enabled" not in category_columns:
                connection.exec_driver_sql("ALTER TABLE category_libraries ADD COLUMN qdrant_enabled BOOLEAN DEFAULT 0 NOT NULL")
            connection.exec_driver_sql(
                "CREATE INDEX IF NOT EXISTS ix_material_libraries_material_library_admin_id "
                "ON material_libraries (material_library_admin_id)"
            )
            connection.exec_driver_sql(
                "CREATE INDEX IF NOT EXISTS ix_material_libraries_category_library_id "
                "ON material_libraries (category_library_id)"
            )
            connection.exec_driver_sql(
                """
                INSERT OR IGNORE INTO material_library_admin_roles (material_library_id, role_id, created_at)
                SELECT id, material_library_admin_id, CURRENT_TIMESTAMP
                FROM material_libraries
                WHERE material_library_admin_id IS NOT NULL
                """
            )
            connection.exec_driver_sql(
                """
                INSERT OR IGNORE INTO material_library_category_libraries (material_library_id, category_library_id, created_at)
                SELECT id, category_library_id, CURRENT_TIMESTAMP
                FROM material_libraries
                WHERE category_library_id IS NOT NULL
                """
            )
            RoleCodeSequence.__table__.create(bind=connection, checkfirst=True)
            return

        Base.metadata.create_all(bind=connection)
        MaterialLibraryAdminRole.__table__.create(bind=connection, checkfirst=True)
        MaterialLibraryCategoryLibrary.__table__.create(bind=connection, checkfirst=True)
        columns = {
            row[0]
            for row in connection.exec_driver_sql(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = 'material_libraries'
                """
            ).fetchall()
        }
        if "material_library_admin_id" not in columns:
            connection.exec_driver_sql("ALTER TABLE material_libraries ADD COLUMN material_library_admin_id INTEGER REFERENCES roles(id)")
            connection.exec_driver_sql(
                "CREATE INDEX IF NOT EXISTS ix_material_libraries_material_library_admin_id "
                "ON material_libraries (material_library_admin_id)"
            )
        if "category_library_id" not in columns:
            connection.exec_driver_sql("ALTER TABLE material_libraries ADD COLUMN category_library_id INTEGER REFERENCES category_libraries(id)")
            connection.exec_driver_sql(
                "CREATE INDEX IF NOT EXISTS ix_material_libraries_category_library_id "
                "ON material_libraries (category_library_id)"
            )
        category_columns = {
            row[0]
            for row in connection.exec_driver_sql(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = 'category_libraries'
                """
            ).fetchall()
        }
        if "qdrant_enabled" not in category_columns:
            connection.exec_driver_sql(
                "ALTER TABLE category_libraries ADD COLUMN qdrant_enabled BOOLEAN NOT NULL DEFAULT false"
            )
        connection.exec_driver_sql(
            """
            INSERT INTO material_library_admin_roles (material_library_id, role_id, created_at)
            SELECT id, material_library_admin_id, NOW()
            FROM material_libraries
            WHERE material_library_admin_id IS NOT NULL
            ON CONFLICT DO NOTHING
            """
        )
        connection.exec_driver_sql(
            """
            INSERT INTO material_library_category_libraries (material_library_id, category_library_id, created_at)
            SELECT id, category_library_id, NOW()
            FROM material_libraries
            WHERE category_library_id IS NOT NULL
            ON CONFLICT DO NOTHING
            """
        )


def ensure_product_name_schema() -> None:
    with engine.begin() as connection:
        if engine.dialect.name == "sqlite":
            Base.metadata.create_all(bind=connection)
            existing = {row[1] for row in connection.exec_driver_sql("PRAGMA table_info(product_names)").fetchall()}
            if "product_name_code" not in existing:
                connection.exec_driver_sql(
                    "ALTER TABLE product_names ADD COLUMN product_name_code VARCHAR(12) DEFAULT '' NOT NULL"
                )
            if "status" not in existing:
                connection.exec_driver_sql(
                    "ALTER TABLE product_names ADD COLUMN status VARCHAR(20) DEFAULT 'active' NOT NULL"
                )
            connection.exec_driver_sql(
                "CREATE INDEX IF NOT EXISTS ix_product_names_status ON product_names (status)"
            )
            ProductNameCodeSequence.__table__.create(bind=connection, checkfirst=True)
            return

        Base.metadata.create_all(bind=connection)
        columns = {
            row[0]
            for row in connection.exec_driver_sql(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = 'product_names'
                """
            ).fetchall()
        }
        if "product_name_code" not in columns:
            connection.exec_driver_sql("ALTER TABLE product_names ADD COLUMN product_name_code VARCHAR(12)")
        if "status" not in columns:
            connection.exec_driver_sql("ALTER TABLE product_names ADD COLUMN status VARCHAR(20) DEFAULT 'active' NOT NULL")
        connection.exec_driver_sql("CREATE INDEX IF NOT EXISTS ix_product_names_status ON product_names (status)")


def ensure_seed_product(db: Session) -> ProductName:
    Base.metadata.create_all(bind=engine)
    product = db.query(ProductName).filter(ProductName.name == SEED_PRODUCT["name"]).first()
    if product:
        if not product.product_name_code:
            product.product_name_code = generate_product_name_code(db)
        if product.status not in PRODUCT_NAME_STATUSES:
            product.status = "active"
        return product
    product = ProductName(**SEED_PRODUCT, product_name_code=generate_product_name_code(db), status="active")
    db.add(product)
    db.commit()
    db.refresh(product)
    return product


def ensure_default_category_library(db: Session) -> CategoryLibrary:
    Base.metadata.create_all(bind=engine)
    ensure_category_schema()
    category_library = db.query(CategoryLibrary).filter(CategoryLibrary.code == SEED_CATEGORY_LIBRARY["code"]).first()
    if category_library:
        return category_library
    category_library = CategoryLibrary(**SEED_CATEGORY_LIBRARY)
    db.add(category_library)
    db.flush()
    return category_library


def ensure_seed_material_context(db: Session) -> tuple[MaterialLibrary, Category]:
    Base.metadata.create_all(bind=engine)
    ensure_category_schema()
    ensure_material_library_association_schema()
    library = db.query(MaterialLibrary).filter(MaterialLibrary.code == SEED_LIBRARY["code"]).first()
    if not library:
        library = MaterialLibrary(**SEED_LIBRARY)
        db.add(library)
    category_library = ensure_default_category_library(db)
    category = db.query(Category).filter(Category.code == SEED_CATEGORY["code"]).first()
    if not category:
        category = Category(**SEED_CATEGORY, category_library_id=category_library.id)
        db.add(category)
    elif category.category_library_id is None:
        category.category_library_id = category_library.id
    db.query(Category).filter(Category.category_library_id.is_(None)).update(
        {Category.category_library_id: category_library.id},
        synchronize_session=False,
    )
    db.commit()
    db.refresh(library)
    db.refresh(category)
    return library, category


def normalize_capabilities(capabilities: list[str] | str | None) -> list[str]:
    if capabilities is None:
        return []
    if isinstance(capabilities, str):
        try:
            loaded = json.loads(capabilities)
            if isinstance(loaded, list):
                capabilities = loaded
            else:
                capabilities = capabilities.split(",")
        except json.JSONDecodeError:
            capabilities = capabilities.split(",")
    return [str(item).strip() for item in capabilities if str(item).strip()]


def encryption_key() -> bytes:
    raw = os.environ.get("LLM_GATEWAY_AES_KEY", "").strip()
    if raw:
        try:
            decoded = base64.urlsafe_b64decode(raw + "=" * (-len(raw) % 4))
            if len(decoded) == 32:
                return decoded
        except (ValueError, binascii.Error):
            pass
        return sha256(raw.encode("utf-8")).digest()
    return sha256(b"material-retrieval-local-aes-256-key").digest()


def encrypt_api_key(api_key: str | None) -> str:
    if not api_key:
        return ""
    aes = AESGCM(encryption_key())
    nonce = secrets.token_bytes(12)
    ciphertext = aes.encrypt(nonce, api_key.encode("utf-8"), None)
    return base64.urlsafe_b64encode(nonce + ciphertext).decode("ascii")


def decrypt_api_key(encrypted_api_key: str | None) -> str:
    if not encrypted_api_key:
        return ""
    try:
        payload = base64.urlsafe_b64decode(encrypted_api_key.encode("ascii"))
        nonce, ciphertext = payload[:12], payload[12:]
        return AESGCM(encryption_key()).decrypt(nonce, ciphertext, None).decode("utf-8")
    except Exception:
        return ""


def masked_api_key(encrypted_api_key: str | None) -> str:
    api_key = decrypt_api_key(encrypted_api_key)
    if not api_key:
        return ""
    if len(api_key) <= 8:
        return "********"
    return f"{api_key[:2]}{'*' * max(6, len(api_key) - 6)}{api_key[-4:]}"


def provider_display_name(provider: str, model_name: str, fallback: str | None = None) -> str:
    name = compact_space(fallback or "")
    if name:
        return name
    return f"{provider}-{model_name}"


def ai_debug_enabled() -> bool:
    return os.environ.get("AI_DEBUG", "true").strip().lower() in {"1", "true", "yes", "on"}


def ensure_provider_configs(db: Session) -> ModelConfig:
    Base.metadata.create_all(bind=engine)
    provider = db.query(ModelConfig).filter(ModelConfig.enabled.is_(True)).first()
    if provider:
        return provider
    provider = ModelConfig(
        display_name=provider_display_name(DEFAULT_PROVIDER["provider"], DEFAULT_PROVIDER["model"], "Default Mock Model"),
        provider=DEFAULT_PROVIDER["provider"],
        model_name=DEFAULT_PROVIDER["model"],
        base_url=DEFAULT_PROVIDER["endpoint"],
        timeout_seconds=5,
        enabled=True,
        connection_status="connected",
        last_test_message="Local deterministic provider is available",
        last_test_at=now(),
    )
    db.add(provider)
    db.flush()
    for capability in DEFAULT_PROVIDER["capabilities"]:
        db.add(CapabilityModelMapping(capability=capability, primary_model_id=provider.id, enabled=True))
    db.commit()
    db.refresh(provider)
    return provider


def enabled_models_for_mapping(db: Session) -> list[ModelConfig]:
    ensure_provider_configs(db)
    return db.query(ModelConfig).filter(ModelConfig.enabled.is_(True)).order_by(ModelConfig.display_name).all()


def mapping_for_capability(db: Session, capability: str) -> CapabilityModelMapping | None:
    ensure_provider_configs(db)
    return (
        db.query(CapabilityModelMapping)
        .filter(CapabilityModelMapping.capability == capability, CapabilityModelMapping.enabled.is_(True))
        .first()
    )


def default_reason_options(values: list[str]) -> list[dict[str, Any]]:
    return [{"name": value, "enabled": True} for value in values]


def default_system_config() -> dict[str, Any]:
    return {
        "system_name": DEFAULT_SYSTEM_NAME,
        "icon": DEFAULT_SYSTEM_ICON,
        "stop_purchase_reasons": default_reason_options(DEFAULT_STOP_PURCHASE_REASONS),
        "stop_use_reasons": default_reason_options(DEFAULT_STOP_USE_REASONS),
        "approval_mode": "multi_node",
    }


def safe_json_loads(value: str | None, fallback: Any) -> Any:
    if not value:
        return fallback
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return fallback


def normalize_reason_option(value: ReasonOption | str | dict[str, Any]) -> dict[str, Any]:
    if isinstance(value, ReasonOption):
        name = value.name.strip()
        enabled = value.enabled
    elif isinstance(value, dict):
        name = str(value.get("name") or value.get("label") or value.get("value") or "").strip()
        enabled = bool(value.get("enabled", True))
    else:
        name = str(value).strip()
        enabled = True
    if not name:
        raise HTTPException(status_code=422, detail="Reason option name is required")
    return {"name": name, "enabled": enabled}


def normalize_reason_options(values: list[ReasonOption | str] | list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for value in values:
        option = normalize_reason_option(value)
        if option["name"] in seen:
            continue
        normalized.append(option)
        seen.add(option["name"])
    return normalized


def sanitize_icon(icon: SystemIcon | dict[str, Any] | None) -> dict[str, str]:
    if not icon:
        return DEFAULT_SYSTEM_ICON.copy()
    if isinstance(icon, SystemIcon):
        data = icon.model_dump()
    else:
        data = dict(icon)
    return {
        "filename": str(data.get("filename") or "system-icon").strip()[:240],
        "content_type": str(data.get("content_type") or "image/png").strip()[:120],
        "data_url": str(data.get("data_url") or "").strip(),
    }


def system_config_payload(config: SystemConfig) -> dict[str, Any]:
    data = default_system_config()
    loaded = safe_json_loads(config.value, {})
    if isinstance(loaded, dict):
        data.update({key: value for key, value in loaded.items() if key in data})
    if data["approval_mode"] not in APPROVAL_MODES:
        data["approval_mode"] = "multi_node"
    data["system_name"] = compact_space(str(data.get("system_name") or DEFAULT_SYSTEM_NAME)) or DEFAULT_SYSTEM_NAME
    data["icon"] = sanitize_icon(data.get("icon"))
    data["stop_purchase_reasons"] = normalize_reason_options(data.get("stop_purchase_reasons") or default_reason_options(DEFAULT_STOP_PURCHASE_REASONS))
    data["stop_use_reasons"] = normalize_reason_options(data.get("stop_use_reasons") or default_reason_options(DEFAULT_STOP_USE_REASONS))
    return data


def ensure_system_config(db: Session) -> SystemConfig:
    Base.metadata.create_all(bind=engine)
    config = db.query(SystemConfig).filter(SystemConfig.key == SYSTEM_CONFIG_KEY).first()
    if config:
        return config
    legacy = db.query(SystemConfig).filter(SystemConfig.key == "approval_mode").first()
    data = default_system_config()
    if legacy and legacy.value in APPROVAL_MODES:
        data["approval_mode"] = legacy.value
    config = SystemConfig(key=SYSTEM_CONFIG_KEY, value=json.dumps(data, ensure_ascii=False), updated_by="system")
    db.add(config)
    db.commit()
    db.refresh(config)
    return config


def ensure_hcm_seed_users(db: Session) -> None:
    Base.metadata.create_all(bind=engine)
    changed = False
    for item in HCM_SEED_USERS:
        user = db.query(User).filter(User.username == item["username"]).first()
        if not user:
            db.add(
                User(
                    **item,
                    account_ownership="HCM",
                    status="active",
                )
            )
            changed = True
        else:
            for field in ["display_name", "hcm_id", "unit", "department", "team", "email"]:
                setattr(user, field, item[field])
            user.account_ownership = "HCM"
            user.status = user.status or "active"
            changed = True
    if changed:
        db.commit()


def json_options(value: Any) -> dict[str, Any] | list[Any]:
    if isinstance(value, dict) or isinstance(value, list):
        return value
    if not value:
        return {}
    try:
        loaded = json.loads(str(value))
    except json.JSONDecodeError:
        return {}
    return loaded if isinstance(loaded, dict) or isinstance(loaded, list) else {}


def options_as_dict(value: Any) -> dict[str, Any]:
    loaded = json_options(value)
    return loaded if isinstance(loaded, dict) else {"values": loaded}


def ensure_rule_engine_seed(db: Session) -> None:
    Base.metadata.create_all(bind=engine)
    changed = False
    categories_by_slug: dict[str, RuleCategory] = {}
    for item in DEFAULT_RULE_CATEGORIES:
        category = db.query(RuleCategory).filter(RuleCategory.slug == item["slug"]).first()
        if not category:
            category = RuleCategory(**item)
            db.add(category)
            changed = True
        else:
            for field in [
                "display_name_zh",
                "display_name_en",
                "description_zh",
                "description_en",
                "icon",
                "sort_order",
            ]:
                if getattr(category, field) != item[field]:
                    setattr(category, field, item[field])
                    changed = True
        categories_by_slug[item["slug"]] = category
    if changed:
        db.flush()
    for item in DEFAULT_RULES:
        category = categories_by_slug[item["category_slug"]]
        existing = db.query(Rule).filter(Rule.category_id == category.id, Rule.name == item["name"]).first()
        if existing:
            continue
        db.add(
            Rule(
                category_id=category.id,
                name=item["name"],
                description=item["description"],
                pattern=item["pattern"],
                value=item["value"],
                options=json.dumps(item["options"], ensure_ascii=False),
                priority=item["priority"],
                enabled=item["enabled"],
            )
        )
        changed = True
    if changed:
        db.commit()


def approval_mode(db: Session) -> str:
    value = system_config_payload(ensure_system_config(db)).get("approval_mode")
    return value if value in APPROVAL_MODES else "multi_node"


def config_to_out(config: SystemConfig) -> SystemConfigOut:
    data = system_config_payload(config)
    return SystemConfigOut(
        system_name=data["system_name"],
        icon=SystemIcon(**data["icon"]),
        stop_purchase_reasons=[ReasonOption(**item) for item in data["stop_purchase_reasons"]],
        stop_use_reasons=[ReasonOption(**item) for item in data["stop_use_reasons"]],
        approval_mode=data["approval_mode"],
        updated_by=config.updated_by,
        updated_at=config.updated_at.isoformat(),
    )


def redact_sensitive(value: Any) -> Any:
    if isinstance(value, dict):
        redacted: dict[str, Any] = {}
        for key, item in value.items():
            lowered = str(key).lower()
            if any(token in lowered for token in ["api_key", "apikey", "secret", "token", "password", "encrypted_api_key"]):
                redacted[str(key)] = "********"
            else:
                redacted[str(key)] = redact_sensitive(item)
        return redacted
    if isinstance(value, list):
        return [redact_sensitive(item) for item in value]
    return value


def audit_to_out(log: AuditLog) -> AuditLogOut:
    return AuditLogOut(
        id=log.id,
        user=log.user,
        resource=log.resource,
        action=log.action,
        before_value=redact_sensitive(safe_json_loads(log.before_value, {})),
        after_value=redact_sensitive(safe_json_loads(log.after_value, {})),
        timestamp=log.timestamp.isoformat(),
        source=log.source,
    )


def slow_query_to_out(log: SlowQueryLog) -> SlowQueryLogOut:
    return SlowQueryLogOut(
        id=log.id,
        timestamp=log.timestamp.isoformat(),
        duration_ms=float(log.duration_ms),
        operation=log.operation,
        statement=log.statement,
    )


def web_vitals_to_out(record: TelemetryWebVital) -> WebVitalsTelemetryOut:
    return WebVitalsTelemetryOut(
        id=record.id,
        metric=record.metric,
        value=float(record.value),
        rating=record.rating,
        client_metric_id=record.client_metric_id,
        navigation_type=record.navigation_type,
        url=record.url,
        path=record.path,
        user_agent=record.user_agent,
        timestamp=record.timestamp,
        created_at=record.created_at.isoformat(),
    )


def add_audit_log(
    db: Session,
    auth: AuthContext | None,
    resource: str,
    action: str,
    before_value: dict[str, Any] | None,
    after_value: dict[str, Any] | None,
    source: str = "human",
) -> None:
    ensure_audit_log_schema()
    db.add(
        AuditLog(
            user=auth.username if auth else "system",
            resource=resource,
            action=action,
            before_value=json.dumps(redact_sensitive(before_value or {}), ensure_ascii=False, sort_keys=True),
            after_value=json.dumps(redact_sensitive(after_value or {}), ensure_ascii=False, sort_keys=True),
            timestamp=now(),
            source=source,
        )
    )


def provider_for_capability(db: Session, capability: str) -> ModelConfig:
    mapping = mapping_for_capability(db, capability)
    if mapping and mapping.primary_model and mapping.primary_model.enabled:
        return mapping.primary_model
    fallback = db.query(ModelConfig).filter(ModelConfig.enabled.is_(True)).order_by(ModelConfig.id.desc()).first()
    if fallback:
        return fallback
    return ensure_provider_configs(db)


def capabilities_for_model(db: Session, provider: ModelConfig) -> list[str]:
    mappings = (
        db.query(CapabilityModelMapping)
        .filter(
            CapabilityModelMapping.enabled.is_(True),
            or_(
                CapabilityModelMapping.primary_model_id == provider.id,
                CapabilityModelMapping.fallback_model_id == provider.id,
            ),
        )
        .order_by(CapabilityModelMapping.capability)
        .all()
    )
    return [mapping.capability for mapping in mappings]


def provider_to_out(provider: ModelConfig, db: Session) -> ProviderConfigOut:
    capabilities = capabilities_for_model(db, provider)
    return ProviderConfigOut(
        id=provider.id,
        display_name=provider.display_name,
        provider=provider.provider,
        model=provider.model_name,
        model_name=provider.model_name,
        endpoint=provider.base_url,
        base_url=provider.base_url,
        api_key_masked=masked_api_key(provider.encrypted_api_key),
        capabilities=capabilities,
        active=provider.enabled,
        enabled=provider.enabled,
        timeout_seconds=provider.timeout_seconds,
        fallback_model_id=provider.fallback_model_id,
        connection_status=provider.connection_status,
        last_test_message=provider.last_test_message,
        last_test_at=provider.last_test_at.isoformat() if provider.last_test_at else None,
        updated_at=provider.updated_at.isoformat(),
    )


def agent_config_to_out(config: AIAgentConfig) -> AgentConfigOut:
    return AgentConfigOut(
        id=config.id,
        config_key=config.config_key,
        provider=config.provider,
        model_name=config.model_name,
        base_url=config.base_url,
        api_key_masked=masked_api_key(config.encrypted_api_key),
        has_api_key=bool(config.encrypted_api_key),
        temperature=config.temperature,
        max_tokens=config.max_tokens,
        timeout=config.timeout,
        enabled=config.enabled,
        connection_status=config.connection_status,
        last_test_message=config.last_test_message,
        last_test_at=config.last_test_at.isoformat() if config.last_test_at else None,
        created_at=config.created_at.isoformat(),
        updated_at=config.updated_at.isoformat(),
    )


def validate_agent_payload(payload: AgentConfigIn) -> dict[str, Any]:
    values = {
        "config_key": compact_space(payload.config_key),
        "provider": compact_space(payload.provider).lower(),
        "model_name": compact_space(payload.model_name),
        "base_url": compact_space(payload.base_url),
        "temperature": float(payload.temperature),
        "max_tokens": int(payload.max_tokens),
        "timeout": int(payload.timeout),
        "enabled": bool(payload.enabled),
    }
    if not values["config_key"]:
        raise HTTPException(status_code=422, detail="config_key is required")
    if not values["provider"]:
        raise HTTPException(status_code=422, detail="provider is required")
    if not values["model_name"]:
        raise HTTPException(status_code=422, detail="model_name is required")
    if not values["base_url"].startswith(("http://", "https://", "local://")):
        raise HTTPException(status_code=422, detail="base_url must be a valid URL")
    return values


def apply_agent_payload(config: AIAgentConfig, payload: AgentConfigIn) -> AIAgentConfig:
    values = validate_agent_payload(payload)
    for field, value in values.items():
        setattr(config, field, value)
    if payload.api_key and not payload.api_key.startswith("**"):
        config.encrypted_api_key = encrypt_api_key(payload.api_key)
    config.updated_at = now()
    return config


def test_agent_config_connection(config: AIAgentConfig) -> dict[str, Any]:
    result = test_model_connection(config)
    config.connection_status = "ok" if result["ok"] else "error"
    config.last_test_message = result["message"]
    config.last_test_at = now()
    config.updated_at = now()
    return {**result, "status": config.connection_status}


def mapping_to_out(mapping: CapabilityModelMapping) -> CapabilityMappingOut:
    return CapabilityMappingOut(
        id=mapping.id,
        capability=mapping.capability,
        primary_model_id=mapping.primary_model_id,
        primary_model_name=mapping.primary_model.display_name if mapping.primary_model else "",
        fallback_model_id=mapping.fallback_model_id,
        fallback_model_name=mapping.fallback_model.display_name if mapping.fallback_model else "",
        enabled=mapping.enabled,
        updated_at=mapping.updated_at.isoformat(),
    )


def agent_mapping_to_out(mapping: CapabilityAgentMapping) -> CapabilityMappingOut:
    return CapabilityMappingOut(
        id=mapping.id,
        capability=mapping.capability,
        primary_model_id=mapping.agent_config_id,
        primary_model_name=mapping.agent_config.config_key if mapping.agent_config else "",
        fallback_model_id=mapping.fallback_agent_config_id,
        fallback_model_name=mapping.fallback_agent_config.config_key if mapping.fallback_agent_config else "",
        agent_config_id=mapping.agent_config_id,
        agent_config_key=mapping.agent_config.config_key if mapping.agent_config else "",
        fallback_agent_config_id=mapping.fallback_agent_config_id,
        fallback_agent_config_key=mapping.fallback_agent_config.config_key if mapping.fallback_agent_config else "",
        enabled=mapping.enabled,
        updated_at=mapping.updated_at.isoformat(),
    )


def gateway_connection_status(status: str | None) -> str:
    normalized = (status or "").strip().lower()
    if normalized in {"ok", "connected", "configured"}:
        return "ok"
    if normalized in {"error", "failed", "failure"}:
        return "error"
    return "untested"


def ensure_model_gateway_schema(db: Session) -> None:
    Base.metadata.create_all(bind=engine)
    # Migration sprint55_run_once already ran during sprint 55 deployment.
    # Re-running on every request would overwrite manual capability-mapping
    # updates (the migration syncs from legacy tables that still reference
    # mock models). Run `python -m app.migrations.sprint55_migrate_ai_config`
    # manually if you need to re-migrate.
    seed_default_capability_mappings(db)


def gateway_model_by_provider_name(db: Session, provider: str, model_name: str) -> Model | None:
    return (
        db.query(Model)
        .filter(Model.provider == provider.strip().lower(), Model.model_name == compact_space(model_name))
        .first()
    )


def create_model_from_legacy_provider(db: Session, provider: ModelConfig) -> Model:
    model = gateway_model_by_provider_name(db, provider.provider, provider.model_name)
    if model:
        model.display_name = provider.display_name
        model.base_url = provider.base_url
        model.api_key_encrypted = provider.encrypted_api_key
        model.timeout = provider.timeout_seconds
        model.enabled = provider.enabled
        model.connection_status = gateway_connection_status(provider.connection_status)
        model.last_tested_at = provider.last_test_at
        model.updated_at = now()
        db.flush()
        return model
    model = Model(
        display_name=provider.display_name,
        provider=provider.provider.strip().lower(),
        model_name=provider.model_name,
        base_url=provider.base_url,
        api_key_encrypted=provider.encrypted_api_key,
        timeout=provider.timeout_seconds,
        temperature=None,
        max_tokens=None,
        enabled=provider.enabled,
        connection_status=gateway_connection_status(provider.connection_status),
        last_tested_at=provider.last_test_at,
        migration_data_version="migrated",
        created_at=provider.created_at,
        updated_at=provider.updated_at,
    )
    db.add(model)
    db.flush()
    return model


def create_model_from_legacy_agent(db: Session, agent: AIAgentConfig) -> Model:
    model = gateway_model_by_provider_name(db, agent.provider, agent.model_name)
    if model:
        model.display_name = agent.config_key or provider_display_name(agent.provider, agent.model_name)
        model.base_url = agent.base_url
        model.temperature = agent.temperature
        model.max_tokens = agent.max_tokens
        model.api_key_encrypted = agent.encrypted_api_key
        model.timeout = agent.timeout
        model.enabled = agent.enabled
        model.connection_status = gateway_connection_status(agent.connection_status)
        model.last_tested_at = agent.last_test_at
        model.updated_at = now()
        db.flush()
        return model
    model = Model(
        display_name=agent.config_key or provider_display_name(agent.provider, agent.model_name),
        provider=agent.provider.strip().lower(),
        model_name=agent.model_name,
        base_url=agent.base_url,
        api_key_encrypted=agent.encrypted_api_key,
        timeout=agent.timeout,
        temperature=agent.temperature,
        max_tokens=agent.max_tokens,
        enabled=agent.enabled,
        connection_status=gateway_connection_status(agent.connection_status),
        last_tested_at=agent.last_test_at,
        migration_data_version="migrated",
        created_at=agent.created_at,
        updated_at=agent.updated_at,
    )
    db.add(model)
    db.flush()
    return model


def migrate_legacy_model_gateway_data(db: Session) -> None:
    migration_state = db.query(SystemConfig).filter(SystemConfig.key == "model_gateway_migration_data_version").first()
    if migration_state and migration_state.value == "migrated":
        return
    for legacy_provider in db.query(ModelConfig).order_by(ModelConfig.id).all():
        create_model_from_legacy_provider(db, legacy_provider)
    agent_models: dict[int, Model] = {}
    for legacy_agent in db.query(AIAgentConfig).order_by(AIAgentConfig.id).all():
        agent_models[legacy_agent.id] = create_model_from_legacy_agent(db, legacy_agent)
    provider_models: dict[int, Model] = {}
    for legacy_provider in db.query(ModelConfig).order_by(ModelConfig.id).all():
        provider_models[legacy_provider.id] = gateway_model_by_provider_name(db, legacy_provider.provider, legacy_provider.model_name)

    for legacy_mapping in db.query(CapabilityModelMapping).order_by(CapabilityModelMapping.id).all():
        primary = provider_models.get(legacy_mapping.primary_model_id)
        fallback = provider_models.get(legacy_mapping.fallback_model_id) if legacy_mapping.fallback_model_id else None
        mapping = db.query(CapabilityMapping).filter(CapabilityMapping.capability == legacy_mapping.capability).first()
        if not mapping:
            mapping = CapabilityMapping(capability=legacy_mapping.capability)
            db.add(mapping)
        mapping.primary_model_id = primary.id if primary else None
        mapping.fallback_model_id = fallback.id if fallback else None
        mapping.enabled = legacy_mapping.enabled
        mapping.migration_data_version = "migrated"
        mapping.updated_at = now()

    for legacy_mapping in db.query(CapabilityAgentMapping).order_by(CapabilityAgentMapping.id).all():
        primary = agent_models.get(legacy_mapping.agent_config_id)
        fallback = agent_models.get(legacy_mapping.fallback_agent_config_id) if legacy_mapping.fallback_agent_config_id else None
        mapping = db.query(CapabilityMapping).filter(CapabilityMapping.capability == legacy_mapping.capability).first()
        if not mapping:
            mapping = CapabilityMapping(capability=legacy_mapping.capability)
            db.add(mapping)
        mapping.primary_model_id = primary.id if primary else None
        mapping.fallback_model_id = fallback.id if fallback else None
        mapping.enabled = legacy_mapping.enabled
        mapping.migration_data_version = "migrated"
        mapping.updated_at = now()
    if not migration_state:
        migration_state = SystemConfig(key="model_gateway_migration_data_version", value="migrated", updated_by="system")
        db.add(migration_state)
    else:
        migration_state.value = "migrated"
        migration_state.updated_by = "system"
        migration_state.updated_at = now()
    db.commit()


def seed_default_capability_mappings(db: Session) -> None:
    changed = False
    for capability in sorted(DEFAULT_CAPABILITY_MAPPINGS):
        mapping = db.query(CapabilityMapping).filter(CapabilityMapping.capability == capability).first()
        if mapping:
            continue
        db.add(CapabilityMapping(capability=capability, primary_model_id=None, fallback_model_id=None, enabled=True))
        changed = True
    if changed:
        db.commit()


def sync_gateway_mapping_from_legacy_models(
    db: Session,
    capability: str,
    primary: ModelConfig | None,
    fallback: ModelConfig | None,
    enabled: bool,
) -> None:
    if not primary:
        return
    gateway_primary = create_model_from_legacy_provider(db, primary)
    gateway_fallback = create_model_from_legacy_provider(db, fallback) if fallback else None
    mapping = db.query(CapabilityMapping).filter(CapabilityMapping.capability == capability).first()
    if not mapping:
        mapping = CapabilityMapping(capability=capability)
        db.add(mapping)
    mapping.primary_model_id = gateway_primary.id
    mapping.fallback_model_id = gateway_fallback.id if gateway_fallback else None
    mapping.enabled = enabled
    mapping.migration_data_version = "migrated"
    mapping.updated_at = now()
    db.flush()


def sync_gateway_mapping_from_legacy_agents(
    db: Session,
    capability: str,
    primary: AIAgentConfig | None,
    fallback: AIAgentConfig | None,
    enabled: bool,
) -> None:
    if not primary:
        return
    gateway_primary = create_model_from_legacy_agent(db, primary)
    gateway_fallback = create_model_from_legacy_agent(db, fallback) if fallback else None
    mapping = db.query(CapabilityMapping).filter(CapabilityMapping.capability == capability).first()
    if not mapping:
        mapping = CapabilityMapping(capability=capability)
        db.add(mapping)
    mapping.primary_model_id = gateway_primary.id
    mapping.fallback_model_id = gateway_fallback.id if gateway_fallback else None
    mapping.enabled = enabled
    mapping.migration_data_version = "migrated"
    mapping.updated_at = now()
    db.flush()


def model_to_read(model: Model) -> ModelRead:
    test_marker = " ".join([model.provider or "", model.model_name or "", model.base_url or "", model.migration_data_version or ""]).lower()
    return ModelRead(
        id=model.id,
        display_name=model.display_name,
        provider=model.provider,
        model_name=model.model_name,
        base_url=model.base_url,
        timeout=model.timeout,
        temperature=model.temperature,
        max_tokens=model.max_tokens,
        enabled=model.enabled,
        connection_status=model.connection_status,
        last_tested_at=model.last_tested_at.isoformat() if model.last_tested_at else None,
        is_test=model.provider == "mock" or "mock" in test_marker or "test" in test_marker or (model.base_url or "").startswith("local://"),
        created_at=model.created_at.isoformat(),
        updated_at=model.updated_at.isoformat(),
    )


@dataclass
class ModelResolution:
    capability: str
    model: Model | ModelConfig | AIAgentConfig
    source: str
    warning: str = ""
    primary_connection_error: str = ""
    primary_model_id: int | None = None
    primary_model_name: str = ""


class CapabilityResolutionError(Exception):
    def __init__(self, capability: str, message: str, status_code: int = 409):
        super().__init__(message)
        self.capability = capability
        self.status_code = status_code
        self.suggestion = "Configure an enabled model for this capability in the Model Gateway."


def model_display_name(model: Model | ModelConfig | AIAgentConfig) -> str:
    return compact_space(getattr(model, "display_name", "") or getattr(model, "config_key", "") or getattr(model, "model_name", ""))


def resolved_model_payload(model: Model | ModelConfig | AIAgentConfig) -> dict[str, Any]:
    return {
        "id": model.id,
        "display_name": model_display_name(model),
        "provider": model.provider,
        "model_name": model.model_name,
        "base_url": model.base_url,
        "timeout": int(getattr(model, "timeout", getattr(model, "timeout_seconds", 30)) or 30),
        "temperature": getattr(model, "temperature", None),
        "max_tokens": getattr(model, "max_tokens", None),
        "enabled": bool(model.enabled),
        "connection_status": gateway_connection_status(getattr(model, "connection_status", "")),
        "last_tested_at": (
            getattr(model, "last_tested_at", None) or getattr(model, "last_test_at", None)
        ).isoformat()
        if (getattr(model, "last_tested_at", None) or getattr(model, "last_test_at", None))
        else None,
    }


def resolution_trace_metadata(resolution: ModelResolution) -> dict[str, Any]:
    return {
        "model_id": resolution.model.id,
        "model_name": resolution.model.model_name,
        "provider": resolution.model.provider,
        "resolution_source": resolution.source,
        "warning": resolution.warning,
        "primary_connection_error": resolution.primary_connection_error,
        "primary_model_id": resolution.primary_model_id,
        "primary_model_name": resolution.primary_model_name,
    }


def resolution_response_payload(resolution: ModelResolution) -> dict[str, Any]:
    payload = {
        "capability": resolution.capability,
        "source": resolution.source,
        "model": resolved_model_payload(resolution.model),
        "warning": resolution.warning,
        "primary_connection_error": resolution.primary_connection_error,
    }
    if resolution.warning:
        payload["metadata"] = {"warning": resolution.warning}
    return payload


def timed_resolution_response_payload(db: Session, capability: str, prefer_fallback: bool = False) -> dict[str, Any]:
    started = time.perf_counter()
    resolution = model_for_capability(db, capability, prefer_fallback)
    payload = resolution_response_payload(resolution)
    payload["lookup_ms"] = max(0, (time.perf_counter() - started) * 1000)
    return payload


def model_connection_probe_url(model: Model | ModelConfig | AIAgentConfig) -> str:
    base_url = (model.base_url or "").rstrip("/")
    if not base_url or base_url.startswith("local://"):
        return ""
    if base_url.endswith("/v1/models"):
        return base_url
    if base_url.endswith("/v1"):
        return f"{base_url}/models"
    return f"{base_url}/v1/models"


def test_gateway_model_connection(model: Model | ModelConfig | AIAgentConfig) -> dict[str, Any]:
    if model.provider == "mock" or (model.base_url or "").startswith("local://"):
        return {"ok": True, "status": "ok", "message": f"Local provider is available for {model.model_name}"}
    url = model_connection_probe_url(model)
    if not url:
        return {"ok": False, "status": "error", "message": "Model base URL is not configured"}
    headers = {"Content-Type": "application/json"}
    api_key = decrypt_api_key(getattr(model, "api_key_encrypted", getattr(model, "encrypted_api_key", "")))
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    get_error = ""
    try:
        response = httpx.get(url, headers=headers, timeout=max(1, int(getattr(model, "timeout", getattr(model, "timeout_seconds", 30)) or 30)))
    except (httpx.TimeoutException, httpx.ConnectError, httpx.RequestError) as exc:
        get_error = f"Connection validation failed: {exc}"
    else:
        if response.status_code < 400:
            return {"ok": True, "status": "ok", "message": f"Connection validation succeeded for {model.model_name}"}
        get_error = f"Connection validation returned HTTP {response.status_code}"
    try:
        result = call_model_config(
            model_config=model,
            prompt="connection test",
            messages=[{"role": "user", "content": "connection test"}],
            capability="connection_test",
        )
        return {"ok": True, "status": "ok", "message": f"Chat validation succeeded for {result['model']}"}
    except Exception as exc:
        return {"ok": False, "status": "error", "message": f"{get_error}; chat validation failed: {exc}"}


def validate_unified_model_connection(db: Session, model: Model) -> str:
    if gateway_connection_status(model.connection_status) == "ok":
        return ""
    result = test_gateway_model_connection(model)
    model.connection_status = "ok" if result["ok"] else "error"
    model.last_tested_at = now()
    model.updated_at = now()
    db.flush()
    if not result["ok"]:
        return str(result["message"])
    return ""


def legacy_resolution_for_capability(db: Session, capability: str, prefer_fallback: bool = False) -> ModelResolution | None:
    warning = "Legacy model configuration is deprecated; configure this capability in the Model Gateway."
    agent_mapping = agent_mapping_for_capability(db, capability)
    if agent_mapping:
        candidates = [agent_mapping.fallback_agent_config, agent_mapping.agent_config] if prefer_fallback else [agent_mapping.agent_config, agent_mapping.fallback_agent_config]
        for candidate in candidates:
            if candidate and candidate.enabled:
                return ModelResolution(capability=capability, model=candidate, source="legacy", warning=warning)
    legacy_mapping = mapping_for_capability(db, capability)
    if legacy_mapping:
        candidates = [legacy_mapping.fallback_model, legacy_mapping.primary_model] if prefer_fallback else [legacy_mapping.primary_model, legacy_mapping.fallback_model]
        for candidate in candidates:
            if candidate and candidate.enabled:
                return ModelResolution(capability=capability, model=candidate, source="legacy", warning=warning)
    return None


def model_for_capability(db: Session, capability: str, prefer_fallback: bool = False) -> ModelResolution:
    capability = compact_space(capability)
    if not capability:
        raise CapabilityResolutionError(capability, "Capability is required", 422)
    ensure_model_gateway_schema(db)

    # Query with eager loading to avoid lazy-load issues in concurrent requests
    mapping = (
        db.query(CapabilityMapping)
        .filter(CapabilityMapping.capability == capability)
        .first()
    )

    configured_unified = bool(mapping and mapping.enabled and (mapping.primary_model_id or mapping.fallback_model_id))
    if configured_unified:
        ordered: list[tuple[str, Model | None]] = []
        if prefer_fallback:
            ordered = [("fallback", mapping.fallback_model), ("primary", mapping.primary_model)]
        else:
            ordered = [("primary", mapping.primary_model), ("fallback", mapping.fallback_model)]
        primary_error = ""
        primary_model_id: int | None = None
        primary_model_name = ""
        for source, candidate in ordered:
            if not candidate:
                continue
            if not candidate.enabled:
                if source == "primary":
                    primary_error = f"Primary model {candidate.model_name} is disabled"
                    primary_model_id = candidate.id
                    primary_model_name = candidate.model_name
                continue
            connection_error = validate_unified_model_connection(db, candidate)
            if connection_error:
                if source == "primary":
                    primary_error = connection_error
                    primary_model_id = candidate.id
                    primary_model_name = candidate.model_name
                continue
            db.commit()
            return ModelResolution(
                capability=capability,
                model=candidate,
                source=source,
                primary_connection_error=primary_error if source == "fallback" else "",
                warning=primary_error if source == "fallback" and primary_error else "",
                primary_model_id=primary_model_id if source == "fallback" else None,
                primary_model_name=primary_model_name if source == "fallback" else "",
            )
        db.commit()
        raise CapabilityResolutionError(capability, f"No usable model is configured for capability {capability}")

    raise CapabilityResolutionError(capability, f"No usable model is configured for capability {capability}")


def model_snapshot(model: Model) -> dict[str, Any]:
    return model_to_read(model).model_dump()


def validate_model_payload_values(values: dict[str, Any]) -> dict[str, Any]:
    if "display_name" in values:
        values["display_name"] = compact_space(values["display_name"])
        if not values["display_name"]:
            raise HTTPException(status_code=422, detail="display_name is required")
    if "provider" in values:
        values["provider"] = compact_space(values["provider"]).lower()
        if values["provider"] not in MODEL_PROVIDERS:
            raise HTTPException(status_code=422, detail="provider must be a supported provider")
    if "model_name" in values:
        values["model_name"] = compact_space(values["model_name"])
        if not values["model_name"]:
            raise HTTPException(status_code=422, detail="model_name is required")
    if "base_url" in values:
        values["base_url"] = compact_space(values["base_url"])
        if values["base_url"] and not values["base_url"].startswith(("http://", "https://", "local://")):
            raise HTTPException(status_code=422, detail="base_url must be a valid URL")
    return values


def apply_gateway_model_values(model: Model, values: dict[str, Any], api_key: str | None = None) -> None:
    values = validate_model_payload_values(values)
    for field, value in values.items():
        setattr(model, field, value)
    if api_key and not api_key.startswith("**"):
        model.api_key_encrypted = encrypt_api_key(api_key)
    model.updated_at = now()


def ensure_distinct_mapping_models(primary_model_id: int | None, fallback_model_id: int | None) -> None:
    if primary_model_id is not None and fallback_model_id is not None and primary_model_id == fallback_model_id:
        raise HTTPException(status_code=422, detail="primary_model_id and fallback_model_id cannot be the same model")


def ensure_gateway_model_reference(db: Session, model_id: int | None, label: str) -> Model | None:
    if model_id is None:
        return None
    model = db.get(Model, model_id)
    if not model:
        raise HTTPException(status_code=404, detail=f"{label} model not found")
    return model


def capability_mapping_to_read(mapping: CapabilityMapping) -> CapabilityMappingRead:
    version = mapping.migration_data_version or ""
    migration_source = version.split(":", 1)[1] if ":" in version else version
    return CapabilityMappingRead(
        id=mapping.id,
        capability=mapping.capability,
        primary_model_id=mapping.primary_model_id,
        fallback_model_id=mapping.fallback_model_id,
        enabled=mapping.enabled,
        migration_source=migration_source,
        created_at=mapping.created_at.isoformat(),
        updated_at=mapping.updated_at.isoformat(),
    )


def capability_mapping_snapshot(mapping: CapabilityMapping) -> dict[str, Any]:
    return capability_mapping_to_read(mapping).model_dump()


def agent_mapping_for_capability(db: Session, capability: str) -> CapabilityAgentMapping | None:
    Base.metadata.create_all(bind=engine)
    return (
        db.query(CapabilityAgentMapping)
        .filter(CapabilityAgentMapping.capability == capability, CapabilityAgentMapping.enabled.is_(True))
        .first()
    )


def agent_for_capability(db: Session, capability: str) -> AIAgentConfig | None:
    mapping = agent_mapping_for_capability(db, capability)
    if mapping and mapping.agent_config and mapping.agent_config.enabled:
        return mapping.agent_config
    return None


class SpanCollector:
    def __init__(self, operation_name: str, capability: str = ""):
        self.trace_id = f"trace-{uuid.uuid4().hex[:20]}"
        self.spans: list[dict[str, Any]] = []
        self.root_span_id = self.start_span(operation_name, "chain", capability=capability)

    def start_span(
        self,
        operation_name: str,
        span_type: str,
        *,
        capability: str = "",
        parent_span_id: str | None = None,
        provider: str = "",
        model: str = "",
        metadata: dict[str, Any] | None = None,
    ) -> str:
        span_id = f"span-{uuid.uuid4().hex[:20]}"
        self.spans.append(
            {
                "trace_id": self.trace_id,
                "span_id": span_id,
                "parent_span_id": parent_span_id if parent_span_id is not None else "",
                "operation_name": operation_name,
                "span_type": span_type,
                "capability": capability,
                "provider": provider,
                "model": model,
                "status": "running",
                "start_time": now(),
                "end_time": None,
                "duration_ms": 0,
                "metadata": metadata or {},
                "error": "",
            }
        )
        return span_id

    def finish_span(self, span_id: str, status: str = "ok", error: str = "", metadata: dict[str, Any] | None = None) -> None:
        ended = now()
        for span in self.spans:
            if span["span_id"] != span_id:
                continue
            span["end_time"] = ended
            span["status"] = status
            span["error"] = error
            if metadata:
                span["metadata"].update(metadata)
            span["duration_ms"] = int((ended - span["start_time"]).total_seconds() * 1000)
            return

    def flush(self, db: Session) -> None:
        for span in self.spans:
            if span["status"] == "running":
                self.finish_span(span["span_id"])
            db.add(
                TracerSpan(
                    trace_id=span["trace_id"],
                    span_id=span["span_id"],
                    parent_span_id=span["parent_span_id"],
                    operation_name=span["operation_name"],
                    span_type=span["span_type"],
                    capability=span["capability"],
                    provider=span["provider"],
                    model=span["model"],
                    status=span["status"],
                    start_time=span["start_time"],
                    end_time=span["end_time"],
                    duration_ms=span["duration_ms"],
                    metadata_json=json.dumps(span["metadata"], ensure_ascii=False),
                    error=span["error"],
                )
            )
        db.commit()


def trace(operation_name: str, span_type: str = "chain") -> Callable:
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args: Any, **kwargs: Any) -> Any:
            collector: SpanCollector | None = kwargs.get("collector")
            if collector is None:
                return func(*args, **kwargs)
            model = kwargs.get("model_config")
            span_id = collector.start_span(
                operation_name,
                span_type,
                capability=kwargs.get("capability", ""),
                parent_span_id=kwargs.get("parent_span_id") or collector.root_span_id,
                provider=getattr(model, "provider", ""),
                model=getattr(model, "model_name", ""),
                metadata=kwargs.get("metadata") or {},
            )
            try:
                result = func(*args, **kwargs)
                collector.finish_span(span_id, "ok")
                return result
            except Exception as exc:
                collector.finish_span(span_id, "error", str(exc))
                raise

        return wrapper

    return decorator


def model_chat_url(model_config: ModelConfig) -> str:
    base_url = (model_config.base_url or "").rstrip("/")
    if not base_url or base_url.startswith("local://"):
        return ""
    if base_url.endswith("/v1/chat/completions"):
        return base_url
    if base_url.endswith("/v1"):
        return f"{base_url}/chat/completions"
    return f"{base_url}/v1/chat/completions"


def local_model_completion(model_config: ModelConfig, prompt: str, capability: str) -> dict[str, Any]:
    return {
        "content": f"{model_config.model_name} handled {capability}: {prompt}",
        "provider": model_config.provider,
        "model": model_config.model_name,
        "raw": {"local": True},
    }


@trace("llm.provider.chat", "llm")
def call_model_config(
    *,
    model_config: ModelConfig,
    prompt: str,
    messages: list[dict[str, Any]],
    capability: str,
    collector: SpanCollector | None = None,
    parent_span_id: str | None = None,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if model_config.provider == "mock" or (model_config.base_url or "").startswith("local://"):
        return local_model_completion(model_config, prompt, capability)
    url = model_chat_url(model_config)
    if not url:
        raise RuntimeError("Model base URL is not configured")
    request_messages = messages or [{"role": "user", "content": prompt}]
    headers = {"Content-Type": "application/json"}
    api_key = decrypt_api_key(model_config.encrypted_api_key)
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    started = time.perf_counter()
    body: dict[str, Any] = {
        "model": model_config.model_name,
        "messages": request_messages,
        "temperature": getattr(model_config, "temperature", 0),
    }
    max_tokens = getattr(model_config, "max_tokens", None)
    if max_tokens:
        body["max_tokens"] = int(max_tokens)
    response = httpx.post(
        url,
        json=body,
        headers=headers,
        timeout=max(1, int(getattr(model_config, "timeout_seconds", 30))),
    )
    elapsed_ms = int((time.perf_counter() - started) * 1000)
    if response.status_code >= 500:
        raise RuntimeError(f"Provider returned HTTP {response.status_code}")
    if response.status_code >= 400:
        raise RuntimeError(f"Provider rejected request with HTTP {response.status_code}")
    body = response.json()
    content = ""
    choices = body.get("choices") if isinstance(body, dict) else None
    if choices and isinstance(choices, list):
        message = choices[0].get("message", {}) if isinstance(choices[0], dict) else {}
        content = str(message.get("content") or choices[0].get("text") or "")
    return {
        "content": content,
        "provider": model_config.provider,
        "model": model_config.model_name,
        "latency_ms": elapsed_ms,
        "raw": body,
    }


def test_model_connection(model_config: ModelConfig) -> dict[str, Any]:
    try:
        result = call_model_config(
            model_config=model_config,
            prompt="connection test",
            messages=[{"role": "user", "content": "connection test"}],
            capability="connection_test",
        )
        return {
            "ok": True,
            "status": "connected",
            "message": f"Connection test succeeded for {result['model']}",
        }
    except (httpx.TimeoutException, httpx.ConnectError) as exc:
        return {"ok": False, "status": "failed", "message": f"Connection failed or timed out: {exc}"}
    except Exception as exc:
        return {"ok": False, "status": "failed", "message": str(exc)}


def invoke_gateway_capability(
    db: Session,
    capability: str,
    prompt: str,
    messages: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    collector = SpanCollector(f"gateway.{capability}", capability)
    attempted: list[dict[str, Any]] = []
    try:
        resolution = model_for_capability(db, capability)
        primary = resolution.model
        mark_root_trace_model(collector, primary, primary.model_name, resolution_trace_metadata(resolution))
        if resolution.source == "fallback" and resolution.primary_connection_error:
            attempted.append(
                {
                    "model_id": resolution.primary_model_id,
                    "model": resolution.primary_model_name,
                    "status": "error",
                    "error": resolution.primary_connection_error,
                }
            )
        try:
            result = call_model_config(
                model_config=primary,
                prompt=prompt,
                messages=messages or [],
                capability=capability,
                collector=collector,
                metadata={"role": resolution.source, **resolution_trace_metadata(resolution)},
            )
            attempted.append({"model_id": primary.id, "model": primary.model_name, "status": "ok"})
            collector.finish_span(
                collector.root_span_id,
                "ok",
                metadata={"fallback_used": resolution.source == "fallback", **resolution_trace_metadata(resolution)},
            )
            return {
                "capability": capability,
                "trace_id": collector.trace_id,
                "provider": primary.provider,
                "model": primary.model_name,
                "model_id": primary.id,
                "resolution_source": "capability_mapping",
                "content": result["content"],
                "raw": result.get("raw", {}),
                "fallback_used": resolution.source == "fallback",
                "fallback_reason": resolution.primary_connection_error,
                "attempted_models": attempted,
            }
        except Exception as primary_error:
            attempted.append({"model_id": primary.id, "model": primary.model_name, "status": "error", "error": str(primary_error)})
            if isinstance(primary, Model):
                primary.connection_status = "error"
                primary.last_tested_at = now()
                primary.updated_at = now()
                db.commit()
            try:
                fallback_resolution = model_for_capability(db, capability, prefer_fallback=True)
            except CapabilityResolutionError:
                collector.finish_span(collector.root_span_id, "error", str(primary_error), {"fallback_used": False})
                raise
            fallback = fallback_resolution.model
            if fallback.id == primary.id and fallback.__class__ is primary.__class__:
                collector.finish_span(collector.root_span_id, "error", str(primary_error), {"fallback_used": False})
                raise
            fallback_span = collector.start_span(
                "gateway.fallback_decision",
                "chain",
                capability=capability,
                parent_span_id=collector.root_span_id,
                metadata={"primary_model": primary.model_name, "reason": str(primary_error), "fallback_model": fallback.model_name},
            )
            collector.finish_span(fallback_span, "ok")
            result = call_model_config(
                model_config=fallback,
                prompt=prompt,
                messages=messages or [],
                capability=capability,
                collector=collector,
                metadata={"role": "fallback", **resolution_trace_metadata(fallback_resolution)},
            )
            attempted.append({"model_id": fallback.id, "model": fallback.model_name, "status": "ok"})
            collector.finish_span(
                collector.root_span_id,
                "ok",
                metadata={"fallback_used": True, "fallback_reason": str(primary_error), **resolution_trace_metadata(fallback_resolution)},
            )
            return {
                "capability": capability,
                "trace_id": collector.trace_id,
                "provider": fallback.provider,
                "model": fallback.model_name,
                "model_id": fallback.id,
                "resolution_source": fallback_resolution.source,
                "content": result["content"],
                "raw": result.get("raw", {}),
                "fallback_used": True,
                "fallback_reason": str(primary_error),
                "attempted_models": attempted,
            }
    except CapabilityResolutionError as exc:
        collector.finish_span(collector.root_span_id, "error", str(exc), {"capability": exc.capability})
        raise
    finally:
        collector.flush(db)


def model_values_from_payload(payload: ProviderConfigIn) -> dict[str, Any]:
    provider = compact_space(payload.provider or "mock")
    model_name = compact_space(payload.model_name or payload.model or "mock-material-governance-v1")
    display_name = provider_display_name(provider, model_name, payload.display_name or payload.name)
    base_url = compact_space(payload.base_url if payload.base_url is not None else payload.endpoint or "")
    enabled = payload.enabled if payload.enabled is not None else payload.active
    if enabled is None:
        enabled = True
    return {
        "display_name": display_name,
        "provider": provider,
        "model_name": model_name,
        "base_url": base_url,
        "enabled": bool(enabled),
        "timeout_seconds": max(1, min(120, int(payload.timeout or payload.timeout_seconds or 10))),
        "fallback_model_id": payload.fallback_model_id,
    }


def provider_payload_capabilities(payload: ProviderConfigIn) -> list[str]:
    capabilities = normalize_capabilities(payload.capabilities)
    if payload.capability:
        capabilities.append(payload.capability)
    return capabilities


def apply_model_payload(db: Session, provider: ModelConfig, payload: ProviderConfigIn) -> ModelConfig:
    values = model_values_from_payload(payload)
    fallback_id = values.pop("fallback_model_id")
    if fallback_id:
        fallback = db.get(ModelConfig, fallback_id)
        if not fallback:
            raise HTTPException(status_code=404, detail="Fallback model not found")
        if fallback.id == provider.id:
            raise HTTPException(status_code=422, detail="Fallback model must be different from primary model")
    for field, value in values.items():
        setattr(provider, field, value)
    provider.fallback_model_id = fallback_id
    if payload.api_key and not payload.api_key.startswith("**"):
        provider.encrypted_api_key = encrypt_api_key(payload.api_key)
    provider.updated_at = now()
    db.flush()
    if provider.provider == "openai-compatible":
        test_result = {
            "status": "configured",
            "message": "OpenAI-compatible provider configured; use explicit test endpoint to verify connectivity",
        }
    else:
        test_result = test_model_connection(provider)
    provider.connection_status = test_result["status"]
    provider.last_test_message = test_result["message"]
    provider.last_test_at = now()
    db.flush()
    sync_model_capabilities(db, provider, provider_payload_capabilities(payload))
    return provider


def sync_model_capabilities(db: Session, provider: ModelConfig, capabilities: list[str] | str | None) -> None:
    requested = [capability for capability in normalize_capabilities(capabilities) if capability in AI_CAPABILITIES]
    if not requested:
        return
    for capability in requested:
        mapping = db.query(CapabilityModelMapping).filter(CapabilityModelMapping.capability == capability).first()
        if not mapping:
            mapping = CapabilityModelMapping(capability=capability, primary_model_id=provider.id, enabled=True)
            db.add(mapping)
        else:
            mapping.primary_model_id = provider.id
            if mapping.fallback_model_id == provider.id:
                mapping.fallback_model_id = None
            mapping.enabled = True
        mapping.updated_at = now()
        sync_gateway_mapping_from_legacy_models(db, capability, provider, None, True)


@dataclass
class AuthContext:
    user: User | None
    username: str
    display_name: str
    permissions: set[str]
    library_scope_ids: set[int] | None
    role_ids: set[int]
    is_super_admin: bool = False

    def has(self, permission_key: str) -> bool:
        return self.is_super_admin or permission_key in self.permissions


NON_SUPER_READ_PERMISSIONS = {
    "api.GET./api/v1/category-libraries",
    "api.GET./api/v1/category-libraries/{library_id}",
    "api.GET./api/v1/categories",
    "api.GET./api/v1/categories/{category_id}/attributes",
    "api.GET./api/v1/categories/{category_id}/attributes/own",
    "api.GET./api/v1/categories/{category_id}/properties",
    "api.GET./api/v1/product-names",
    "api.GET./api/v1/product-names/{product_name_id}",
    "api.GET./api/v1/brands",
    "api.GET./api/v1/attributes",
    "api.GET./api/v1/material-libraries",
    "api.GET./api/v1/material-libraries/{library_id}",
    "api.GET./api/v1/materials",
    "api.GET./api/v1/materials/{material_id}",
}

LIBRARY_ADMIN_PERMISSIONS = {
    "api.POST./api/v1/materials",
    "api.PUT./api/v1/materials/{material_id}",
    "api.DELETE./api/v1/materials/{material_id}",
    "api.PATCH./api/v1/materials/{material_id}/stop-purchase",
    "api.POST./api/v1/materials/{material_id}/transition",
    "api.POST./api/v1/materials/governance/preview",
    "api.POST./api/v1/materials/governance/import",
    "api.POST./api/v1/materials/ai-add/preview",
    "api.POST./api/v1/materials/ai-add/confirm",
    "api.POST./api/v1/materials/match",
    "api.PUT./api/v1/material-libraries/{library_id}",
    "button.material_archives.create",
    "button.material_archives.edit",
    "button.material_archives.delete",
    "button.material_archives.approval",
    "button.material_archives.import",
    "button.material_library.edit",
}


def permission_catalog_entries(db: Session | None = None) -> list[PermissionEntry]:
    entries = [PermissionEntry(**item) for item in PERMISSION_CATALOG]
    if db is not None:
        ensure_seed_material_context(db)
        libraries = db.query(MaterialLibrary).order_by(MaterialLibrary.id).all()
        entries.extend(
            PermissionEntry(
                module="material_library",
                permission_type="scope",
                permission_key=f"scope.material_library.{library.id}",
                label=f"Material Library Scope: {library.name}",
            )
            for library in libraries
        )
    return entries


def permission_catalog_by_key(db: Session | None = None) -> dict[str, PermissionEntry]:
    return {item.permission_key: item for item in permission_catalog_entries(db)}


def super_admin_auth(db: Session | None = None) -> AuthContext:
    scope_ids = None
    permissions = set(permission_catalog_by_key(db).keys())
    return AuthContext(
        user=None,
        username="super_admin",
        display_name="Seeded Administrator",
        permissions=permissions,
        library_scope_ids=scope_ids,
        role_ids=set(),
        is_super_admin=True,
    )


def regular_user_auth() -> AuthContext:
    return AuthContext(
        user=None,
        username="regular_user",
        display_name="Regular User",
        permissions=set(NON_SUPER_READ_PERMISSIONS),
        library_scope_ids=set(),
        role_ids=set(),
        is_super_admin=False,
    )


def effective_auth_for_user(user: User, db: Session) -> AuthContext:
    if user.status != "active":
        raise HTTPException(status_code=403, detail="User account is disabled")
    permissions: set[str] = set(NON_SUPER_READ_PERMISSIONS) if user.account_ownership == "HCM" else set()
    scope_ids: set[int] = set()
    enabled_roles = [link.role for link in user.role_links if link.role.enabled]
    role_ids = {role.id for role in enabled_roles}
    for role in enabled_roles:
        for permission in role.permissions:
            if not permission.enabled:
                continue
            permissions.add(permission.permission_key)
            if permission.permission_key.startswith("scope.material_library."):
                try:
                    scope_ids.add(int(permission.permission_key.rsplit(".", 1)[1]))
                except ValueError:
                    continue
    if role_ids:
        administered_library_ids = {
            library_id
            for (library_id,) in db.query(MaterialLibraryAdminRole.material_library_id)
            .filter(MaterialLibraryAdminRole.role_id.in_(role_ids))
            .all()
        }
        legacy_administered_library_ids = {
            library_id
            for (library_id,) in db.query(MaterialLibrary.id)
            .filter(MaterialLibrary.material_library_admin_id.in_(role_ids))
            .all()
        }
        administered_library_ids.update(legacy_administered_library_ids)
        scope_ids.update(administered_library_ids)
        if administered_library_ids:
            permissions.update(NON_SUPER_READ_PERMISSIONS)
            permissions.update(LIBRARY_ADMIN_PERMISSIONS)
    return AuthContext(
        user=user,
        username=user.username,
        display_name=user.display_name,
        permissions=permissions,
        library_scope_ids=scope_ids,
        role_ids=role_ids,
        is_super_admin=False,
    )


def current_auth(request: Request, db: Session) -> AuthContext:
    role_header = request.headers.get("X-User-Role", "").strip()
    if role_header == "super_admin":
        return super_admin_auth(db)
    if role_header:
        return regular_user_auth()
    user_id = request.headers.get("X-User-Id", "").strip()
    username = request.headers.get("X-Username", "").strip()
    if not user_id and not username:
        return super_admin_auth(db)
    user = db.get(User, int(user_id)) if user_id.isdigit() else None
    if not user and username:
        if username == "super_admin":
            return super_admin_auth(db)
        if username == "regular_user":
            return regular_user_auth()
        user = db.query(User).filter(User.username == username).first()
    if not user:
        raise HTTPException(status_code=403, detail="Authenticated user not found")
    return effective_auth_for_user(user, db)


def require_api_permission(permission_key: str):
    def dependency(request: Request, db: Session = Depends(get_db)) -> AuthContext:
        auth = current_auth(request, db)
        if not auth.has(permission_key):
            raise HTTPException(status_code=403, detail=f"Missing permission: {permission_key}")
        return auth

    return dependency


def require_button_permission(auth: AuthContext, permission_key: str) -> None:
    if not auth.has(permission_key):
        raise HTTPException(status_code=403, detail=f"Missing permission: {permission_key}")


def require_super_admin(auth: AuthContext) -> None:
    if not auth.is_super_admin:
        raise HTTPException(status_code=403, detail="super_admin role is required")


def is_library_in_scope(auth: AuthContext, library_id: int) -> bool:
    return auth.is_super_admin or library_id in (auth.library_scope_ids or set())


def require_library_scope(auth: AuthContext, library_id: int) -> None:
    if not is_library_in_scope(auth, library_id):
        raise HTTPException(status_code=403, detail="Material library is outside the user's permission scope")


def role_summary(role: Role) -> dict[str, Any]:
    return {"id": role.id, "name": role.name, "code": role.code, "enabled": role.enabled}


def user_summary(user: User) -> UserSummaryOut:
    return UserSummaryOut(
        id=user.id,
        username=user.username,
        display_name=user.display_name,
        unit=user.unit,
        department=user.department,
        team=user.team,
        account_ownership=user.account_ownership,
        status=user.status,
    )


def user_to_out(user: User) -> UserOut:
    roles = [role_summary(link.role) for link in sorted(user.role_links, key=lambda link: link.role.name)]
    return UserOut(
        id=user.id,
        username=user.username,
        display_name=user.display_name,
        hcm_id=user.hcm_id,
        unit=user.unit,
        department=user.department,
        team=user.team,
        email=user.email,
        account_ownership=user.account_ownership,
        account_owner=user.account_ownership,
        status=user.status,
        roles=roles,
        created_at=user.created_at.isoformat(),
        updated_at=user.updated_at.isoformat(),
    )


def permission_to_entry(permission: FeaturePermission) -> PermissionEntry:
    return PermissionEntry(
        module=permission.module,
        permission_type=permission.permission_type,
        permission_key=permission.permission_key,
        label=permission.label,
    )


def role_to_out(role: Role) -> RoleOut:
    users = [user_summary(link.user) for link in sorted(role.user_links, key=lambda link: link.user.username)]
    permissions = [
        permission_to_entry(permission)
        for permission in sorted(role.permissions, key=lambda item: (item.permission_type, item.permission_key))
        if permission.enabled
    ]
    return RoleOut(
        id=role.id,
        name=role.name,
        code=role.code,
        description=role.description,
        enabled=role.enabled,
        users=users,
        user_count=len(users),
        permissions=permissions,
        created_at=role.created_at.isoformat(),
        updated_at=role.updated_at.isoformat(),
    )


def require_local_user(user: User) -> None:
    if user.account_ownership != "local":
        raise HTTPException(status_code=409, detail="HCM-managed users cannot be locally edited, reset, or deleted")


def validate_user_status(status: str) -> str:
    if status not in USER_STATUSES:
        raise HTTPException(status_code=422, detail="User status must be active or disabled")
    return status


def validate_role_uniqueness(db: Session, name: str, code: str, role_id: int | None = None) -> None:
    query = db.query(Role).filter(or_(Role.name == name, Role.code == code))
    if role_id is not None:
        query = query.filter(Role.id != role_id)
    existing = query.first()
    if existing:
        field = "name" if existing.name == name else "code"
        raise HTTPException(status_code=409, detail=f"Role {field} must be unique")


def highest_generated_role_code_value(db: Session) -> int:
    highest = 0
    for (code,) in db.query(Role.code).filter(Role.code.like("ROLE_%")).all():
        match = ROLE_CODE_PATTERN.match(code or "")
        if match:
            highest = max(highest, int(match.group(1)))
    return highest


def ensure_role_code_sequence(db: Session) -> RoleCodeSequence:
    Base.metadata.create_all(bind=engine)
    sequence = (
        db.query(RoleCodeSequence)
        .filter(RoleCodeSequence.id == 1)
        .with_for_update()
        .one_or_none()
    )
    highest_existing = highest_generated_role_code_value(db)
    if sequence is None:
        sequence = RoleCodeSequence(id=1, current_value=highest_existing, updated_at=now())
        db.add(sequence)
        db.flush()
        return sequence
    if sequence.current_value < highest_existing:
        sequence.current_value = highest_existing
        sequence.updated_at = now()
        db.flush()
    return sequence


def generate_role_code(db: Session) -> str:
    sequence = ensure_role_code_sequence(db)
    while True:
        sequence.current_value += 1
        sequence.updated_at = now()
        code = f"ROLE_{sequence.current_value:03d}"
        if not db.query(Role).filter(Role.code == code).first():
            db.flush()
            return code


def highest_generated_product_name_code_value(db: Session) -> int:
    highest = 0
    for (code,) in db.query(ProductName.product_name_code).filter(ProductName.product_name_code.like("PM%")).all():
        match = PRODUCT_NAME_CODE_PATTERN.match(code or "")
        if match:
            highest = max(highest, int(match.group(1)))
    return highest


def ensure_product_name_code_sequence(db: Session) -> ProductNameCodeSequence:
    ensure_product_name_schema()
    sequence = (
        db.query(ProductNameCodeSequence)
        .filter(ProductNameCodeSequence.id == 1)
        .with_for_update()
        .one_or_none()
    )
    highest_existing = highest_generated_product_name_code_value(db)
    if sequence is None:
        sequence = ProductNameCodeSequence(id=1, current_value=highest_existing, updated_at=now())
        db.add(sequence)
        db.flush()
    elif sequence.current_value < highest_existing:
        sequence.current_value = highest_existing
        sequence.updated_at = now()
        db.flush()

    for product in db.query(ProductName).order_by(ProductName.id).all():
        if product.product_name_code:
            if product.status not in PRODUCT_NAME_STATUSES:
                product.status = "active"
            continue
        sequence.current_value += 1
        product.product_name_code = f"PM{sequence.current_value:08d}"
        product.status = product.status if product.status in PRODUCT_NAME_STATUSES else "active"
        sequence.updated_at = now()
    db.flush()

    db.connection().exec_driver_sql(
        "CREATE UNIQUE INDEX IF NOT EXISTS ix_product_names_product_name_code "
        "ON product_names (product_name_code)"
    )
    return sequence


def generate_product_name_code(db: Session) -> str:
    sequence = ensure_product_name_code_sequence(db)
    while True:
        sequence.current_value += 1
        sequence.updated_at = now()
        code = f"PM{sequence.current_value:08d}"
        if not db.query(ProductName).filter(ProductName.product_name_code == code).first():
            db.flush()
            return code


def validate_product_name_status(status: str) -> str:
    normalized = status.strip().lower()
    if normalized not in PRODUCT_NAME_STATUSES:
        raise HTTPException(status_code=422, detail="status must be active or inactive")
    return normalized


def product_name_to_out(product: ProductName) -> ProductNameOut:
    return ProductNameOut(
        id=product.id,
        product_name_code=product.product_name_code,
        status=product.status,
        name=product.name,
        unit=product.unit,
        category=product.category,
    )


def product_name_audit_value(product: ProductName) -> dict[str, Any]:
    return {
        "id": product.id,
        "product_name_code": product.product_name_code,
        "status": product.status,
        "name": product.name,
        "unit": product.unit,
        "category": product.category,
    }


def get_product_name_or_404(db: Session, product_name_id: int) -> ProductName:
    product = db.get(ProductName, product_name_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product name not found")
    return product


def get_role_or_404(db: Session, role_id: int) -> Role:
    role = db.get(Role, role_id)
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    return role


def get_user_or_404(db: Session, user_id: int) -> User:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user


def validate_bindable_role(role: Role) -> None:
    if not role.enabled:
        raise HTTPException(status_code=409, detail="Disabled roles cannot be bound to users")


def normalize_permission_payload(payload: RolePermissionsIn, db: Session) -> list[PermissionEntry]:
    catalog = permission_catalog_by_key(db)
    entries: list[PermissionEntry] = []
    seen: set[str] = set()
    for key in payload.permission_keys:
        if key not in catalog:
            raise HTTPException(status_code=422, detail=f"Unknown permission identifier: {key}")
        if key not in seen:
            entries.append(catalog[key])
            seen.add(key)
    for item in payload.permissions:
        if item.permission_key not in catalog:
            raise HTTPException(status_code=422, detail=f"Unknown permission identifier: {item.permission_key}")
        if item.permission_key not in seen:
            entries.append(catalog[item.permission_key])
            seen.add(item.permission_key)
    return entries


def now() -> datetime:
    return datetime.now(timezone.utc)


def normalize_options(value: list[str] | str | None) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [part.strip() for part in re.split(r"[,，、\s]+", value) if part.strip()]


def code_for(prefix: str, seed: str) -> str:
    digest = sha1(seed.encode("utf-8")).hexdigest()[:8].upper()
    return f"{prefix}-{digest}"


def next_unique_code(db: Session, model: type[Any], prefix: str, seed: str) -> str:
    base = code_for(prefix, seed)
    code = base
    suffix = 2
    while db.query(model).filter(model.code == code).first():
        code = f"{base}-{suffix}"
        suffix += 1
    return code


def product_by_payload(db: Session, product_name_id: int | None, product_name: str | None) -> ProductName:
    query = db.query(ProductName)
    product = query.filter(ProductName.id == product_name_id).first() if product_name_id else None
    if not product and product_name:
        product = query.filter(ProductName.name == product_name).first()
    if not product and product_name == SEED_PRODUCT["name"]:
        product = ensure_seed_product(db)
    if not product:
        raise HTTPException(status_code=404, detail="Product name not found")
    return product


def attribute_to_out(attribute: Attribute) -> AttributeOut:
    return AttributeOut(
        id=attribute.id,
        code=attribute.code,
        product_name_id=attribute.product_name_id,
        product_name=attribute.product_name.name,
        name=attribute.name,
        data_type=attribute.data_type,
        unit=attribute.unit,
        required=attribute.required,
        default_value=attribute.default_value,
        options=normalize_options(attribute.options),
        description=attribute.description,
        source=attribute.source,
        version=attribute.version,
        enabled=attribute.enabled,
    )


def change_to_out(change: AttributeChange) -> ChangeOut:
    return ChangeOut(
        id=change.id,
        attribute_id=change.attribute_id,
        attribute_code=change.attribute_code,
        attribute_name=change.attribute_name,
        version=change.version,
        operator=change.operator,
        changed_fields=json.loads(change.changed_fields or "[]"),
        before_values=json.loads(change.before_values or "{}"),
        after_values=json.loads(change.after_values or "{}"),
        created_at=change.created_at.isoformat(),
    )


def logo_to_model(brand: Brand, logo: BrandLogo) -> None:
    brand.logo_filename = logo.filename
    brand.logo_content_type = logo.content_type
    brand.logo_data_url = logo.data_url


def brand_to_out(brand: Brand) -> BrandOut:
    return BrandOut(
        id=brand.id,
        code=brand.code,
        name=brand.name,
        description=brand.description,
        enabled=brand.enabled,
        logo=BrandLogo(
            filename=brand.logo_filename,
            content_type=brand.logo_content_type,
            data_url=brand.logo_data_url,
        ),
    )


CODE_RULE_ALLOWED_RE = re.compile(r"^[A-Z0-9_-]+$")
CODE_RULE_SEGMENT_TYPES = {"fixed", "fixed_text", "category_path", "attribute_code", "date", "serial", "serial_number"}


def normalize_segment_type(segment: dict[str, Any]) -> str:
    raw = str(segment.get("type") or segment.get("segment_type") or "").strip().lower()
    if raw == "fixed_text":
        return "fixed"
    if raw == "serial_number":
        return "serial"
    if raw in CODE_RULE_SEGMENT_TYPES:
        return raw
    raise HTTPException(status_code=422, detail=f"Unsupported code rule segment type: {raw or '<empty>'}")


def normalize_code_rule_config(payload: dict[str, Any] | MaterialCodeRuleVersionIn | None) -> dict[str, Any]:
    if payload is None:
        raise HTTPException(status_code=422, detail="code_rule is required when auto_code_enabled is true")
    data = payload.model_dump(exclude_none=True) if isinstance(payload, MaterialCodeRuleVersionIn) else dict(payload)
    nested = data.get("rule_config") if isinstance(data.get("rule_config"), dict) else {}
    separator = str(data.get("separator", nested.get("separator", "")) or "")
    segments = data.get("segments", nested.get("segments", []))
    if not isinstance(segments, list) or not segments:
        raise HTTPException(status_code=422, detail="code rule must include at least one segment")
    normalized_segments: list[dict[str, Any]] = []
    for index, item in enumerate(segments):
        if not isinstance(item, dict):
            raise HTTPException(status_code=422, detail="code rule segments must be objects")
        segment = dict(item)
        segment["type"] = normalize_segment_type(segment)
        segment.setdefault("order", index + 1)
        normalized_segments.append(segment)
    return {"separator": separator, "segments": normalized_segments}


def code_rule_name(payload: dict[str, Any] | MaterialCodeRuleVersionIn | None, fallback: str = "Material code rule") -> str:
    if payload is None:
        return fallback
    data = payload.model_dump(exclude_none=True) if isinstance(payload, MaterialCodeRuleVersionIn) else dict(payload)
    return str(data.get("rule_name") or fallback).strip() or fallback


def validate_separator(separator: str) -> None:
    if not separator:
        return
    if len(separator) > 1 or not CODE_RULE_ALLOWED_RE.fullmatch(separator):
        raise HTTPException(status_code=422, detail="Code format only allows uppercase letters, digits, hyphen, and underscore")


def validate_code_rule_config(config: dict[str, Any]) -> None:
    separator = str(config.get("separator") or "")
    validate_separator(separator)
    segments = config.get("segments") or []
    has_uniqueness_segment = False
    estimated_length = 0
    for segment in segments:
        segment_type = normalize_segment_type(segment)
        if estimated_length and separator:
            estimated_length += len(separator)
        if segment_type == "fixed":
            literal = str(segment.get("value") or segment.get("text") or segment.get("literal") or "")
            if not literal or not CODE_RULE_ALLOWED_RE.fullmatch(literal):
                raise HTTPException(
                    status_code=422,
                    detail="Code format only allows uppercase letters, digits, hyphen, and underscore",
                )
            estimated_length += len(literal)
        elif segment_type == "date":
            fmt = str(segment.get("format") or "YYYYMMDD").upper()
            if fmt not in {"YYYY", "YYMM", "YYYYMM", "YYYYMMDD"}:
                raise HTTPException(status_code=422, detail="Unsupported date format")
            estimated_length += len(fmt.replace("Y", "0").replace("M", "0").replace("D", "0"))
        elif segment_type == "serial":
            has_uniqueness_segment = True
            length = int(segment.get("length") or segment.get("padding_length") or 3)
            if length < 1 or length > 10:
                raise HTTPException(status_code=422, detail="Serial length must be between 1 and 10")
            step = int(segment.get("step") or 1)
            if step < 1:
                raise HTTPException(status_code=422, detail="Serial step must be at least 1")
            estimated_length += length
        elif segment_type == "category_path":
            has_uniqueness_segment = True
            estimated_length += int(segment.get("length") or segment.get("max_length") or 8)
        elif segment_type == "attribute_code":
            has_uniqueness_segment = True
            attribute_name = str(segment.get("attribute") or segment.get("attribute_name") or segment.get("name") or "")
            if not attribute_name:
                raise HTTPException(status_code=422, detail="Attribute-code segment requires an attribute name")
            estimated_length += int(segment.get("length") or segment.get("max_length") or 8)
    if not has_uniqueness_segment:
        raise HTTPException(status_code=422, detail="At least one uniqueness-producing segment is required")
    if estimated_length > 64:
        raise HTTPException(status_code=422, detail="Generated material code maximum length is 64 characters")


def rule_config_dict(rule_version: MaterialCodeRuleVersion) -> dict[str, Any]:
    try:
        loaded = json.loads(rule_version.rule_config or "{}")
    except json.JSONDecodeError:
        loaded = {}
    if not isinstance(loaded, dict):
        loaded = {}
    return normalize_code_rule_config(loaded)


def code_rule_summary(rule_version: MaterialCodeRuleVersion | None) -> dict[str, Any] | None:
    if not rule_version:
        return None
    return {
        "id": rule_version.id,
        "version_no": rule_version.version_no,
        "version": rule_version.version_no,
        "version_label": f"V{rule_version.version_no}",
        "rule_name": rule_version.rule_name,
        "status": rule_version.status,
        "created_by": rule_version.created_by,
        "effective_time": rule_version.effective_time.isoformat() if rule_version.effective_time else None,
    }


def active_rule_for_library(db: Session, library: MaterialLibrary) -> MaterialCodeRuleVersion | None:
    if library.current_rule_version_id:
        rule = db.get(MaterialCodeRuleVersion, library.current_rule_version_id)
        if rule and rule.library_id == library.id:
            return rule
    return (
        db.query(MaterialCodeRuleVersion)
        .filter(MaterialCodeRuleVersion.library_id == library.id, MaterialCodeRuleVersion.status == "active")
        .order_by(MaterialCodeRuleVersion.version_no.desc(), MaterialCodeRuleVersion.id.desc())
        .first()
    )


def code_rule_version_to_out(rule_version: MaterialCodeRuleVersion) -> MaterialCodeRuleVersionOut:
    config = rule_config_dict(rule_version)
    return MaterialCodeRuleVersionOut(
        id=rule_version.id,
        library_id=rule_version.library_id,
        version_no=rule_version.version_no,
        version=rule_version.version_no,
        version_label=f"V{rule_version.version_no}",
        rule_name=rule_version.rule_name,
        rule_config=config,
        segments=list(config.get("segments") or []),
        separator=str(config.get("separator") or ""),
        status=rule_version.status,
        change_reason=rule_version.change_reason,
        created_by=rule_version.created_by,
        effective_time=rule_version.effective_time.isoformat() if rule_version.effective_time else None,
        created_at=rule_version.created_at.isoformat(),
        updated_at=rule_version.updated_at.isoformat(),
    )


def create_code_rule_version(
    db: Session,
    library: MaterialLibrary,
    payload: dict[str, Any] | MaterialCodeRuleVersionIn,
    status: str,
    created_by: str,
) -> MaterialCodeRuleVersion:
    config = normalize_code_rule_config(payload)
    validate_code_rule_config(config)
    latest = (
        db.query(func.max(MaterialCodeRuleVersion.version_no))
        .filter(MaterialCodeRuleVersion.library_id == library.id)
        .scalar()
        or 0
    )
    data = payload.model_dump(exclude_none=True) if isinstance(payload, MaterialCodeRuleVersionIn) else dict(payload)
    rule_version = MaterialCodeRuleVersion(
        library_id=library.id,
        version_no=int(latest) + 1,
        rule_name=code_rule_name(payload, f"{library.name} code rule"),
        rule_config=json.dumps(config, ensure_ascii=False),
        status=status,
        change_reason=str(data.get("change_reason") or "").strip(),
        created_by=created_by,
        effective_time=now() if status == "active" else None,
    )
    db.add(rule_version)
    db.flush()
    return rule_version


def unique_int_ids(values: list[int] | None) -> list[int]:
    if values is None:
        return []
    seen: set[int] = set()
    ids: list[int] = []
    for value in values:
        item_id = int(value)
        if item_id not in seen:
            ids.append(item_id)
            seen.add(item_id)
    return ids


def material_library_ids_from_payload(
    payload: MaterialLibraryIn | MaterialLibraryUpdate,
    current_admin_ids: list[int] | None = None,
    current_category_ids: list[int] | None = None,
) -> tuple[list[int] | None, list[int] | None]:
    fields_set = payload.model_fields_set
    admin_ids: list[int] | None = None
    category_ids: list[int] | None = None

    if "material_library_admin_ids" in fields_set:
        admin_ids = unique_int_ids(payload.material_library_admin_ids)
        if not admin_ids:
            raise HTTPException(status_code=422, detail="material_library_admin_ids is required")
    elif "material_library_admin_id" in fields_set:
        admin_ids = [] if payload.material_library_admin_id is None else [payload.material_library_admin_id]
    elif current_admin_ids is not None:
        admin_ids = current_admin_ids

    if "category_library_ids" in fields_set:
        category_ids = unique_int_ids(payload.category_library_ids)
        if not category_ids:
            raise HTTPException(status_code=422, detail="category_library_ids is required")
    elif "category_library_id" in fields_set:
        category_ids = [] if payload.category_library_id is None else [payload.category_library_id]
    elif current_category_ids is not None:
        category_ids = current_category_ids

    return admin_ids, category_ids


def validate_material_library_associations(
    db: Session,
    material_library_admin_ids: list[int],
    category_library_ids: list[int],
) -> tuple[list[Role], list[CategoryLibrary]]:
    roles = db.query(Role).filter(Role.id.in_(material_library_admin_ids)).order_by(Role.id).all() if material_library_admin_ids else []
    if len(roles) != len(set(material_library_admin_ids)):
        raise HTTPException(status_code=404, detail="Material library admin role not found")
    libraries = (
        db.query(CategoryLibrary).filter(CategoryLibrary.id.in_(category_library_ids)).order_by(CategoryLibrary.id).all()
        if category_library_ids
        else []
    )
    if len(libraries) != len(set(category_library_ids)):
        raise HTTPException(status_code=404, detail="Category library not found")
    return roles, libraries


def apply_material_library_associations(
    library: MaterialLibrary,
    roles: list[Role],
    category_libraries: list[CategoryLibrary],
) -> None:
    library.material_library_admins = roles
    library.category_libraries = category_libraries
    library.material_library_admin_id = roles[0].id if roles else None
    library.category_library_id = category_libraries[0].id if category_libraries else None


def material_library_association_snapshot(library: MaterialLibrary) -> dict[str, Any]:
    admin_ids = [role.id for role in library.material_library_admins]
    category_ids = [category_library.id for category_library in library.category_libraries]
    return {
        "material_library_admin_ids": admin_ids,
        "category_library_ids": category_ids,
        "material_library_admin_id": admin_ids[0] if admin_ids else library.material_library_admin_id,
        "category_library_id": category_ids[0] if category_ids else library.category_library_id,
    }


def library_access_role(library: MaterialLibrary, auth: AuthContext | None = None) -> tuple[str, str]:
    if auth is None or auth.is_super_admin or is_library_in_scope(auth, library.id):
        return "admin", "Admin"
    return "no_access", "No access"


def library_to_out(library: MaterialLibrary, db: Session | None = None, auth: AuthContext | None = None) -> MaterialLibraryOut:
    active_rule = active_rule_for_library(db, library) if db is not None else None
    material_count = db.query(Material).filter(Material.material_library_id == library.id).count() if db is not None else len(library.materials)
    admins = list(library.material_library_admins)
    category_libraries = list(library.category_libraries)
    admin = admins[0] if admins else library.material_library_admin
    category_library = category_libraries[0] if category_libraries else library.category_library
    access_role, access_role_label = library_access_role(library, auth)
    return MaterialLibraryOut(
        id=library.id,
        code=library.code,
        name=library.name,
        description=library.description,
        enabled=library.enabled,
        auto_code_enabled=library.auto_code_enabled,
        recode_enabled=library.recode_enabled,
        current_rule_version_id=library.current_rule_version_id,
        code_rule_summary=code_rule_summary(active_rule),
        material_count=material_count,
        material_library_admin_ids=[role.id for role in admins],
        material_library_admin_names=[role.name for role in admins],
        material_library_admin_codes=[role.code for role in admins],
        material_library_admin_id=admin.id if admin else library.material_library_admin_id,
        material_library_admin_name=admin.name if admin else None,
        material_library_admin_code=admin.code if admin else None,
        category_library_ids=[item.id for item in category_libraries],
        category_library_names=[item.name for item in category_libraries],
        category_library_codes=[item.code for item in category_libraries],
        category_library_id=category_library.id if category_library else library.category_library_id,
        category_library_name=category_library.name if category_library else None,
        category_library_code=category_library.code if category_library else None,
        access_role=access_role,
        access_role_label=access_role_label,
    )


def category_library_to_out(library: CategoryLibrary) -> CategoryLibraryOut:
    return CategoryLibraryOut(
        id=library.id,
        code=library.code,
        name=library.name,
        description=library.description,
        enabled=library.enabled,
        qdrant_enabled=library.qdrant_enabled,
    )


def category_to_out(category: Category) -> CategoryOut:
    return CategoryOut(
        id=category.id,
        code=category.code,
        name=category.name,
        category_library_id=category.category_library_id,
        category_library=category.category_library.name if category.category_library else "",
        parent_category_id=category.parent_category_id,
        description=category.description,
        enabled=category.enabled,
    )


CATEGORY_ATTRIBUTE_TYPES = {"string", "number", "enum", "date"}


def category_attribute_options(value: str | list[str] | None) -> list[str]:
    if isinstance(value, list):
        return [compact_space(str(item)) for item in value if compact_space(str(item))]
    if not value:
        return []
    try:
        loaded = json.loads(value)
    except json.JSONDecodeError:
        return []
    if isinstance(loaded, list):
        return [compact_space(str(item)) for item in loaded if compact_space(str(item))]
    return []


def normalize_category_attribute_type(payload: CategoryAttributeCreate | CategoryAttributeUpdate) -> str | None:
    raw_type = getattr(payload, "attr_type", None) if getattr(payload, "attr_type", None) is not None else getattr(payload, "data_type", None)
    if raw_type is None:
        return None
    attr_type = compact_space(str(raw_type)).lower()
    aliases = {"text": "string", "str": "string", "integer": "number", "float": "number", "decimal": "number"}
    attr_type = aliases.get(attr_type, attr_type)
    if attr_type not in CATEGORY_ATTRIBUTE_TYPES:
        raise HTTPException(status_code=422, detail=f"Invalid attr_type: {raw_type}")
    return attr_type


def validate_category_attribute_payload(
    db: Session,
    category: Category,
    payload: CategoryAttributeCreate | CategoryAttributeUpdate,
    existing: CategoryAttribute | None = None,
) -> dict[str, Any]:
    ensure_category_attribute_schema()
    values: dict[str, Any] = {}
    fields_set = payload.model_fields_set
    if isinstance(payload, CategoryAttributeCreate) or "name" in fields_set:
        name = compact_space(str(payload.name or ""))
        if not name:
            raise HTTPException(status_code=422, detail="Attribute name is required")
        duplicate = (
            db.query(CategoryAttribute)
            .filter(CategoryAttribute.category_id == category.id, CategoryAttribute.name == name)
        )
        if existing is not None:
            duplicate = duplicate.filter(CategoryAttribute.id != existing.id)
        if duplicate.first():
            raise HTTPException(status_code=409, detail="Attribute name already exists for this category")
        values["name"] = name
    attr_type = normalize_category_attribute_type(payload)
    if attr_type is not None:
        values["attr_type"] = attr_type
    elif isinstance(payload, CategoryAttributeCreate):
        values["attr_type"] = "string"

    for field in ["display_name_zh", "display_name_en", "default_value"]:
        if isinstance(payload, CategoryAttributeCreate) or field in fields_set:
            value = getattr(payload, field)
            values[field] = None if value is None else str(value).strip()
    for field in ["required", "allow_empty", "sort_order"]:
        if isinstance(payload, CategoryAttributeCreate) or field in fields_set:
            value = getattr(payload, field)
            if value is not None:
                values[field] = value
    if isinstance(payload, CategoryAttributeCreate) or "options" in fields_set:
        options = getattr(payload, "options", None) or []
        values["options"] = json.dumps(category_attribute_options(options), ensure_ascii=False)
    if values.get("attr_type") == "enum" and not category_attribute_options(values.get("options")):
        raise HTTPException(status_code=422, detail="Enum attributes require at least one option")
    if "sort_order" not in values and isinstance(payload, CategoryAttributeCreate):
        max_sort = (
            db.query(func.max(CategoryAttribute.sort_order))
            .filter(CategoryAttribute.category_id == category.id)
            .scalar()
        )
        values["sort_order"] = int(max_sort or 0) + 10
    return values


def category_attribute_to_read(
    attribute: CategoryAttribute,
    target_category_id: int,
    source_category: Category,
) -> CategoryAttributeRead:
    is_own = source_category.id == target_category_id
    inherited_from = None if is_own else source_category.id
    return CategoryAttributeRead(
        id=attribute.id,
        category_id=target_category_id,
        name=attribute.name,
        attr_type=attribute.attr_type,
        data_type=attribute.attr_type,
        display_name_zh=attribute.display_name_zh,
        display_name_en=attribute.display_name_en,
        options=category_attribute_options(attribute.options),
        required=attribute.required,
        allow_empty=attribute.allow_empty,
        default_value=attribute.default_value,
        sort_order=attribute.sort_order,
        inherited_from=inherited_from,
        inherited_from_category_id=inherited_from,
        inherited_from_category_name=None if is_own else source_category.name,
        source_attribute_id=attribute.id,
        source_category_id=source_category.id,
        source_category_name=source_category.name,
        is_own=is_own,
        is_inherited=not is_own,
        readonly=not is_own,
        created_at=attribute.created_at.isoformat(),
        updated_at=attribute.updated_at.isoformat(),
    )


def category_ancestor_chain(db: Session, category: Category) -> list[Category]:
    chain = [category]
    seen = {category.id}
    parent_id = category.parent_category_id
    while parent_id and parent_id not in seen:
        parent = db.get(Category, parent_id)
        if not parent:
            break
        chain.append(parent)
        seen.add(parent.id)
        parent_id = parent.parent_category_id
    return list(reversed(chain))


def compute_category_properties(db: Session, category_id: int) -> list[CategoryAttributeRead]:
    ensure_category_attribute_schema()
    category = db.get(Category, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    properties: list[CategoryAttributeRead] = []
    seen_source_ids: set[int] = set()
    for source_category in category_ancestor_chain(db, category):
        attributes = (
            db.query(CategoryAttribute)
            .filter(CategoryAttribute.category_id == source_category.id)
            .order_by(CategoryAttribute.sort_order, CategoryAttribute.id)
            .all()
        )
        for attribute in attributes:
            if attribute.id in seen_source_ids:
                continue
            seen_source_ids.add(attribute.id)
            properties.append(category_attribute_to_read(attribute, category.id, source_category))
    return properties


def category_property_list(db: Session, category_id: int) -> CategoryPropertyList:
    properties = compute_category_properties(db, category_id)
    own = [item for item in properties if item.is_own]
    inherited = [item for item in properties if item.is_inherited]
    return CategoryPropertyList(
        category_id=category_id,
        own=own,
        inherited=inherited,
        attributes=properties,
        properties=properties,
    )


def linked_category_library_ids(library: MaterialLibrary) -> set[int]:
    ids = {item.id for item in library.category_libraries}
    if library.category_library_id:
        ids.add(library.category_library_id)
    return ids


def validate_required_category_properties(
    db: Session,
    library: MaterialLibrary,
    category: Category,
    attributes: dict[str, Any],
) -> None:
    if not category.category_library_id or category.category_library_id not in linked_category_library_ids(library):
        return
    missing: list[str] = []
    for prop in compute_category_properties(db, category.id):
        if not (prop.required or not prop.allow_empty):
            continue
        value = attributes.get(prop.name)
        if value is None or (isinstance(value, str) and not value.strip()):
            missing.append(prop.name)
    if missing:
        raise HTTPException(
            status_code=422,
            detail={
                "error": "Missing required category properties",
                "missing_properties": missing,
            },
        )


QDRANT_URL = os.environ.get("QDRANT_URL", "http://localhost:6333").rstrip("/")
CATEGORY_EMBEDDING_DIM = int(os.environ.get("CATEGORY_EMBEDDING_DIM", "64"))


class QdrantSyncError(Exception):
    pass


def qdrant_collection_name(library_id: int) -> str:
    return f"category_library_{library_id}"


def qdrant_request(method: str, path: str, **kwargs: Any) -> httpx.Response:
    url = f"{QDRANT_URL}{path}"
    try:
        response = httpx.request(method, url, timeout=5, **kwargs)
    except httpx.RequestError as exc:
        raise QdrantSyncError(str(exc)) from exc
    return response


def qdrant_collection_exists(library_id: int) -> bool:
    name = qdrant_collection_name(library_id)
    response = qdrant_request("GET", f"/collections/{name}")
    if response.status_code == 200:
        return True
    if response.status_code == 404:
        return False
    raise QdrantSyncError(f"Qdrant collection lookup failed with HTTP {response.status_code}: {response.text[:200]}")


def trace_qdrant_error(db: Session, operation: str, error: str, metadata: dict[str, Any] | None = None) -> None:
    collector = SpanCollector(operation, "category_vector")
    collector.finish_span(collector.root_span_id, "error", error, metadata or {})
    collector.flush(db)


def qdrant_http_exception(operation: str, exc: QdrantSyncError) -> HTTPException:
    return HTTPException(status_code=502, detail=f"Qdrant {operation} failed: {exc}")


def category_embedding(text: str, dimension: int = CATEGORY_EMBEDDING_DIM) -> list[float]:
    tokens = re.findall(r"[a-z0-9]+|[\u4e00-\u9fff]", text.lower())
    if not tokens:
        tokens = [text.lower().strip() or "empty"]
    features = tokens[:]
    features.extend("".join(tokens[index : index + 2]) for index in range(max(0, len(tokens) - 1)))
    features.extend("".join(tokens[index : index + 3]) for index in range(max(0, len(tokens) - 2)))
    vector = [0.0] * dimension
    for feature in features:
        digest = sha256(feature.encode("utf-8")).digest()
        slot = int.from_bytes(digest[:4], "big") % dimension
        sign = 1.0 if digest[4] % 2 == 0 else -1.0
        vector[slot] += sign
    norm = sum(value * value for value in vector) ** 0.5
    if norm == 0:
        return [0.0] * dimension
    return [round(value / norm, 6) for value in vector]


def qdrant_health_payload() -> dict[str, Any]:
    try:
        response = qdrant_request("GET", "/collections")
    except QdrantSyncError as exc:
        return {"status": "unavailable", "available": False, "message": str(exc)}
    if response.status_code >= 400:
        return {"status": "unavailable", "available": False, "message": f"Qdrant HTTP {response.status_code}"}
    return {"status": "available", "available": True, "url": QDRANT_URL}


def create_qdrant_collection(library_id: int) -> None:
    name = qdrant_collection_name(library_id)
    if qdrant_collection_exists(library_id):
        return
    response = qdrant_request(
        "PUT",
        f"/collections/{name}",
        json={"vectors": {"size": CATEGORY_EMBEDDING_DIM, "distance": "Cosine"}},
    )
    if response.status_code == 409:
        return
    if response.status_code >= 400:
        raise QdrantSyncError(f"Qdrant collection create failed with HTTP {response.status_code}: {response.text[:200]}")


def delete_qdrant_collection(library_id: int) -> None:
    name = qdrant_collection_name(library_id)
    response = qdrant_request("DELETE", f"/collections/{name}")
    if response.status_code not in {200, 202, 404}:
        raise QdrantSyncError(f"Qdrant collection delete failed with HTTP {response.status_code}: {response.text[:200]}")


def category_payload(category: Category, by_id: dict[int, Category]) -> dict[str, Any]:
    path = category_path_for(category, by_id)
    return {
        "category_id": category.id,
        "level1": path[0] if len(path) > 0 else "",
        "level2": path[1] if len(path) > 1 else None,
        "level3": path[2] if len(path) > 2 else None,
        "level4": path[3] if len(path) > 3 else None,
        "level5": path[4] if len(path) > 4 else None,
        "path_string": " > ".join(path),
    }


def category_embedding_text(category: Category, payload: dict[str, Any]) -> str:
    parts = [
        str(payload.get("path_string") or ""),
        category.description or "",
        category.code or "",
    ]
    return " ".join(part for part in parts if part)


def upsert_category_point(library_id: int, category: Category, by_id: dict[int, Category]) -> None:
    payload = category_payload(category, by_id)
    vector = category_embedding(category_embedding_text(category, payload))
    response = qdrant_request(
        "PUT",
        f"/collections/{qdrant_collection_name(library_id)}/points?wait=true",
        json={"points": [{"id": category.id, "vector": vector, "payload": payload}]},
    )
    if response.status_code >= 400:
        raise QdrantSyncError(f"Qdrant point upsert failed with HTTP {response.status_code}: {response.text[:200]}")


def delete_category_point(library_id: int, category_id: int) -> None:
    response = qdrant_request(
        "POST",
        f"/collections/{qdrant_collection_name(library_id)}/points/delete?wait=true",
        json={"points": [category_id]},
    )
    if response.status_code not in {200, 202, 404}:
        raise QdrantSyncError(f"Qdrant point delete failed with HTTP {response.status_code}: {response.text[:200]}")


def enabled_library_for_category(db: Session, category: Category) -> CategoryLibrary | None:
    library = category.category_library or db.get(CategoryLibrary, category.category_library_id)
    if library and library.qdrant_enabled:
        return library
    return None


def qdrant_sync_category(db: Session, category: Category, operation: str = "upsert") -> bool:
    library = enabled_library_for_category(db, category)
    if not library:
        return False
    categories = db.query(Category).filter(Category.category_library_id == library.id).all()
    by_id = {item.id: item for item in categories}
    try:
        create_qdrant_collection(library.id)
        upsert_category_point(library.id, category, by_id)
        return True
    except QdrantSyncError as exc:
        trace_qdrant_error(db, f"qdrant.category.{operation}", str(exc), {"category_id": category.id, "library_id": library.id})
        raise


def qdrant_sync_category_subtree(db: Session, category: Category, operation: str = "update") -> int:
    library = enabled_library_for_category(db, category)
    if not library:
        return 0
    categories = db.query(Category).filter(Category.category_library_id == library.id).all()
    by_id = {item.id: item for item in categories}
    child_ids_by_parent: dict[int | None, list[int]] = {}
    for item in categories:
        child_ids_by_parent.setdefault(item.parent_category_id, []).append(item.id)
    pending = [category.id]
    affected: list[Category] = []
    while pending:
        current_id = pending.pop()
        current = by_id.get(current_id)
        if current:
            affected.append(current)
            pending.extend(child_ids_by_parent.get(current.id, []))
    try:
        create_qdrant_collection(library.id)
        for item in affected:
            upsert_category_point(library.id, item, by_id)
        return len(affected)
    except QdrantSyncError as exc:
        trace_qdrant_error(db, f"qdrant.category.{operation}", str(exc), {"category_id": category.id, "library_id": library.id})
        raise


def reembed_category_library(db: Session, library: CategoryLibrary) -> dict[str, Any]:
    categories = db.query(Category).filter(Category.category_library_id == library.id).order_by(Category.id).all()
    by_id = {item.id: item for item in categories}
    processed = 0
    failed = 0
    errors: list[str] = []
    try:
        create_qdrant_collection(library.id)
        for category in categories:
            try:
                upsert_category_point(library.id, category, by_id)
                processed += 1
            except QdrantSyncError as exc:
                failed += 1
                errors.append(str(exc))
    except QdrantSyncError as exc:
        failed = len(categories)
        errors.append(str(exc))
    if errors:
        trace_qdrant_error(db, "qdrant.category.reembed", errors[0], {"library_id": library.id, "failed": failed})
    status = "succeeded" if failed == 0 else "failed"
    return {
        "job_id": f"reembed-{library.id}-{uuid.uuid4().hex[:12]}",
        "status": status,
        "category_library_id": library.id,
        "total": len(categories),
        "processed": processed,
        "succeeded": processed,
        "failed": failed,
        "errors": errors[:3],
    }


def search_category_collection(library_id: int, query_vector: list[float], limit: int = 3) -> list[dict[str, Any]]:
    response = qdrant_request(
        "POST",
        f"/collections/{qdrant_collection_name(library_id)}/points/search",
        json={"vector": query_vector, "limit": limit, "with_payload": True},
    )
    if response.status_code == 404:
        return []
    if response.status_code >= 400:
        raise QdrantSyncError(f"Qdrant search failed with HTTP {response.status_code}: {response.text[:200]}")
    data = response.json()
    return data.get("result", []) if isinstance(data, dict) else []


def category_match_item(raw: dict[str, Any]) -> dict[str, Any] | None:
    payload = raw.get("payload") if isinstance(raw, dict) else None
    if not isinstance(payload, dict):
        return None
    category_id = int(payload.get("category_id") or raw.get("id") or 0)
    if category_id <= 0:
        return None
    score = clamp_confidence(raw.get("score"))
    item = {
        "id": category_id,
        "category_id": category_id,
        "path_string": str(payload.get("path_string") or ""),
        "level1": str(payload.get("level1") or ""),
        "level2": payload.get("level2"),
        "level3": payload.get("level3"),
        "level4": payload.get("level4"),
        "level5": payload.get("level5"),
        "score": score,
        "confidence": score,
    }
    item["category"] = {
        "id": category_id,
        "path_string": item["path_string"],
        "level1": item["level1"],
        "level2": item["level2"],
        "level3": item["level3"],
        "level4": item["level4"],
        "level5": item["level5"],
    }
    return item


def rule_category_to_out(category: RuleCategory, rule_count: int | None = None) -> RuleCategoryRead:
    count = len(category.rules) if rule_count is None else rule_count
    return RuleCategoryRead(
        id=category.id,
        slug=category.slug,
        display_name_zh=category.display_name_zh,
        display_name_en=category.display_name_en,
        description_zh=category.description_zh,
        description_en=category.description_en,
        icon=category.icon,
        sort_order=category.sort_order,
        created_at=category.created_at.isoformat(),
        rule_count=count,
    )


def rule_to_out(rule: Rule) -> RuleRead:
    return RuleRead(
        id=rule.id,
        category_id=rule.category_id,
        category_slug=rule.category.slug,
        category=rule_category_to_out(rule.category),
        name=rule.name,
        description=rule.description,
        pattern=rule.pattern,
        value=rule.value,
        options=json_options(rule.options),
        priority=rule.priority,
        enabled=rule.enabled,
        created_at=rule.created_at.isoformat(),
        updated_at=rule.updated_at.isoformat(),
    )


def material_attributes(value: str | dict[str, Any] | None) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if not value:
        return {}
    try:
        loaded = json.loads(value)
    except json.JSONDecodeError:
        return {}
    return loaded if isinstance(loaded, dict) else {}


def sanitize_code_part(value: str, field_name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "", str(value or "")).upper()
    if not cleaned or not CODE_RULE_ALLOWED_RE.fullmatch(cleaned):
        raise HTTPException(status_code=422, detail=f"Code format invalid for {field_name}")
    return cleaned


def render_date_segment(fmt: str) -> str:
    today = datetime.now()
    normalized = fmt.upper()
    if normalized == "YYYY":
        return today.strftime("%Y")
    if normalized == "YYMM":
        return today.strftime("%y%m")
    if normalized == "YYYYMM":
        return today.strftime("%Y%m")
    if normalized == "YYYYMMDD":
        return today.strftime("%Y%m%d")
    raise HTTPException(status_code=422, detail="Unsupported date format")


def attribute_mapping(segment: dict[str, Any]) -> dict[str, str]:
    mapping = (
        segment.get("value_to_code")
        or segment.get("value_to_code_mapping")
        or segment.get("mapping")
        or segment.get("mappings")
        or {}
    )
    if isinstance(mapping, list):
        result: dict[str, str] = {}
        for item in mapping:
            if isinstance(item, dict):
                key = str(item.get("value") or item.get("name") or "").strip()
                code = str(item.get("code") or "").strip()
                if key and code:
                    result[key] = code
        return result
    if isinstance(mapping, dict):
        return {str(key): str(value) for key, value in mapping.items()}
    return {}


def serial_scope_key(segment: dict[str, Any], material_data: dict[str, Any]) -> str:
    scope = str(segment.get("scope") or "global").strip().lower()
    if scope == "global":
        base = "global"
    elif scope == "category":
        base = f"category:{material_data['category'].id}"
    elif scope == "category_attribute":
        attribute_name = str(segment.get("scope_attribute") or segment.get("attribute") or segment.get("attribute_name") or "")
        value = material_data["attributes"].get(attribute_name, "")
        base = f"category_attribute:{material_data['category'].id}:{attribute_name}:{value}"
    elif scope == "year":
        base = f"year:{datetime.now().strftime('%Y')}"
    elif scope == "month":
        base = f"month:{datetime.now().strftime('%Y%m')}"
    else:
        raise HTTPException(status_code=422, detail="Unsupported serial scope")
    reset = str(segment.get("reset") or segment.get("reset_by") or "").strip().lower()
    if reset == "year" and not base.startswith("year:"):
        base = f"{base}:year:{datetime.now().strftime('%Y')}"
    if reset == "month" and not base.startswith("month:"):
        base = f"{base}:month:{datetime.now().strftime('%Y%m')}"
    return base


def next_serial_value(
    db: Session,
    library_id: int,
    rule_version_id: int,
    segment: dict[str, Any],
    material_data: dict[str, Any],
) -> str:
    length = int(segment.get("length") or segment.get("padding_length") or 3)
    start = int(segment.get("start") or segment.get("start_value") or 1)
    step = int(segment.get("step") or 1)
    scope_key = serial_scope_key(segment, material_data)
    serial = (
        db.query(MaterialCodeSerial)
        .filter(
            MaterialCodeSerial.library_id == library_id,
            MaterialCodeSerial.rule_version_id == rule_version_id,
            MaterialCodeSerial.scope_key == scope_key,
        )
        .first()
    )
    if not serial:
        serial = MaterialCodeSerial(
            library_id=library_id,
            rule_version_id=rule_version_id,
            scope_key=scope_key,
            current_value=start - step,
        )
        db.add(serial)
        db.flush()
    serial.current_value += step
    serial.updated_at = now()
    rendered = str(serial.current_value)
    if bool(segment.get("padding", True)) and str(segment.get("padding")) != "none":
        rendered = rendered.zfill(length)
    if len(rendered) > length:
        raise HTTPException(status_code=422, detail="Serial value exceeds configured serial length")
    return rendered


def generate_material_code(
    db: Session,
    tenant_id: str,
    library_id: int,
    material: dict[str, Any],
    rule_version: MaterialCodeRuleVersion,
) -> str:
    del tenant_id
    config = rule_config_dict(rule_version)
    validate_code_rule_config(config)
    parts: list[str] = []
    for segment in config["segments"]:
        segment_type = normalize_segment_type(segment)
        if segment_type == "fixed":
            parts.append(sanitize_code_part(str(segment.get("value") or segment.get("text") or segment.get("literal") or ""), "fixed segment"))
        elif segment_type == "date":
            parts.append(render_date_segment(str(segment.get("format") or "YYYYMMDD")))
        elif segment_type == "category_path":
            category = material["category"]
            source = str(segment.get("source") or "code")
            raw = category.name if source == "name" else category.code
            value = sanitize_code_part(raw, "category path")
            length = int(segment.get("length") or segment.get("max_length") or 0)
            parts.append(value[:length] if length else value)
        elif segment_type == "attribute_code":
            attribute_name = str(segment.get("attribute") or segment.get("attribute_name") or segment.get("name") or "")
            attributes = material["attributes"]
            if attribute_name not in attributes or attributes.get(attribute_name) in (None, ""):
                raise HTTPException(status_code=422, detail=f"Missing attribute for code generation: {attribute_name}")
            attribute_value = str(attributes[attribute_name])
            mapping = attribute_mapping(segment)
            raw = mapping.get(attribute_value) or mapping.get(attribute_value.strip()) or attribute_value
            parts.append(sanitize_code_part(raw, f"attribute {attribute_name}"))
        elif segment_type == "serial":
            parts.append(next_serial_value(db, library_id, rule_version.id, segment, material))
    code = str(config.get("separator") or "").join(parts)
    if len(code) > 64:
        raise HTTPException(status_code=422, detail="Generated material code maximum length is 64 characters")
    if not CODE_RULE_ALLOWED_RE.fullmatch(code):
        raise HTTPException(status_code=422, detail="Generated code format only allows uppercase letters, digits, hyphen, and underscore")
    duplicate = db.query(Material).filter(Material.code == code).first()
    if duplicate:
        raise HTTPException(status_code=409, detail="Generated material code must be unique")
    return code


def preview_serial_value(
    db: Session,
    library_id: int,
    rule_version_id: int,
    segment: dict[str, Any],
    material_data: dict[str, Any],
    serial_state: dict[tuple[int, str], int],
) -> str:
    length = int(segment.get("length") or segment.get("padding_length") or 3)
    start = int(segment.get("start") or segment.get("start_value") or 1)
    step = int(segment.get("step") or 1)
    scope_key = serial_scope_key(segment, material_data)
    key = (rule_version_id, scope_key)
    if key not in serial_state:
        serial = (
            db.query(MaterialCodeSerial)
            .filter(
                MaterialCodeSerial.library_id == library_id,
                MaterialCodeSerial.rule_version_id == rule_version_id,
                MaterialCodeSerial.scope_key == scope_key,
            )
            .first()
        )
        serial_state[key] = serial.current_value if serial else start - step
    serial_state[key] += step
    rendered = str(serial_state[key])
    if bool(segment.get("padding", True)) and str(segment.get("padding")) != "none":
        rendered = rendered.zfill(length)
    if len(rendered) > length:
        raise HTTPException(status_code=422, detail="Serial value exceeds configured serial length")
    return rendered


def generate_material_code_candidate(
    db: Session,
    library_id: int,
    material_model: Material,
    rule_version: MaterialCodeRuleVersion,
    serial_state: dict[tuple[int, str], int],
) -> str:
    config = rule_config_dict(rule_version)
    validate_code_rule_config(config)
    material_data = {
        "product": material_model.product_name,
        "library": material_model.material_library,
        "category": material_model.category,
        "attributes": material_attributes(material_model.attributes),
    }
    parts: list[str] = []
    for segment in config["segments"]:
        segment_type = normalize_segment_type(segment)
        if segment_type == "fixed":
            parts.append(sanitize_code_part(str(segment.get("value") or segment.get("text") or segment.get("literal") or ""), "fixed segment"))
        elif segment_type == "date":
            parts.append(render_date_segment(str(segment.get("format") or "YYYYMMDD")))
        elif segment_type == "category_path":
            category = material_data["category"]
            source = str(segment.get("source") or "code")
            raw = category.name if source == "name" else category.code
            value = sanitize_code_part(raw, "category path")
            length = int(segment.get("length") or segment.get("max_length") or 0)
            parts.append(value[:length] if length else value)
        elif segment_type == "attribute_code":
            attribute_name = str(segment.get("attribute") or segment.get("attribute_name") or segment.get("name") or "")
            attributes = material_data["attributes"]
            if attribute_name not in attributes or attributes.get(attribute_name) in (None, ""):
                raise HTTPException(status_code=422, detail=f"Missing attribute for code generation: {attribute_name}")
            attribute_value = str(attributes[attribute_name])
            mapping = attribute_mapping(segment)
            raw = mapping.get(attribute_value) or mapping.get(attribute_value.strip()) or attribute_value
            parts.append(sanitize_code_part(raw, f"attribute {attribute_name}"))
        elif segment_type == "serial":
            parts.append(preview_serial_value(db, library_id, rule_version.id, segment, material_data, serial_state))
    code = str(config.get("separator") or "").join(parts)
    if len(code) > 64:
        raise HTTPException(status_code=422, detail="Generated material code maximum length is 64 characters")
    if not CODE_RULE_ALLOWED_RE.fullmatch(code):
        raise HTTPException(status_code=422, detail="Generated code format only allows uppercase letters, digits, hyphen, and underscore")
    duplicate = db.query(Material).filter(Material.code == code, Material.id != material_model.id).first()
    if duplicate:
        raise HTTPException(status_code=409, detail="Generated material code must be unique")
    return code


def code_change_row_to_out(detail: MaterialCodeChangeDetail, material: Material | None = None) -> MaterialCodeChangeRowOut:
    return MaterialCodeChangeRowOut(
        id=detail.id,
        batch_id=detail.batch_id,
        material_id=detail.material_id,
        material_name=material.name if material else "",
        old_code=detail.old_code,
        new_code=detail.new_code,
        status=detail.status,
        error_message=detail.error_message,
    )


def code_change_batch_to_out(db: Session, batch: MaterialCodeChangeBatch, include_rows: bool = False) -> MaterialCodeChangeBatchOut:
    rows: list[MaterialCodeChangeRowOut] = []
    if include_rows:
        details = (
            db.query(MaterialCodeChangeDetail)
            .filter(MaterialCodeChangeDetail.batch_id == batch.id)
            .order_by(MaterialCodeChangeDetail.id)
            .all()
        )
        materials = {item.id: item for item in db.query(Material).filter(Material.id.in_([detail.material_id for detail in details] or [-1])).all()}
        rows = [code_change_row_to_out(detail, materials.get(detail.material_id)) for detail in details]
    return MaterialCodeChangeBatchOut(
        batch_id=batch.id,
        id=batch.id,
        library_id=batch.library_id,
        old_rule_version_id=batch.old_rule_version_id,
        new_rule_version_id=batch.new_rule_version_id,
        change_mode=batch.change_mode,
        total_count=batch.total_count,
        success_count=batch.success_count,
        failed_count=batch.failed_count,
        status=batch.status,
        rows=rows,
        created_at=batch.created_at.isoformat(),
        updated_at=batch.updated_at.isoformat(),
    )


def code_mapping_to_out(mapping: MaterialCodeMapping, material: Material | None = None) -> MaterialCodeMappingOut:
    return MaterialCodeMappingOut(
        id=mapping.id,
        library_id=mapping.library_id,
        material_id=mapping.material_id,
        material_name=material.name if material else "",
        old_code=mapping.old_code,
        new_code=mapping.new_code,
        old_rule_version_id=mapping.old_rule_version_id,
        new_rule_version_id=mapping.new_rule_version_id,
        batch_id=mapping.batch_id,
        status=mapping.status,
        created_at=mapping.created_at.isoformat(),
    )


def material_to_out(material: Material) -> MaterialOut:
    attributes = material_attributes(material.attributes)
    lifecycle_history = attributes.get("_lifecycle_history", [])
    if not isinstance(lifecycle_history, list):
        lifecycle_history = []
    return MaterialOut(
        id=material.id,
        code=material.code,
        name=material.name,
        product_name_id=material.product_name_id,
        product_name=material.product_name.name,
        material_library_id=material.material_library_id,
        material_library=material.material_library.name,
        category_id=material.category_id,
        category=material.category.name,
        unit=material.unit,
        brand_id=material.brand_id,
        brand=material.brand.name if material.brand else "",
        status=material.status,
        description=material.description,
        attributes=attributes,
        lifecycle_history=lifecycle_history,
        original_code=material.original_code,
        previous_code=material.previous_code,
        code_rule_version_id=material.code_rule_version_id,
        code_change_count=material.code_change_count,
        code_status=material.code_status,
        enabled=material.enabled,
        created_at=material.created_at.isoformat(),
        updated_at=material.updated_at.isoformat(),
    )


def workflow_payload(value: str | dict[str, Any] | None) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if not value:
        return {}
    try:
        loaded = json.loads(value)
    except json.JSONDecodeError:
        return {}
    return loaded if isinstance(loaded, dict) else {}


def history_to_out(item: WorkflowHistory) -> WorkflowHistoryOut:
    return WorkflowHistoryOut(
        id=item.id,
        actor=item.actor,
        node=item.node,
        action=item.action,
        from_status=item.from_status,
        to_status=item.to_status,
        comment=item.comment,
        created_at=item.created_at.isoformat(),
    )


def workflow_to_out(application: WorkflowApplication) -> WorkflowApplicationOut:
    return WorkflowApplicationOut(
        id=application.id,
        application_no=application.application_no,
        type=application.type,
        status=application.status,
        applicant=application.applicant,
        current_node=application.current_node,
        business_reason=application.business_reason,
        rejection_reason=application.rejection_reason,
        data=workflow_payload(application.payload),
        approval_history=[history_to_out(item) for item in application.history],
        created_resource_type=application.created_resource_type,
        created_resource_id=application.created_resource_id,
        created_at=application.created_at.isoformat(),
        updated_at=application.updated_at.isoformat(),
    )


def add_workflow_history(
    application: WorkflowApplication,
    action: str,
    actor: str,
    node: str,
    from_status: str,
    to_status: str,
    comment: str = "",
) -> None:
    application.history.append(
        WorkflowHistory(
            actor=actor,
            node=node,
            action=action,
            from_status=from_status,
            to_status=to_status,
            comment=comment,
        )
    )


def initial_workflow_state(mode: str) -> tuple[str, str]:
    if mode == "simple":
        return "pending_approval", "approver"
    return "pending_department_head", "department_head"


def next_approval_state(current_node: str, mode: str) -> tuple[str, str]:
    if mode == "simple" or current_node == "approver":
        return "approved", "approved"
    if current_node == "department_head":
        return "pending_asset_management", "asset_management"
    if current_node == "asset_management":
        return "approved", "approved"
    raise HTTPException(status_code=409, detail=f"Invalid workflow node for approval: {current_node}")


def application_no(seed: str) -> str:
    return f"APP-{sha1(seed.encode('utf-8')).hexdigest()[:10].upper()}"


def validate_reference_url(value: str) -> str:
    link = value.strip()
    if not re.match(r"^https?://[^\s]+$", link):
        raise HTTPException(status_code=422, detail="A valid reference mall link URL is required")
    return link


def normalize_reference_images(images: list[dict[str, Any] | str]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for index, image in enumerate(images, start=1):
        if isinstance(image, dict):
            data_url = str(image.get("data_url") or image.get("url") or image.get("content") or "").strip()
            filename = str(image.get("filename") or f"reference-{index}.png")
            content_type = str(image.get("content_type") or "image/png")
        else:
            data_url = str(image).strip()
            filename = f"reference-{index}.png"
            content_type = "image/png"
        if data_url:
            normalized.append({"filename": filename, "content_type": content_type, "data_url": data_url})
    if len(normalized) < 3:
        raise HTTPException(status_code=422, detail="Three required reference images must be uploaded before submission")
    return normalized


def material_summary(material: Material) -> dict[str, Any]:
    return {
        "material_id": material.id,
        "material_code": material.code,
        "material_name": material.name,
        "material_library_id": material.material_library_id,
        "material_library": material.material_library.name,
        "category_id": material.category_id,
        "category": material.category.name,
        "product_name_id": material.product_name_id,
        "product_name": material.product_name.name,
        "current_material_status": material.status,
    }


def required_stop_reason(payload: WorkflowApplicationIn) -> tuple[str, str]:
    reason = (payload.reason or payload.reason_code or "").strip()
    reason_code = (payload.reason_code or reason).strip()
    if not reason:
        raise HTTPException(status_code=422, detail="A stop workflow reason option is required")
    return reason_code, reason


def record_material_lifecycle(
    material: Material,
    from_status: str,
    to_status: str,
    reason: str,
    source: str,
    actor: str,
    application_no: str = "",
) -> None:
    attrs = material_attributes(material.attributes)
    history = attrs.get("_lifecycle_history")
    if not isinstance(history, list):
        history = []
    history.append(
        {
            "from_status": from_status,
            "to_status": to_status,
            "reason": reason,
            "source": source,
            "actor": actor,
            "application_no": application_no,
            "created_at": now().isoformat(),
        }
    )
    attrs["_lifecycle_history"] = history
    material.attributes = json.dumps(attrs, ensure_ascii=False)


def build_stop_workflow_payload(payload: WorkflowApplicationIn, db: Session) -> dict[str, Any]:
    if payload.material_id is None:
        raise HTTPException(status_code=422, detail="Target material is required")
    material = db.get(Material, payload.material_id)
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    reason_code, reason = required_stop_reason(payload)
    if payload.type == "stop_purchase":
        if material.status != "normal":
            raise HTTPException(status_code=409, detail="Stop purchase requires a material in normal status")
        target_status = "stop_purchase"
    elif payload.type == "stop_use":
        if material.status != "stop_purchase":
            raise HTTPException(status_code=409, detail="Stop use requires prior stop_purchase status")
        target_status = "stop_use"
    else:
        raise HTTPException(status_code=422, detail="Unsupported stop workflow application type")
    data = material_summary(material)
    data.update(
        {
            "reason_code": reason_code,
            "reason": reason,
            "business_reason": payload.business_reason,
            "from_status": material.status,
            "target_status": target_status,
            "irreversible": payload.type == "stop_use",
            "acknowledge_terminal": payload.acknowledge_terminal,
        }
    )
    return data


def build_workflow_payload(payload: WorkflowApplicationIn, db: Session) -> dict[str, Any]:
    if payload.type not in APPLICATION_TYPES:
        raise HTTPException(status_code=422, detail="Unsupported workflow application type")
    if not payload.business_reason.strip():
        raise HTTPException(status_code=422, detail="Business reason is required")

    if payload.type == "new_category":
        library = db.get(MaterialLibrary, payload.material_library_id) if payload.material_library_id else None
        if not library:
            raise HTTPException(status_code=404, detail="Material library not found")
        name = (payload.proposed_category_name or "").strip()
        if not name:
            raise HTTPException(status_code=422, detail="Proposed category name is required")
        parent = db.get(Category, payload.parent_category_id) if payload.parent_category_id else None
        if payload.parent_category_id and not parent:
            raise HTTPException(status_code=404, detail="Parent category not found")
        code = (payload.proposed_category_code or "").strip() or next_unique_code(db, Category, "CAT", name)
        return {
            "material_library_id": library.id,
            "material_library": library.name,
            "parent_category_id": parent.id if parent else None,
            "parent_category": parent.name if parent else "",
            "proposed_category_name": name,
            "proposed_category_code": code,
            "category_path_preview": f"{parent.name} / {name}" if parent else name,
            "description": payload.description,
            "business_reason": payload.business_reason,
        }

    if payload.type in {"stop_purchase", "stop_use"}:
        return build_stop_workflow_payload(payload, db)

    product, material_library, category = material_context_by_payload(
        db,
        payload.product_name_id,
        payload.material_library_id,
        payload.category_id,
    )
    name = (payload.material_name or "").strip()
    if not name:
        raise HTTPException(status_code=422, detail="Material name is required")
    if not payload.unit.strip() and not product.unit:
        raise HTTPException(status_code=422, detail="Unit is required")
    existing = db.query(Material).filter(Material.product_name_id == product.id, Material.name == name).first()
    duplicate_matches = material_matches(
        db,
        material_library.id,
        material_search_text(name, "", payload.description, payload.attributes),
        "",
        payload.attributes,
        3,
    )
    brand = db.get(Brand, payload.brand_id) if payload.brand_id else None
    if payload.brand_id and not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    return {
        "material_library_id": material_library.id,
        "material_library": material_library.name,
        "category_id": category.id,
        "category": category.name,
        "product_name_id": product.id,
        "product_name": product.name,
        "material_name": name,
        "unit": payload.unit.strip() or product.unit,
        "brand_id": brand.id if brand else None,
        "brand": brand.name if brand else "",
        "attributes": payload.attributes,
        "description": payload.description,
        "reference_mall_link": validate_reference_url(payload.reference_mall_link),
        "reference_images": normalize_reference_images(payload.reference_images),
        "duplicate_warning": {
            "existing_material": material_to_out(existing).model_dump() if existing else None,
            "top_matches": duplicate_matches,
        },
        "business_reason": payload.business_reason,
    }


def complete_workflow_application(application: WorkflowApplication, db: Session) -> None:
    data = workflow_payload(application.payload)
    if application.type == "new_category":
        name = str(data.get("proposed_category_name", "")).strip()
        existing = db.query(Category).filter(Category.name == name).first()
        proposed_code = str(data.get("proposed_category_code") or "").strip()
        accepted_code = proposed_code if proposed_code and not db.query(Category).filter(Category.code == proposed_code).first() else ""
        category_library = ensure_default_category_library(db)
        category = existing or Category(
            code=accepted_code or next_unique_code(db, Category, "CAT", name),
            name=name,
            category_library_id=category_library.id,
            description=str(data.get("description") or data.get("business_reason") or ""),
            enabled=True,
        )
        if existing and existing.category_library_id is None:
            existing.category_library_id = category_library.id
        if not existing:
            db.add(category)
            db.flush()
        application.created_resource_type = "category"
        application.created_resource_id = category.id
        return

    if application.type in {"stop_purchase", "stop_use"}:
        material = db.get(Material, int(data.get("material_id") or 0))
        if not material:
            raise HTTPException(status_code=404, detail="Material not found")
        expected_from = "normal" if application.type == "stop_purchase" else "stop_purchase"
        target_status = "stop_purchase" if application.type == "stop_purchase" else "stop_use"
        if material.status != expected_from:
            detail = (
                "Stop purchase requires a material in normal status"
                if application.type == "stop_purchase"
                else "Stop use requires prior stop_purchase status"
            )
            raise HTTPException(status_code=409, detail=detail)
        material.status = target_status
        material.updated_at = now()
        record_material_lifecycle(
            material,
            expected_from,
            target_status,
            str(data.get("reason") or data.get("business_reason") or ""),
            "workflow",
            application.current_node,
            application.application_no,
        )
        data["current_material_status"] = target_status
        data["approved_at"] = material.updated_at.isoformat()
        application.payload = json.dumps(data, ensure_ascii=False)
        application.created_resource_type = "material"
        application.created_resource_id = material.id
        return

    product, library, category = material_context_by_payload(
        db,
        int(data.get("product_name_id") or 0),
        int(data.get("material_library_id") or 0),
        int(data.get("category_id") or 0),
    )
    name = str(data.get("material_name", "")).strip()
    existing = db.query(Material).filter(Material.product_name_id == product.id, Material.name == name).first()
    if existing:
        raise HTTPException(status_code=409, detail="Material already exists for this product name")
    attrs = material_attributes(data.get("attributes"))
    attrs["_reference_mall_link"] = data.get("reference_mall_link")
    attrs["_reference_images"] = data.get("reference_images", [])
    material = Material(
        code=next_unique_code(db, Material, "MAT", f"{product.id}:{name}:{application.application_no}"),
        name=name,
        product_name_id=product.id,
        material_library_id=library.id,
        category_id=category.id,
        unit=str(data.get("unit") or product.unit),
        brand_id=data.get("brand_id"),
        status="normal",
        description=str(data.get("description") or data.get("business_reason") or ""),
        attributes=json.dumps(attrs, ensure_ascii=False),
        enabled=True,
    )
    db.add(material)
    db.flush()
    application.created_resource_type = "material"
    application.created_resource_id = material.id


def validate_material_status(status: str) -> str:
    if status not in MATERIAL_STATUSES:
        raise HTTPException(status_code=422, detail=f"Unsupported material status: {status}")
    return status


def enforce_material_transition(current: str, target: str, reason: str | None = None) -> None:
    validate_material_status(target)
    if current == target:
        return
    if (current, target) not in MATERIAL_TRANSITIONS:
        raise HTTPException(
            status_code=400,
            detail="Material status is non-reversible and must follow normal -> stop_purchase -> stop_use",
        )
    if not reason or not reason.strip():
        raise HTTPException(status_code=422, detail="A transition or exemption reason is required")


def material_context_by_payload(
    db: Session,
    product_name_id: int | None,
    material_library_id: int | None,
    category_id: int | None,
) -> tuple[ProductName, MaterialLibrary, Category]:
    product = product_by_payload(db, product_name_id, None)
    library, category = ensure_seed_material_context(db)
    if material_library_id:
        library = db.get(MaterialLibrary, material_library_id)
    if category_id:
        category = db.get(Category, category_id)
    if not library:
        raise HTTPException(status_code=404, detail="Material library not found")
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    return product, library, category


def material_row_value(row: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def attributes_from_row(row: dict[str, Any]) -> dict[str, Any]:
    known = {
        "name",
        "material_name",
        "物料名称",
        "unit",
        "单位",
        "brand",
        "品牌",
        "description",
        "描述",
        "attributes",
        "属性",
        "product_name",
        "product",
        "产品名称",
    }
    attributes: dict[str, Any] = {}
    raw = material_row_value(row, ["attributes", "属性"])
    if raw:
        try:
            loaded = json.loads(raw)
            if isinstance(loaded, dict):
                attributes.update({str(key): value for key, value in loaded.items()})
        except json.JSONDecodeError:
            for part in re.split(r"[;；|]+", raw):
                if not part.strip():
                    continue
                key, _, value = re.split(r"[:：=]", part, maxsplit=1)[0].strip(), "", ""
                if ":" in part:
                    key, value = part.split(":", 1)
                elif "：" in part:
                    key, value = part.split("：", 1)
                elif "=" in part:
                    key, value = part.split("=", 1)
                if key.strip() and value.strip():
                    attributes[key.strip()] = value.strip()
    for key, value in row.items():
        if key not in known and value is not None and str(value).strip():
            attributes[str(key).strip()] = str(value).strip()
    return attributes


def compact_space(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip(" ，,。.;；")


def text_matches(value: str, pattern: str, options: dict[str, Any]) -> bool:
    if not pattern:
        return False
    mode = str(options.get("match") or "regex").strip()
    if mode == "exact":
        return value == pattern
    if mode == "exact_ignore_case":
        return value.casefold() == pattern.casefold()
    try:
        return re.search(pattern, value, flags=re.IGNORECASE) is not None
    except re.error:
        return pattern.casefold() in value.casefold()


def attr_field(options: dict[str, Any], pattern: str, value: str = "") -> str:
    return str(options.get("field") or options.get("attribute") or pattern or value).strip()


def allowed_values(options: dict[str, Any], value: str) -> list[str]:
    raw = options.get("allowed") or options.get("values") or options.get("options")
    if isinstance(raw, list):
        return [str(item) for item in raw if str(item)]
    if isinstance(raw, str):
        return normalize_options(raw)
    return normalize_options(value)


def evaluate_rule(rule: Rule, payload: EvaluateRequest) -> EvaluateResult:
    options = options_as_dict(rule.options)
    slug = rule.category.slug
    passed = True
    message = "规则校验通过"
    suggestion = ""

    if slug == "unit_normalization":
        unit = payload.unit.strip()
        if text_matches(unit, rule.pattern, options) and unit != rule.value:
            passed = False
            suggestion = rule.value
            message = f"单位「{unit}」应标准化为「{rule.value}」。"

    elif slug == "brand_alias":
        brand = payload.brand.strip()
        if text_matches(brand, rule.pattern, options) and brand != rule.value:
            passed = False
            suggestion = rule.value
            message = f"品牌「{brand}」应归一为「{rule.value}」。"

    elif slug == "title_cleaning":
        name = payload.name
        try:
            cleaned = re.sub(rule.pattern or r"\s+", rule.value, name)
        except re.error:
            cleaned = name.replace(rule.pattern, rule.value)
        if options.get("strip", True):
            cleaned = compact_space(cleaned)
        if cleaned != name:
            passed = False
            suggestion = cleaned
            message = "物料标题存在多余空格或非法格式，建议使用清洗后的名称。"

    elif slug == "enum_validation":
        field = attr_field(options, rule.pattern)
        allowed = allowed_values(options, rule.value)
        actual = payload.attributes.get(field)
        if allowed and str(actual) not in allowed:
            passed = False
            suggestion = allowed[0]
            message = f"属性「{field}」的值「{actual}」不在允许范围：{', '.join(allowed)}。"

    elif slug == "required_field_check":
        field = attr_field(options, rule.pattern, rule.value)
        actual = payload.attributes.get(field)
        if actual is None or str(actual).strip() == "":
            passed = False
            suggestion = field
            message = f"缺少必填属性「{field}」，请补充该字段。"

    elif slug == "blackwhite_list":
        text = " ".join([payload.name, payload.brand, payload.unit, json.dumps(payload.attributes, ensure_ascii=False)])
        keywords = options.get("keywords")
        if isinstance(keywords, str):
            keyword_list = normalize_options(keywords)
        elif isinstance(keywords, list):
            keyword_list = [str(item) for item in keywords if str(item)]
        else:
            keyword_list = normalize_options(rule.pattern)
        mode = str(options.get("mode") or "blacklist")
        if mode == "whitelist":
            if keyword_list and not any(keyword.casefold() in text.casefold() for keyword in keyword_list):
                passed = False
                suggestion = "补充白名单关键词或调整物料描述"
                message = "物料未命中白名单关键词，请确认是否允许入库。"
        elif any(keyword.casefold() in text.casefold() for keyword in keyword_list):
            passed = False
            hit = next(keyword for keyword in keyword_list if keyword.casefold() in text.casefold())
            suggestion = f"移除或替换受限关键词「{hit}」"
            message = f"物料命中黑名单关键词「{hit}」，请移除或提交人工复核。"

    return EvaluateResult(
        category_slug=slug,
        rule_id=rule.id,
        rule_name=rule.name,
        passed=passed,
        message=message,
        suggestion=suggestion,
    )


def infer_product_category(text: str) -> tuple[str, str, str]:
    lowered = text.lower()
    if any(token in text for token in ["交换机", "端口", "千兆"]) or "switch" in lowered:
        return "交换机", "网络设备 / 交换机", "台"
    if any(token in text for token in ["打印机", "打印", "a4"]) or "printer" in lowered:
        return SEED_PRODUCT["name"], SEED_CATEGORY["name"], "台"
    if any(token in text for token in ["电缆", "线缆", "网线"]) or "cable" in lowered:
        return "线缆", "网络设备 / 线缆", "米"
    return "通用物料", "未分类 / 通用物料", "件"


def get_or_create_category(db: Session, name: str) -> Category:
    category = db.query(Category).filter(Category.name == name).first()
    if category:
        if category.category_library_id is None:
            category.category_library_id = ensure_default_category_library(db).id
        return category
    category_library = ensure_default_category_library(db)
    category = Category(
        code=next_unique_code(db, Category, "CAT", name),
        name=name,
        category_library_id=category_library.id,
        description="Created by AI material addition recommendation",
        enabled=True,
    )
    db.add(category)
    db.flush()
    return category


def get_or_create_product_name(db: Session, name: str, unit: str, category: str) -> ProductName:
    ensure_product_name_code_sequence(db)
    product = db.query(ProductName).filter(ProductName.name == name).first()
    if product:
        return product
    product = ProductName(
        name=name,
        product_name_code=generate_product_name_code(db),
        status="active",
        unit=unit,
        category=category,
    )
    db.add(product)
    db.flush()
    return product


def get_or_create_brand(db: Session, brand_name: str) -> Brand | None:
    if not brand_name:
        return None
    brand = db.query(Brand).filter(Brand.name == brand_name).first()
    if brand:
        return brand
    brand = Brand(
        code=next_unique_code(db, Brand, "BRAND", brand_name),
        name=brand_name,
        description="Created by AI material addition",
        enabled=True,
    )
    db.add(brand)
    db.flush()
    return brand


def extract_after_pattern(text: str, pattern: str) -> str:
    match = re.search(pattern, text, re.IGNORECASE)
    return compact_space(match.group(1)) if match else ""


def infer_material_name(text: str) -> str:
    patterns = [
        r"(?:申请新增|新增|添加|创建)\s*([^，,。；;\n]+)",
        r"material\s*[:：]\s*([^，,。；;\n]+)",
    ]
    for pattern in patterns:
        value = extract_after_pattern(text, pattern)
        if value:
            return value
    return compact_space(re.split(r"[，,。；;\n]", text.strip(), maxsplit=1)[0])[:80]


def infer_unit(text: str, data_type: str = "") -> str:
    unit = extract_after_pattern(text, r"单位\s*[:：]?\s*([台件个米套只箱包卷A-Za-z]+)")
    if unit:
        return unit
    if data_type == "number" and ("页" in text or "速度" in text):
        return "页/分钟"
    return ""


def infer_brand_name(text: str) -> str:
    brand = extract_after_pattern(text, r"品牌\s*[:：]?\s*([\u4e00-\u9fa5A-Za-z0-9_-]+)")
    if brand:
        return brand
    lowered = text.lower()
    for candidate in KNOWN_BRANDS:
        if candidate.lower() in lowered:
            return "华为" if candidate.lower() == "huawei" else candidate
    return ""


def extract_attribute_value(pattern: str, text: str, suffix: str = "") -> str:
    match = re.search(pattern, text, re.IGNORECASE)
    if not match:
        return ""
    value = compact_space(match.group(1))
    return f"{value}{suffix}" if suffix and value and suffix not in value else value


def extract_material_attributes(text: str) -> tuple[dict[str, Any], dict[str, str]]:
    attributes: dict[str, Any] = {}
    sources: dict[str, str] = {}

    port_count = extract_attribute_value(r"(?:端口数|端口|ports?)\s*[:：]?\s*(\d{1,4})", text)
    if not port_count:
        port_count = extract_attribute_value(r"(\d{1,4})\s*口", text)
    if port_count:
        attributes["端口数"] = port_count
        sources["端口数"] = "regex:端口数/口"

    speed = extract_attribute_value(r"(?:速率|速度|speed)\s*[:：]?\s*([0-9.]+\s*(?:Mbps|Gbps|MB/s|GB/s|兆|千兆)?)", text)
    if not speed and "千兆" in text:
        speed = "1000Mbps"
    if speed:
        attributes["速率"] = speed.replace(" ", "")
        sources["速率"] = "regex:速率/千兆"

    model = extract_attribute_value(r"(?:型号|model)\s*[:：]?\s*([A-Za-z0-9][A-Za-z0-9._-]{2,})", text)
    if not model:
        model_match = re.search(r"\b([A-Z][A-Z0-9]+(?:-[A-Z0-9]+){1,})\b", text)
        model = model_match.group(1) if model_match else ""
    if model:
        attributes["型号"] = model
        sources["型号"] = "regex:型号/model"

    scenario = extract_after_pattern(text, r"适用(?:于|场景)?\s*([^，,。；;\n]+)")
    if scenario:
        attributes["适用场景"] = scenario
        sources["适用场景"] = "regex:适用场景"

    print_speed = extract_attribute_value(r"打印速度\s*[:：]?\s*([0-9.]+\s*页/分钟)", text)
    if print_speed:
        attributes["打印速度"] = print_speed.replace(" ", "")
        sources["打印速度"] = "regex:打印速度"

    color_mode = extract_attribute_value(r"颜色模式\s*[:：]?\s*([\u4e00-\u9fa5A-Za-z]+)", text)
    if not color_mode and "彩色" in text:
        color_mode = "彩色"
    if color_mode:
        attributes["颜色模式"] = color_mode
        sources["颜色模式"] = "regex:颜色模式/彩色"

    paper_size = extract_attribute_value(r"(A[0-9])", text)
    if paper_size:
        attributes["纸张尺寸"] = paper_size
        sources["纸张尺寸"] = "regex:纸张尺寸"

    return attributes, sources


def material_search_text(name: str, brand: str, description: str, attributes: dict[str, Any]) -> str:
    attr_text = " ".join(f"{key} {value}" for key, value in attributes.items())
    return compact_space(f"{name} {brand} {description} {attr_text}")


def token_set(text: str) -> set[str]:
    lowered = text.lower()
    tokens = set(re.findall(r"[a-z0-9._-]+", lowered))
    tokens.update(re.findall(r"\d+\s*(?:口|端口|mbps|gbps|页/分钟|ppm|米|台)", lowered))
    chinese_phrases = [
        "交换机",
        "千兆",
        "端口",
        "网络",
        "接入",
        "打印机",
        "打印",
        "彩色",
        "黑白",
        "华为",
        "联想",
        "惠普",
        "办公",
    ]
    tokens.update(phrase for phrase in chinese_phrases if phrase in text)
    synonym_tokens = {
        "switch": "交换机",
        "gigabit": "千兆",
        "huawei": "华为",
        "printer": "打印机",
        "color": "彩色",
        "access": "接入",
        "network": "网络",
        "port": "端口",
        "ports": "端口",
    }
    tokens.update(mapped for token, mapped in synonym_tokens.items() if token in lowered)
    port_match = re.search(r"(\d{1,4})\s*(?:口|端口|ports?|port)", lowered)
    if port_match:
        tokens.add(f"端口数:{port_match.group(1)}")
    return {token.strip() for token in tokens if token.strip()}


def ngrams(text: str, size: int = 2) -> set[str]:
    cleaned = re.sub(r"\s+", "", text.lower())
    return {cleaned[index : index + size] for index in range(max(len(cleaned) - size + 1, 0))}


def jaccard(left: set[str], right: set[str]) -> float:
    if not left or not right:
        return 0.0
    return len(left & right) / len(left | right)


def semantic_features(text: str, attributes: dict[str, Any]) -> set[str]:
    lowered = text.lower()
    features: set[str] = set()
    if any(token in text for token in ["交换机", "端口", "网络接入"]) or "switch" in lowered:
        features.add("concept:network_switch")
    if any(token in text for token in ["千兆", "1000mbps", "1gbps"]) or "gigabit" in lowered:
        features.add("concept:gigabit")
    if any(token in text for token in ["打印机", "打印"]) or "printer" in lowered:
        features.add("concept:printer")
    if "彩色" in text or "color" in lowered:
        features.add("concept:color")
    if "a4" in lowered:
        features.add("concept:a4")
    for key, value in attributes.items():
        key_text = str(key)
        value_text = str(value)
        combined = f"{key_text} {value_text}".lower()
        if "端口" in key_text or "口" in value_text or "port" in combined:
            number = re.search(r"\d+", value_text)
            features.add(f"ports:{number.group(0) if number else value_text}")
        if "速率" in key_text or "speed" in combined or "mbps" in combined or "gbps" in combined:
            if "1000" in combined or "千兆" in value_text or "1g" in combined:
                features.add("concept:gigabit")
            normalized_speed = re.sub(r"\s+", "", value_text.lower())
            features.add(f"speed:{normalized_speed}")
        if "型号" in key_text or "model" in combined:
            features.add(f"model:{value_text.lower()}")
    return features


def text_similarity(left: str, right: str) -> float:
    token_score = jaccard(token_set(left), token_set(right))
    ngram_score = jaccard(ngrams(left), ngrams(right))
    left_clean = re.sub(r"\s+", "", left.lower())
    right_clean = re.sub(r"\s+", "", right.lower())
    containment = 0.0
    if left_clean and right_clean:
        shorter, longer = sorted([left_clean, right_clean], key=len)
        containment = len(shorter) / len(longer) if shorter in longer else 0.0
    return round(max(token_score, ngram_score, containment), 4)


def classify_match(score: float) -> str:
    if score >= 0.90:
        return "highly_duplicate"
    if score >= 0.75:
        return "suspicious"
    return "normal"


def match_score(
    query_text: str,
    query_brand: str,
    query_attributes: dict[str, Any],
    material: Material,
) -> dict[str, float | str | dict[str, Any]]:
    material_brand = material.brand.name if material.brand else ""
    material_attrs = material_attributes(material.attributes)
    candidate_text = material_search_text(material.name, material_brand, material.description, material_attrs)
    semantic_score = jaccard(semantic_features(query_text, query_attributes), semantic_features(candidate_text, material_attrs))
    text_score = text_similarity(query_text, candidate_text)
    if query_brand and material_brand:
        brand_score = 1.0 if query_brand.lower() == material_brand.lower() else 0.0
    else:
        brand_score = 0.5 if not query_brand and not material_brand else 0.0
    total_score = round(min(1.0, semantic_score * 0.4 + text_score * 0.4 + brand_score * 0.2), 4)
    return {
        "material": material_to_out(material).model_dump(),
        "score": total_score,
        "total_score": total_score,
        "semantic_score": round(semantic_score, 4),
        "text_score": round(text_score, 4),
        "brand_score": round(brand_score, 4),
        "classification": classify_match(total_score),
        "evidence": {
            "hybrid_search": "semantic + BM25 token overlap",
            "engine": "deterministic local fallback for Qdrant hybrid search",
        },
    }


def material_matches(
    db: Session,
    material_library_id: int,
    query_text: str,
    brand: str,
    attributes: dict[str, Any],
    top_k: int = 3,
) -> list[dict[str, Any]]:
    candidates = (
        db.query(Material)
        .filter(Material.material_library_id == material_library_id, Material.enabled.is_(True))
        .order_by(Material.id.desc())
        .all()
    )
    scored = [match_score(query_text, brand, attributes, material) for material in candidates]
    scored.sort(key=lambda item: (float(item["total_score"]), item["material"]["id"]), reverse=True)
    return scored[: max(1, min(top_k, 3))]


def decode_uploaded_rows(file_name: str, file_content: str) -> list[dict[str, Any]]:
    if not file_name or not file_content:
        return []
    encoded = file_content.split(",", 1)[1] if "," in file_content[:80] else file_content
    try:
        data = base64.b64decode(encoded)
    except (binascii.Error, ValueError) as exc:
        raise HTTPException(status_code=422, detail="Unable to decode uploaded material file") from exc
    if file_name.lower().endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="Excel parser is unavailable") from exc
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell or "").strip() for cell in rows[0]]
        return [
            {headers[index] or f"column_{index + 1}": value for index, value in enumerate(row)}
            for row in rows[1:]
            if any(cell is not None and str(cell).strip() for cell in row)
        ]
    text = data.decode("utf-8-sig")
    return list(csv.DictReader(StringIO(text)))


def parse_material_rows(rows: str | list[str] | list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    if not rows:
        return []
    if isinstance(rows, list) and all(isinstance(row, dict) for row in rows):
        return [dict(row) for row in rows]
    if isinstance(rows, str):
        text = rows.strip()
        if not text:
            return []
        first_line = text.splitlines()[0]
        if "," in first_line and any(header in first_line for header in ["name", "物料名称", "unit", "单位"]):
            return list(csv.DictReader(StringIO(text)))
        raw_rows = text.splitlines()
    else:
        raw_rows = [str(row) for row in rows]
    parsed: list[dict[str, Any]] = []
    for row in raw_rows:
        parts = [part.strip() for part in re.split(r"[,，\t|]+", str(row)) if part.strip()]
        if not parts:
            continue
        parsed.append(
            {
                "name": parts[0],
                "unit": parts[1] if len(parts) > 1 else "",
                "brand": parts[2] if len(parts) > 2 else "",
                "description": parts[3] if len(parts) > 3 else "",
                "attributes": parts[4] if len(parts) > 4 else "",
            }
        )
    return parsed


def material_governance_items(
    payload: MaterialGovernancePreviewIn,
    db: Session,
) -> list[dict[str, Any]]:
    product = product_by_payload(db, payload.product_name_id, payload.product_name) if payload.product_name_id or payload.product_name else ensure_seed_product(db)
    library, category = ensure_seed_material_context(db)
    if payload.material_library_id:
        library = db.get(MaterialLibrary, payload.material_library_id) or library
    if payload.category_id:
        category = db.get(Category, payload.category_id) or category
    rows = decode_uploaded_rows(payload.file_name, payload.file_content) or parse_material_rows(payload.rows)
    items: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        name = material_row_value(row, ["name", "material_name", "物料名称"])
        unit = material_row_value(row, ["unit", "单位"]) or product.unit
        brand_name = material_row_value(row, ["brand", "品牌"])
        description = material_row_value(row, ["description", "描述"])
        attributes = attributes_from_row(row)
        errors: list[str] = []
        if not name:
            errors.append("Material name is required")
        validation_status = "valid" if not errors else "invalid"
        confidence = 0.93 if validation_status == "valid" and attributes else 0.86 if validation_status == "valid" else 0.42
        items.append(
            {
                "source_row": index,
                "name": name,
                "code": code_for("MAT", f"{product.id}:{name}:{index}"),
                "product_name_id": product.id,
                "product_name": product.name,
                "material_library_id": library.id,
                "material_library": library.name,
                "category_id": category.id,
                "category": category.name,
                "unit": unit,
                "brand_name": brand_name,
                "description": description,
                "attributes": attributes,
                "status": "normal",
                "validation_status": validation_status,
                "errors": errors,
                "selectable": validation_status == "valid",
                "confidence": confidence,
            }
        )
    return items


def governance_items(rows: str | list[str]) -> list[dict[str, Any]]:
    raw_rows = rows.splitlines() if isinstance(rows, str) else rows
    items: list[dict[str, Any]] = []
    for index, row in enumerate(raw_rows, start=1):
        text = str(row).strip()
        if not text:
            continue
        parts = [part.strip() for part in re.split(r"[/,，\t|]+", text) if part.strip()]
        name = parts[0] if parts else text
        raw_options = parts[1] if len(parts) > 1 else ""
        raw_type = parts[2] if len(parts) > 2 else raw_options
        data_type = normalize_data_type(raw_type)
        options = normalize_options(raw_options) if data_type == "enum" else []
        standardized_name = standardize_attribute_name(name)
        items.append(
            {
                "source_row": index,
                "source_text": text,
                "name": standardized_name,
                "data_type": data_type,
                "unit": infer_unit(text, data_type),
                "required": False,
                "default_value": options[0] if options else "",
                "options": options,
                "description": f"AI governance standardized from row {index}",
                "source": "AI governance import",
                "code": code_for("ATTR", standardized_name + text),
                "confidence": 0.92 if data_type in {"number", "enum"} else 0.84,
            }
        )
    return items


def standardize_attribute_name(name: str) -> str:
    mapping = {
        "速度": "打印速度",
        "每分钟页数": "打印速度",
        "打印颜色": "颜色模式",
        "纸张尺寸": "纸张尺寸",
    }
    return mapping.get(name, name)


def normalize_data_type(value: str) -> str:
    text = value.lower()
    if any(token in text for token in ["数值", "数字", "number", "每分钟"]):
        return "number"
    if any(token in text for token in ["枚举", "enum", "黑白", "彩色", "a4"]):
        return "enum"
    if any(token in text for token in ["日期", "date"]):
        return "date"
    return "text"


def infer_unit(text: str, data_type: str = "") -> str:
    unit = extract_after_pattern(text, r"单位\s*[:：]?\s*([台件个米套只箱包卷A-Za-z]+)")
    if unit:
        return unit
    if data_type == "number" and ("页" in text or "速度" in text):
        return "页/分钟"
    return ""


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok", "version": API_VERSION}


@app.get("/metrics", response_class=PlainTextResponse)
def metrics() -> PlainTextResponse:
    return PlainTextResponse(render_prometheus_metrics(), media_type="text/plain; version=0.0.4")


@app.get("/api/v1/health/qdrant")
def qdrant_health(db: Session = Depends(get_db)) -> dict[str, Any]:
    payload = qdrant_health_payload()
    if not payload["available"]:
        trace_qdrant_error(db, "qdrant.health", str(payload.get("message") or "unavailable"))
    return payload


def auth_to_out(auth: AuthContext) -> AuthUserOut:
    roles = [role_summary(link.role) for link in sorted(auth.user.role_links, key=lambda link: link.role.name)] if auth.user else []
    return AuthUserOut(
        id=auth.user.id if auth.user else None,
        username=auth.username,
        display_name=auth.display_name,
        is_super_admin=auth.is_super_admin,
        permissions=sorted(auth.permissions),
        material_library_scope_ids=None if auth.library_scope_ids is None else sorted(auth.library_scope_ids),
        roles=roles,
    )


@app.get("/api/v1/auth/me", response_model=AuthUserOut)
def get_current_user(request: Request, db: Session = Depends(get_db)) -> AuthUserOut:
    return auth_to_out(current_auth(request, db))


@app.post("/api/v1/auth/login", response_model=AuthUserOut)
def login(payload: AuthLoginIn, db: Session = Depends(get_db)) -> AuthUserOut:
    username = payload.username.strip()
    if username == "super_admin":
        return auth_to_out(super_admin_auth(db))
    if username == "regular_user":
        return auth_to_out(regular_user_auth())
    user = db.query(User).filter(User.username == username).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return auth_to_out(effective_auth_for_user(user, db))


def get_rule_or_404(db: Session, rule_id: int) -> Rule:
    rule = db.get(Rule, rule_id)
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    return rule


def get_rule_category_or_404(db: Session, category_id: int) -> RuleCategory:
    category = db.get(RuleCategory, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Rule category not found")
    return category


@app.get("/api/v1/rules/categories", response_model=list[RuleCategoryRead])
def list_rule_categories(
    request: Request,
    db: Session = Depends(get_db),
) -> list[RuleCategoryRead]:
    current_auth(request, db)
    ensure_rule_engine_seed(db)
    count_rows = (
        db.query(Rule.category_id, func.count(Rule.id))
        .group_by(Rule.category_id)
        .all()
    )
    counts = {category_id: count for category_id, count in count_rows}
    categories = db.query(RuleCategory).order_by(RuleCategory.sort_order, RuleCategory.id).all()
    return [rule_category_to_out(category, counts.get(category.id, 0)) for category in categories]


@app.get("/api/v1/rules", response_model=RuleListResponse)
def list_rules(
    request: Request,
    category_id: int | None = None,
    search: str = "",
    enabled: bool | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> RuleListResponse:
    current_auth(request, db)
    ensure_rule_engine_seed(db)
    query = db.query(Rule).join(RuleCategory)
    if category_id is not None:
        query = query.filter(Rule.category_id == category_id)
    if search.strip():
        like = f"%{search.strip()}%"
        query = query.filter(
            or_(
                Rule.name.like(like),
                Rule.description.like(like),
                Rule.pattern.like(like),
                Rule.value.like(like),
            )
        )
    if enabled is not None:
        query = query.filter(Rule.enabled.is_(enabled))
    total = query.count()
    pages = (total + page_size - 1) // page_size if total else 0
    rules = (
        query.order_by(RuleCategory.sort_order, Rule.priority, Rule.id)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return RuleListResponse(
        items=[rule_to_out(rule) for rule in rules],
        total=total,
        page=page,
        page_size=page_size,
        pages=pages,
    )


@app.post("/api/v1/rules", response_model=RuleRead)
def create_rule(
    payload: RuleCreate,
    request: Request,
    db: Session = Depends(get_db),
) -> RuleRead:
    require_super_admin(current_auth(request, db))
    ensure_rule_engine_seed(db)
    category = get_rule_category_or_404(db, payload.category_id)
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=422, detail="Rule name is required")
    rule = Rule(
        category_id=category.id,
        name=name,
        description=payload.description.strip(),
        pattern=payload.pattern.strip(),
        value=payload.value.strip(),
        options=json.dumps(payload.options, ensure_ascii=False),
        priority=payload.priority,
        enabled=payload.enabled,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return rule_to_out(rule)


@app.post("/api/v1/rules/evaluate", response_model=EvaluateResponse)
def evaluate_rules(
    payload: EvaluateRequest,
    request: Request,
    db: Session = Depends(get_db),
) -> EvaluateResponse:
    current_auth(request, db)
    ensure_rule_engine_seed(db)
    collector = SpanCollector("rules.evaluate", "material_governance")
    try:
        resolution = model_for_capability(db, "material_governance")
    except CapabilityResolutionError as exc:
        collector.finish_span(collector.root_span_id, "error", str(exc), {"capability": exc.capability})
        collector.flush(db)
        raise capability_resolution_http_error(exc) from exc
    mark_root_trace_model(collector, resolution.model, resolution.model.model_name, resolution_trace_metadata(resolution))
    rules = (
        db.query(Rule)
        .join(RuleCategory)
        .filter(Rule.enabled.is_(True))
        .order_by(RuleCategory.sort_order, Rule.priority, Rule.id)
        .all()
    )
    results = [evaluate_rule(rule, payload) for rule in rules]
    collector.finish_span(collector.root_span_id, "ok", metadata={"rule_count": len(rules), **resolution_trace_metadata(resolution)})
    collector.flush(db)
    return EvaluateResponse(
        results=results,
        trace_id=collector.trace_id,
        provider=resolution.model.provider,
        model=resolution.model.model_name,
        resolution_source="capability_mapping",
    )


@app.get("/api/v1/rules/{rule_id}", response_model=RuleRead)
def get_rule(
    rule_id: int,
    request: Request,
    db: Session = Depends(get_db),
) -> RuleRead:
    current_auth(request, db)
    ensure_rule_engine_seed(db)
    return rule_to_out(get_rule_or_404(db, rule_id))


@app.put("/api/v1/rules/{rule_id}", response_model=RuleRead)
def update_rule(
    rule_id: int,
    payload: RuleUpdate,
    request: Request,
    db: Session = Depends(get_db),
) -> RuleRead:
    require_super_admin(current_auth(request, db))
    ensure_rule_engine_seed(db)
    rule = get_rule_or_404(db, rule_id)
    if payload.category_id is not None:
        rule.category_id = get_rule_category_or_404(db, payload.category_id).id
    if payload.name is not None:
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=422, detail="Rule name is required")
        rule.name = name
    if payload.description is not None:
        rule.description = payload.description.strip()
    if payload.pattern is not None:
        rule.pattern = payload.pattern.strip()
    if payload.value is not None:
        rule.value = payload.value.strip()
    if payload.options is not None:
        rule.options = json.dumps(payload.options, ensure_ascii=False)
    if payload.priority is not None:
        rule.priority = payload.priority
    if payload.enabled is not None:
        rule.enabled = payload.enabled
    rule.updated_at = now()
    db.commit()
    db.refresh(rule)
    return rule_to_out(rule)


@app.patch("/api/v1/rules/{rule_id}/toggle", response_model=RuleRead)
def toggle_rule(
    rule_id: int,
    payload: RuleToggle,
    request: Request,
    db: Session = Depends(get_db),
) -> RuleRead:
    require_super_admin(current_auth(request, db))
    ensure_rule_engine_seed(db)
    rule = get_rule_or_404(db, rule_id)
    rule.enabled = payload.enabled
    rule.updated_at = now()
    db.commit()
    db.refresh(rule)
    return rule_to_out(rule)


@app.delete("/api/v1/rules/{rule_id}")
def delete_rule(
    rule_id: int,
    request: Request,
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    require_super_admin(current_auth(request, db))
    ensure_rule_engine_seed(db)
    rule = get_rule_or_404(db, rule_id)
    db.delete(rule)
    db.commit()
    return {"deleted": True, "id": rule_id}


@app.get("/api/v1/product-names", response_model=list[ProductNameOut])
def list_product_names(
    status: str = Query("all"),
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/product-names")),
) -> list[ProductNameOut]:
    ensure_product_name_code_sequence(db)
    ensure_seed_product(db)
    normalized_status = status.strip().lower()
    if normalized_status not in {"all", *PRODUCT_NAME_STATUSES}:
        raise HTTPException(status_code=422, detail="status must be all, active, or inactive")
    query = db.query(ProductName)
    if normalized_status != "all":
        query = query.filter(ProductName.status == normalized_status)
    products = query.order_by(ProductName.id).all()
    return [product_name_to_out(product) for product in products]


@app.post("/api/v1/product-names", response_model=ProductNameOut)
def create_product_name(
    payload: ProductNameIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/product-names")),
) -> ProductNameOut:
    require_button_permission(auth, "button.product_names.create")
    ensure_product_name_code_sequence(db)
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=422, detail="Product name is required")
    if db.query(ProductName).filter(ProductName.name == name).first():
        raise HTTPException(status_code=409, detail="Product name must be unique")
    product = ProductName(
        name=name,
        product_name_code=generate_product_name_code(db),
        status="active",
        unit=payload.unit.strip(),
        category=payload.category.strip(),
    )
    db.add(product)
    db.flush()
    add_audit_log(db, auth, "product_name", "create", {}, product_name_audit_value(product))
    db.commit()
    db.refresh(product)
    return product_name_to_out(product)


@app.get("/api/v1/product-names/{product_name_id}", response_model=ProductNameOut)
def get_product_name(
    product_name_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/product-names/{product_name_id}")),
) -> ProductNameOut:
    ensure_product_name_code_sequence(db)
    return product_name_to_out(get_product_name_or_404(db, product_name_id))


@app.put("/api/v1/product-names/{product_name_id}", response_model=ProductNameOut)
def update_product_name(
    product_name_id: int,
    payload: ProductNameUpdate,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/product-names/{product_name_id}")),
) -> ProductNameOut:
    require_button_permission(auth, "button.product_names.edit")
    ensure_product_name_code_sequence(db)
    product = get_product_name_or_404(db, product_name_id)
    before = product_name_audit_value(product)
    if payload.name is not None:
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=422, detail="Product name is required")
        existing = db.query(ProductName).filter(ProductName.name == name, ProductName.id != product.id).first()
        if existing:
            raise HTTPException(status_code=409, detail="Product name must be unique")
        product.name = name
    if payload.unit is not None:
        product.unit = payload.unit.strip()
    if payload.category is not None:
        product.category = payload.category.strip()
    db.flush()
    after = product_name_audit_value(product)
    if before != after:
        add_audit_log(db, auth, "product_name", "update", before, after)
    db.commit()
    db.refresh(product)
    return product_name_to_out(product)


@app.patch("/api/v1/product-names/{product_name_id}/status", response_model=ProductNameOut)
def update_product_name_status(
    product_name_id: int,
    payload: ProductNameStatusUpdate,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PATCH./api/v1/product-names/{product_name_id}/status")),
) -> ProductNameOut:
    require_button_permission(auth, "button.product_names.edit")
    ensure_product_name_code_sequence(db)
    product = get_product_name_or_404(db, product_name_id)
    status = validate_product_name_status(payload.status)
    before = product_name_audit_value(product)
    product.status = status
    db.flush()
    after = product_name_audit_value(product)
    if before != after:
        add_audit_log(db, auth, "product_name", "status", before, after)
    db.commit()
    db.refresh(product)
    return product_name_to_out(product)


@app.delete("/api/v1/product-names/{product_name_id}")
def delete_product_name(
    product_name_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.DELETE./api/v1/product-names/{product_name_id}")),
) -> dict[str, Any]:
    require_button_permission(auth, "button.product_names.delete")
    ensure_product_name_code_sequence(db)
    product = get_product_name_or_404(db, product_name_id)
    before = product_name_audit_value(product)
    product.status = "inactive"
    db.flush()
    add_audit_log(db, auth, "product_name", "delete", before, product_name_audit_value(product))
    db.commit()
    return {"deleted": True, "id": product_name_id, "soft_deleted": True}


@app.get("/api/v1/material-libraries", response_model=list[MaterialLibraryOut])
def list_material_libraries(
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/material-libraries")),
) -> list[MaterialLibraryOut]:
    ensure_seed_material_context(db)
    query = db.query(MaterialLibrary)
    if not auth.is_super_admin:
        query = query.filter(MaterialLibrary.id.in_(auth.library_scope_ids or {-1}))
    libraries = query.order_by(MaterialLibrary.id).all()
    return [library_to_out(library, db, auth) for library in libraries]


@app.post("/api/v1/material-libraries", response_model=MaterialLibraryOut)
def create_material_library(
    payload: MaterialLibraryIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/material-libraries")),
) -> MaterialLibraryOut:
    require_button_permission(auth, "button.material_library.create")
    if payload.auto_code_enabled or payload.code_rule is not None:
        require_super_admin(auth)
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=422, detail="Material library name is required")
    if db.query(MaterialLibrary).filter(MaterialLibrary.name == name).first():
        raise HTTPException(status_code=409, detail="Material library name must be unique")
    if payload.auto_code_enabled:
        normalize_code_rule_config(payload.code_rule)
    admin_ids, category_ids = material_library_ids_from_payload(payload)
    roles, category_libraries = validate_material_library_associations(db, admin_ids or [], category_ids or [])
    library = MaterialLibrary(
        code=next_unique_code(db, MaterialLibrary, "MLIB", f"{name}:{now().isoformat()}"),
        name=name,
        description=payload.description.strip(),
        enabled=payload.enabled,
        auto_code_enabled=payload.auto_code_enabled,
        recode_enabled=payload.recode_enabled,
    )
    db.add(library)
    db.flush()
    apply_material_library_associations(library, roles, category_libraries)
    db.query(MaterialCodeRuleVersion).filter(MaterialCodeRuleVersion.library_id == library.id).delete()
    if payload.auto_code_enabled:
        rule_version = create_code_rule_version(db, library, payload.code_rule or {}, "active", auth.username)
        library.current_rule_version_id = rule_version.id
    add_audit_log(
        db,
        auth,
        "material_library",
        "create",
        {},
        {
            "id": library.id,
            "name": library.name,
            **material_library_association_snapshot(library),
        },
    )
    db.commit()
    db.refresh(library)
    return library_to_out(library, db, auth)


@app.get("/api/v1/material-libraries/{library_id}", response_model=MaterialLibraryOut)
def get_material_library(
    library_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/material-libraries/{library_id}")),
) -> MaterialLibraryOut:
    library = db.get(MaterialLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Material library not found")
    require_library_scope(auth, library.id)
    return library_to_out(library, db, auth)


@app.put("/api/v1/material-libraries/{library_id}", response_model=MaterialLibraryOut)
def update_material_library(
    library_id: int,
    payload: MaterialLibraryUpdate,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/material-libraries/{library_id}")),
) -> MaterialLibraryOut:
    require_button_permission(auth, "button.material_library.edit")
    library = db.get(MaterialLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Material library not found")
    require_library_scope(auth, library.id)
    before = material_library_association_snapshot(library)
    if payload.name is not None:
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=422, detail="Material library name is required")
        duplicate = db.query(MaterialLibrary).filter(MaterialLibrary.name == name, MaterialLibrary.id != library.id).first()
        if duplicate:
            raise HTTPException(status_code=409, detail="Material library name must be unique")
        library.name = name
    if payload.description is not None:
        library.description = payload.description.strip()
    if payload.enabled is not None:
        library.enabled = payload.enabled
    fields_set = payload.model_fields_set
    if {
        "material_library_admin_ids",
        "material_library_admin_id",
        "category_library_ids",
        "category_library_id",
    } & fields_set:
        next_admin_ids, next_category_library_ids = material_library_ids_from_payload(
            payload,
            [role.id for role in library.material_library_admins],
            [category_library.id for category_library in library.category_libraries],
        )
        next_admin_ids = next_admin_ids or []
        next_category_library_ids = next_category_library_ids or []
        if not auth.is_super_admin and not set(next_admin_ids).issubset(auth.role_ids):
            raise HTTPException(status_code=403, detail="Material library admin roles are outside the user's permission scope")
        roles, category_libraries = validate_material_library_associations(db, next_admin_ids, next_category_library_ids)
        apply_material_library_associations(library, roles, category_libraries)
    library.updated_at = now()
    after = material_library_association_snapshot(library)
    if before != after:
        add_audit_log(
            db,
            auth,
            "material_library",
            "update",
            before,
            {
                "id": library.id,
                "name": library.name,
                **after,
            },
        )
    db.commit()
    db.refresh(library)
    return library_to_out(library, db, auth)


@app.get("/api/v1/material-libraries/{library_id}/code-rules/current", response_model=MaterialCodeRuleVersionOut)
def get_current_code_rule(
    library_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/material-libraries/{library_id}")),
) -> MaterialCodeRuleVersionOut:
    library = db.get(MaterialLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Material library not found")
    require_library_scope(auth, library.id)
    rule_version = active_rule_for_library(db, library)
    if not rule_version:
        raise HTTPException(status_code=404, detail="Active code rule not found")
    return code_rule_version_to_out(rule_version)


@app.get("/api/v1/material-libraries/{library_id}/code-rules/versions", response_model=MaterialCodeRuleVersionListOut)
def list_code_rule_versions(
    library_id: int,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=10, ge=1, le=100),
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/material-libraries/{library_id}")),
) -> MaterialCodeRuleVersionListOut:
    library = db.get(MaterialLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Material library not found")
    require_library_scope(auth, library.id)
    query = db.query(MaterialCodeRuleVersion).filter(MaterialCodeRuleVersion.library_id == library.id)
    total = query.count()
    versions = (
        query.order_by(MaterialCodeRuleVersion.version_no.desc(), MaterialCodeRuleVersion.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return MaterialCodeRuleVersionListOut(
        items=[code_rule_version_to_out(rule_version) for rule_version in versions],
        total=total,
        page=page,
        page_size=page_size,
    )


@app.post("/api/v1/material-libraries/{library_id}/code-rules/versions", response_model=MaterialCodeRuleVersionOut)
def create_code_rule_version_endpoint(
    library_id: int,
    payload: MaterialCodeRuleVersionIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/material-libraries")),
) -> MaterialCodeRuleVersionOut:
    require_super_admin(auth)
    library = db.get(MaterialLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Material library not found")
    require_library_scope(auth, library.id)
    status = "active" if payload.activate else "draft"
    if payload.activate:
        current_rule = active_rule_for_library(db, library)
        if current_rule:
            current_rule.status = "deprecated"
            current_rule.updated_at = now()
    rule_version = create_code_rule_version(db, library, payload, status, auth.username)
    if payload.activate:
        library.current_rule_version_id = rule_version.id
        library.updated_at = now()
    db.commit()
    db.refresh(rule_version)
    return code_rule_version_to_out(rule_version)


@app.get("/api/v1/material-libraries/{library_id}/code-rules/versions/{version_id}", response_model=MaterialCodeRuleVersionOut)
def get_code_rule_version(
    library_id: int,
    version_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/material-libraries/{library_id}")),
) -> MaterialCodeRuleVersionOut:
    library = db.get(MaterialLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Material library not found")
    require_library_scope(auth, library.id)
    rule_version = db.get(MaterialCodeRuleVersion, version_id)
    if not rule_version or rule_version.library_id != library.id:
        raise HTTPException(status_code=404, detail="Code rule version not found")
    return code_rule_version_to_out(rule_version)


@app.post(
    "/api/v1/material-libraries/{library_id}/code-rules/versions/{version_id}/recode-preview",
    response_model=MaterialCodeChangeBatchOut,
)
def create_recode_preview(
    library_id: int,
    version_id: int,
    payload: RecodePreviewIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/material-libraries")),
) -> MaterialCodeChangeBatchOut:
    require_super_admin(auth)
    library = db.get(MaterialLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Material library not found")
    require_library_scope(auth, library.id)
    if not library.recode_enabled:
        raise HTTPException(status_code=409, detail="Material library recoding is not enabled")
    new_rule = db.get(MaterialCodeRuleVersion, version_id)
    if not new_rule or new_rule.library_id != library.id:
        raise HTTPException(status_code=404, detail="Code rule version not found")
    old_rule = active_rule_for_library(db, library)
    scope = payload.scope.strip().lower()
    query = db.query(Material).filter(Material.material_library_id == library.id).order_by(Material.id)
    if scope == "selected":
        if not payload.material_ids:
            raise HTTPException(status_code=422, detail="material_ids are required when scope is selected")
        requested_ids = list(dict.fromkeys(payload.material_ids))
        materials = query.filter(Material.id.in_(requested_ids)).all()
        found_ids = {item.id for item in materials}
        missing_ids = [item_id for item_id in requested_ids if item_id not in found_ids]
        if missing_ids:
            raise HTTPException(status_code=404, detail=f"Selected materials not found in library: {missing_ids}")
        materials.sort(key=lambda item: requested_ids.index(item.id))
    elif scope == "all":
        materials = query.all()
    else:
        raise HTTPException(status_code=422, detail="scope must be all or selected")

    batch = MaterialCodeChangeBatch(
        library_id=library.id,
        old_rule_version_id=old_rule.id if old_rule else None,
        new_rule_version_id=new_rule.id,
        change_mode=scope,
        total_count=len(materials),
        status="preview",
    )
    db.add(batch)
    db.flush()

    serial_state: dict[tuple[int, str], int] = {}
    seen_codes: set[str] = set()
    success_count = 0
    failed_count = 0
    for material in materials:
        new_code = ""
        status = "success"
        error_message = ""
        try:
            new_code = generate_material_code_candidate(db, library.id, material, new_rule, serial_state)
            if new_code in seen_codes:
                raise HTTPException(status_code=409, detail="Generated material code must be unique within preview batch")
            seen_codes.add(new_code)
            success_count += 1
        except HTTPException as exc:
            status = "failed"
            error_message = str(exc.detail)
            failed_count += 1
        db.add(
            MaterialCodeChangeDetail(
                batch_id=batch.id,
                material_id=material.id,
                old_code=material.code,
                new_code=new_code,
                status=status,
                error_message=error_message,
            )
        )
    batch.success_count = success_count
    batch.failed_count = failed_count
    batch.updated_at = now()
    add_audit_log(
        db,
        auth,
        "material_code_change_batch",
        "preview",
        {"library_id": library.id, "new_rule_version_id": new_rule.id},
        {"batch_id": batch.id, "total_count": batch.total_count, "success_count": success_count, "failed_count": failed_count},
    )
    db.commit()
    db.refresh(batch)
    return code_change_batch_to_out(db, batch, include_rows=True)


@app.get("/api/v1/material-code-change-batches/{batch_id}", response_model=MaterialCodeChangeBatchOut)
def get_code_change_batch(
    batch_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/material-libraries/{library_id}")),
) -> MaterialCodeChangeBatchOut:
    batch = db.get(MaterialCodeChangeBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Material code change batch not found")
    require_library_scope(auth, batch.library_id)
    return code_change_batch_to_out(db, batch, include_rows=False)


@app.get("/api/v1/material-code-change-batches/{batch_id}/preview", response_model=MaterialCodeChangePreviewListOut)
def get_code_change_preview(
    batch_id: int,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=10, ge=1, le=100),
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/material-libraries/{library_id}")),
) -> MaterialCodeChangePreviewListOut:
    batch = db.get(MaterialCodeChangeBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Material code change batch not found")
    require_library_scope(auth, batch.library_id)
    query = db.query(MaterialCodeChangeDetail).filter(MaterialCodeChangeDetail.batch_id == batch.id)
    total = query.count()
    details = query.order_by(MaterialCodeChangeDetail.id).offset((page - 1) * page_size).limit(page_size).all()
    materials = {item.id: item for item in db.query(Material).filter(Material.id.in_([detail.material_id for detail in details] or [-1])).all()}
    return MaterialCodeChangePreviewListOut(
        items=[code_change_row_to_out(detail, materials.get(detail.material_id)) for detail in details],
        total=total,
        page=page,
        page_size=page_size,
    )


@app.post("/api/v1/material-code-change-batches/{batch_id}/execute", response_model=MaterialCodeChangeBatchOut)
def execute_code_change_batch(
    batch_id: int,
    payload: BatchActionIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/material-libraries")),
) -> MaterialCodeChangeBatchOut:
    require_super_admin(auth)
    if not payload.confirm:
        raise HTTPException(status_code=422, detail="Explicit confirmation is required to execute a recode batch")
    batch = db.get(MaterialCodeChangeBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Material code change batch not found")
    require_library_scope(auth, batch.library_id)
    if batch.status != "preview":
        raise HTTPException(status_code=409, detail=f"Batch is already {batch.status} and cannot be executed again")
    if batch.failed_count:
        raise HTTPException(status_code=409, detail="Failed preview rows or validation errors block execution")
    library = db.get(MaterialLibrary, batch.library_id)
    new_rule = db.get(MaterialCodeRuleVersion, batch.new_rule_version_id) if batch.new_rule_version_id else None
    old_rule = db.get(MaterialCodeRuleVersion, batch.old_rule_version_id) if batch.old_rule_version_id else None
    if not library or not new_rule:
        raise HTTPException(status_code=404, detail="Batch library or target rule version not found")
    details = (
        db.query(MaterialCodeChangeDetail)
        .filter(MaterialCodeChangeDetail.batch_id == batch.id, MaterialCodeChangeDetail.status == "success")
        .order_by(MaterialCodeChangeDetail.id)
        .all()
    )
    material_ids = [detail.material_id for detail in details]
    materials = {item.id: item for item in db.query(Material).filter(Material.id.in_(material_ids or [-1])).all()}
    before = {"batch_id": batch.id, "status": batch.status, "codes": {str(detail.material_id): detail.old_code for detail in details}}
    for detail in details:
        material = materials.get(detail.material_id)
        if not material:
            raise HTTPException(status_code=404, detail=f"Material not found for preview row: {detail.material_id}")
        generated = generate_material_code(
            db,
            "default",
            library.id,
            {
                "product": material.product_name,
                "library": material.material_library,
                "category": material.category,
                "attributes": material_attributes(material.attributes),
            },
            new_rule,
        )
        if generated != detail.new_code:
            raise HTTPException(status_code=409, detail="Preview is stale because serial counters or material data changed")
        material.original_code = material.original_code or material.code
        material.previous_code = material.code
        material.code = detail.new_code
        material.code_rule_version_id = new_rule.id
        material.code_change_count = (material.code_change_count or 0) + 1
        material.code_status = "active"
        material.updated_at = now()
        detail.status = "executed"
        db.add(
            MaterialCodeMapping(
                library_id=library.id,
                material_id=material.id,
                old_code=detail.old_code,
                new_code=detail.new_code,
                old_rule_version_id=batch.old_rule_version_id,
                new_rule_version_id=batch.new_rule_version_id,
                batch_id=batch.id,
                status="active",
            )
        )
    if old_rule and old_rule.id != new_rule.id:
        old_rule.status = "deprecated"
        old_rule.updated_at = now()
    new_rule.status = "active"
    new_rule.effective_time = new_rule.effective_time or now()
    new_rule.updated_at = now()
    library.current_rule_version_id = new_rule.id
    library.updated_at = now()
    batch.status = "executed"
    batch.updated_at = now()
    add_audit_log(
        db,
        auth,
        "material_code_change_batch",
        "execute",
        before,
        {"batch_id": batch.id, "status": batch.status, "codes": {str(detail.material_id): detail.new_code for detail in details}},
    )
    db.commit()
    db.refresh(batch)
    return code_change_batch_to_out(db, batch, include_rows=True)


@app.post("/api/v1/material-code-change-batches/{batch_id}/rollback", response_model=MaterialCodeChangeBatchOut)
def rollback_code_change_batch(
    batch_id: int,
    payload: BatchActionIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/material-libraries")),
) -> MaterialCodeChangeBatchOut:
    require_super_admin(auth)
    if not payload.confirm:
        raise HTTPException(status_code=422, detail="Explicit confirmation is required to rollback a recode batch")
    batch = db.get(MaterialCodeChangeBatch, batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Material code change batch not found")
    require_library_scope(auth, batch.library_id)
    if batch.status != "executed":
        raise HTTPException(status_code=409, detail=f"Batch is {batch.status} and is no longer rollbackable")
    library = db.get(MaterialLibrary, batch.library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Batch library not found")
    details = (
        db.query(MaterialCodeChangeDetail)
        .filter(MaterialCodeChangeDetail.batch_id == batch.id, MaterialCodeChangeDetail.status == "executed")
        .order_by(MaterialCodeChangeDetail.id)
        .all()
    )
    material_ids = [detail.material_id for detail in details]
    materials = {item.id: item for item in db.query(Material).filter(Material.id.in_(material_ids or [-1])).all()}
    before = {"batch_id": batch.id, "status": batch.status, "codes": {str(detail.material_id): detail.new_code for detail in details}}
    for detail in details:
        material = materials.get(detail.material_id)
        if not material:
            raise HTTPException(status_code=404, detail=f"Material not found for preview row: {detail.material_id}")
        duplicate = db.query(Material).filter(Material.code == detail.old_code, Material.id != material.id).first()
        if duplicate:
            raise HTTPException(status_code=409, detail="Rollback would create a duplicate material code")
        material.code = detail.old_code
        material.previous_code = detail.new_code
        material.original_code = material.original_code or detail.old_code
        material.code_rule_version_id = batch.old_rule_version_id
        material.code_change_count = max(0, (material.code_change_count or 0) - 1)
        material.code_status = "active"
        material.updated_at = now()
        detail.status = "rolled_back"
    mappings = db.query(MaterialCodeMapping).filter(MaterialCodeMapping.batch_id == batch.id).all()
    for mapping in mappings:
        mapping.status = "rolled_back"
    old_rule = db.get(MaterialCodeRuleVersion, batch.old_rule_version_id) if batch.old_rule_version_id else None
    new_rule = db.get(MaterialCodeRuleVersion, batch.new_rule_version_id) if batch.new_rule_version_id else None
    if old_rule:
        old_rule.status = "active"
        old_rule.updated_at = now()
        library.current_rule_version_id = old_rule.id
    if new_rule and old_rule and new_rule.id != old_rule.id:
        new_rule.status = "deprecated"
        new_rule.updated_at = now()
    library.updated_at = now()
    batch.status = "rolled_back"
    batch.updated_at = now()
    add_audit_log(
        db,
        auth,
        "material_code_change_batch",
        "rollback",
        before,
        {
            "batch_id": batch.id,
            "status": batch.status,
            "reason": payload.reason,
            "codes": {str(detail.material_id): detail.old_code for detail in details},
        },
    )
    db.commit()
    db.refresh(batch)
    return code_change_batch_to_out(db, batch, include_rows=True)


@app.get("/api/v1/material-libraries/{library_id}/code-mappings", response_model=None)
def list_code_mappings(
    library_id: int,
    batch_id: int | None = None,
    old_code: str = "",
    new_code: str = "",
    export: str = "",
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=10, ge=1, le=100),
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/material-libraries/{library_id}")),
) -> Any:
    library = db.get(MaterialLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Material library not found")
    require_library_scope(auth, library.id)
    query = db.query(MaterialCodeMapping).filter(MaterialCodeMapping.library_id == library.id)
    if batch_id is not None:
        query = query.filter(MaterialCodeMapping.batch_id == batch_id)
    if old_code:
        query = query.filter(MaterialCodeMapping.old_code == old_code)
    if new_code:
        query = query.filter(MaterialCodeMapping.new_code == new_code)
    query = query.order_by(MaterialCodeMapping.id.desc())
    if export.strip().lower() == "csv":
        mappings = query.all()
        materials = {item.id: item for item in db.query(Material).filter(Material.id.in_([mapping.material_id for mapping in mappings] or [-1])).all()}
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(["id", "library_id", "material_id", "material_name", "old_code", "new_code", "old_rule_version_id", "new_rule_version_id", "batch_id", "status", "created_at"])
        for mapping in mappings:
            material = materials.get(mapping.material_id)
            writer.writerow(
                [
                    mapping.id,
                    mapping.library_id,
                    mapping.material_id,
                    material.name if material else "",
                    mapping.old_code,
                    mapping.new_code,
                    mapping.old_rule_version_id or "",
                    mapping.new_rule_version_id or "",
                    mapping.batch_id or "",
                    mapping.status,
                    mapping.created_at.isoformat(),
                ]
            )
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="material-code-mappings-{library.id}.csv"'},
        )
    total = query.count()
    mappings = query.offset((page - 1) * page_size).limit(page_size).all()
    materials = {item.id: item for item in db.query(Material).filter(Material.id.in_([mapping.material_id for mapping in mappings] or [-1])).all()}
    return MaterialCodeMappingListOut(
        items=[code_mapping_to_out(mapping, materials.get(mapping.material_id)) for mapping in mappings],
        total=total,
        page=page,
        page_size=page_size,
    )


@app.delete("/api/v1/material-libraries/{library_id}")
def delete_material_library(
    library_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.DELETE./api/v1/material-libraries/{library_id}")),
) -> dict[str, Any]:
    require_button_permission(auth, "button.material_library.delete")
    library = db.get(MaterialLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Material library not found")
    require_library_scope(auth, library.id)
    if db.query(Material).filter(Material.material_library_id == library.id).first():
        raise HTTPException(status_code=409, detail="Material library cannot be deleted while it contains materials")
    db.delete(library)
    db.commit()
    return {"deleted": True, "id": library_id}


@app.get("/api/v1/category-libraries", response_model=list[CategoryLibraryOut])
def list_category_libraries(
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/category-libraries")),
) -> list[CategoryLibraryOut]:
    ensure_seed_material_context(db)
    libraries = db.query(CategoryLibrary).order_by(CategoryLibrary.id).all()
    return [category_library_to_out(library) for library in libraries]


@app.post("/api/v1/category-libraries", response_model=CategoryLibraryOut)
def create_category_library(
    payload: CategoryLibraryIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/category-libraries")),
) -> CategoryLibraryOut:
    require_button_permission(auth, "button.category_library.create")
    name = compact_space(payload.name)
    if not name:
        raise HTTPException(status_code=422, detail="Category library name is required")
    code = compact_space(payload.code).upper() or next_unique_code(db, CategoryLibrary, "CLIB", f"{name}:{now().isoformat()}")
    if db.query(CategoryLibrary).filter(CategoryLibrary.name == name).first():
        raise HTTPException(status_code=409, detail="Category library name must be unique")
    if db.query(CategoryLibrary).filter(CategoryLibrary.code == code).first():
        raise HTTPException(status_code=409, detail="Category library code must be unique")
    library = CategoryLibrary(
        code=code,
        name=name,
        description=payload.description.strip(),
        enabled=payload.enabled,
        qdrant_enabled=payload.qdrant_enabled,
    )
    db.add(library)
    db.commit()
    db.refresh(library)
    if library.qdrant_enabled:
        try:
            create_qdrant_collection(library.id)
        except QdrantSyncError as exc:
            trace_qdrant_error(db, "qdrant.collection.create", str(exc), {"library_id": library.id})
    return category_library_to_out(library)


@app.get("/api/v1/category-libraries/{library_id}", response_model=CategoryLibraryOut)
def get_category_library(
    library_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/category-libraries/{library_id}")),
) -> CategoryLibraryOut:
    library = db.get(CategoryLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Category library not found")
    return category_library_to_out(library)


@app.put("/api/v1/category-libraries/{library_id}", response_model=CategoryLibraryOut)
def update_category_library(
    library_id: int,
    payload: CategoryLibraryUpdate,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/category-libraries/{library_id}")),
) -> CategoryLibraryOut:
    require_button_permission(auth, "button.category_library.edit")
    library = db.get(CategoryLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Category library not found")
    if payload.name is not None:
        name = compact_space(payload.name)
        if not name:
            raise HTTPException(status_code=422, detail="Category library name is required")
        duplicate = db.query(CategoryLibrary).filter(CategoryLibrary.name == name, CategoryLibrary.id != library.id).first()
        if duplicate:
            raise HTTPException(status_code=409, detail="Category library name must be unique")
        library.name = name
    if payload.code is not None:
        code = compact_space(payload.code).upper()
        if not code:
            raise HTTPException(status_code=422, detail="Category library code is required")
        duplicate = db.query(CategoryLibrary).filter(CategoryLibrary.code == code, CategoryLibrary.id != library.id).first()
        if duplicate:
            raise HTTPException(status_code=409, detail="Category library code must be unique")
        library.code = code
    if payload.description is not None:
        library.description = payload.description.strip()
    if payload.enabled is not None:
        library.enabled = payload.enabled
    qdrant_before = library.qdrant_enabled
    if payload.qdrant_enabled is not None:
        library.qdrant_enabled = payload.qdrant_enabled
    library.updated_at = now()
    db.commit()
    db.refresh(library)
    if payload.qdrant_enabled is not None and payload.qdrant_enabled != qdrant_before:
        if library.qdrant_enabled:
            reembed_category_library(db, library)
        else:
            try:
                delete_qdrant_collection(library.id)
            except QdrantSyncError as exc:
                trace_qdrant_error(db, "qdrant.collection.delete", str(exc), {"library_id": library.id})
    return category_library_to_out(library)


@app.delete("/api/v1/category-libraries/{library_id}")
def delete_category_library(
    library_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.DELETE./api/v1/category-libraries/{library_id}")),
) -> dict[str, Any]:
    require_button_permission(auth, "button.category_library.delete")
    library = db.get(CategoryLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Category library not found")
    if db.query(Category).filter(Category.category_library_id == library.id).first():
        raise HTTPException(status_code=409, detail="Category library cannot be deleted while it contains categories")
    qdrant_enabled = library.qdrant_enabled
    db.delete(library)
    db.commit()
    if qdrant_enabled:
        try:
            delete_qdrant_collection(library_id)
        except QdrantSyncError as exc:
            trace_qdrant_error(db, "qdrant.collection.delete", str(exc), {"library_id": library_id})
    return {"deleted": True, "id": library_id}


@app.post("/api/v1/category-libraries/{library_id}/re-embed")
def reembed_category_library_endpoint(
    library_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/category-libraries/{library_id}/re-embed")),
) -> dict[str, Any]:
    require_super_admin(auth)
    library = db.get(CategoryLibrary, library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Category library not found")
    if not library.qdrant_enabled:
        raise HTTPException(status_code=422, detail="Qdrant is not enabled for this category library")
    return reembed_category_library(db, library)


@app.get("/api/v1/categories", response_model=list[CategoryOut])
def list_categories(
    category_library_id: int | None = None,
    parent_id: int | None = None,
    level: int | None = Query(default=None, ge=1),
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    request: Request = None,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/categories")),
) -> list[CategoryOut]:
    ensure_seed_material_context(db)
    query = db.query(Category)
    if category_library_id is not None:
        query = query.filter(Category.category_library_id == category_library_id)
    if parent_id is not None:
        query = query.filter(Category.parent_category_id == parent_id)
    if level is not None:
        if level == 1:
            query = query.filter(Category.parent_category_id.is_(None))
        else:
            categories = query.order_by(Category.id).all()
            by_id = {category.id: category for category in categories}

            def depth(category: Category) -> int:
                current_depth = 1
                parent_id = category.parent_category_id
                while parent_id:
                    parent = by_id.get(parent_id)
                    if not parent:
                        break
                    current_depth += 1
                    parent_id = parent.parent_category_id
                return current_depth

            matching_categories = [category for category in categories if depth(category) == level]
            return [category_to_out(category) for category in matching_categories[offset : offset + limit]]
    is_bare_default = (
        request is not None
        and not request.query_params
        and category_library_id is None
        and parent_id is None
        and level is None
    )
    order_column = Category.id.desc() if is_bare_default else Category.id
    categories = query.order_by(order_column).offset(offset).limit(limit).all()
    return [category_to_out(category) for category in categories]


CATEGORY_IMPORT_HEADERS = ["一级类目", "二级类目", "三级类目", "四级类目", "五级类目"]


def normalize_category_import_row(raw: Any, row_number: int) -> dict[str, Any]:
    if not isinstance(raw, dict):
        return {"row_number": row_number, "levels": [], "errors": ["Row must be an object"]}
    levels = [compact_space(str(raw.get(header) or "")) for header in CATEGORY_IMPORT_HEADERS]
    errors: list[str] = []
    if not levels[0]:
        errors.append("一级类目 is required")
    for index in range(1, len(levels)):
        if levels[index] and not all(levels[previous] for previous in range(index)):
            errors.append(f"{CATEGORY_IMPORT_HEADERS[index - 1]} is required when {CATEGORY_IMPORT_HEADERS[index]} is provided")
    return {"row_number": row_number, "levels": levels, "errors": errors}


def parse_category_import_csv(text: str) -> list[dict[str, str]]:
    reader = csv.DictReader(StringIO(text))
    fieldnames = [name.strip() for name in (reader.fieldnames or [])]
    if CATEGORY_IMPORT_HEADERS[0] not in fieldnames:
        raise HTTPException(status_code=422, detail=f"Missing CSV headers: {CATEGORY_IMPORT_HEADERS[0]}")
    return [{header: row.get(header, "") for header in CATEGORY_IMPORT_HEADERS} for row in reader]


def xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def category_import_rows_from_table(table_rows: list[list[Any]], source: str) -> list[dict[str, str]]:
    non_empty_rows = [
        [compact_space(str(cell or "")) for cell in row]
        for row in table_rows
        if any(compact_space(str(cell or "")) for cell in row)
    ]
    if not non_empty_rows:
        raise HTTPException(status_code=422, detail=f"{source} category import file is empty")
    headers = [cell.strip() for cell in non_empty_rows[0]]
    if CATEGORY_IMPORT_HEADERS[0] not in headers:
        raise HTTPException(status_code=422, detail=f"Missing {source} headers: {CATEGORY_IMPORT_HEADERS[0]}")
    indexes = [headers.index(header) if header in headers else -1 for header in CATEGORY_IMPORT_HEADERS]
    rows: list[dict[str, str]] = []
    for raw_row in non_empty_rows[1:]:
        rows.append(
            {
                header: raw_row[index] if index >= 0 and index < len(raw_row) else ""
                for header, index in zip(CATEGORY_IMPORT_HEADERS, indexes)
            }
        )
    return rows


def xlsx_column_index(cell_ref: str, fallback: int) -> int:
    letters = re.match(r"([A-Za-z]+)", cell_ref or "")
    if not letters:
        return fallback
    value = 0
    for char in letters.group(1).upper():
        value = value * 26 + (ord(char) - ord("A") + 1)
    return max(0, value - 1)


def xlsx_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        data = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ElementTree.fromstring(data)
    values: list[str] = []
    for item in root:
        if xml_local_name(item.tag) != "si":
            continue
        values.append("".join(text.text or "" for text in item.iter() if xml_local_name(text.tag) == "t"))
    return values


def xlsx_first_sheet_path(archive: zipfile.ZipFile) -> str:
    try:
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    except KeyError as exc:
        raise ValueError("workbook.xml is missing") from exc
    relationship_targets: dict[str, str] = {}
    try:
        rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        for rel in rels:
            rel_id = rel.attrib.get("Id")
            target = rel.attrib.get("Target")
            if rel_id and target:
                relationship_targets[rel_id] = target
    except KeyError:
        relationship_targets = {}
    for element in workbook.iter():
        if xml_local_name(element.tag) != "sheet":
            continue
        rel_id = element.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = relationship_targets.get(rel_id or "")
        if target:
            return f"xl/{target.lstrip('/')}" if not target.startswith("xl/") else target
        sheet_id = compact_space(str(element.attrib.get("sheetId") or "1"))
        return f"xl/worksheets/sheet{sheet_id}.xml"
    return "xl/worksheets/sheet1.xml"


def parse_xlsx_category_import(content: bytes) -> list[dict[str, str]]:
    with zipfile.ZipFile(BytesIO(content)) as archive:
        shared_strings = xlsx_shared_strings(archive)
        sheet = ElementTree.fromstring(archive.read(xlsx_first_sheet_path(archive)))
    table_rows: list[list[str]] = []
    for row in [element for element in sheet.iter() if xml_local_name(element.tag) == "row"]:
        cells: dict[int, str] = {}
        fallback_index = 0
        for cell in [element for element in row if xml_local_name(element.tag) == "c"]:
            column_index = xlsx_column_index(cell.attrib.get("r", ""), fallback_index)
            fallback_index = column_index + 1
            cell_type = cell.attrib.get("t", "")
            if cell_type == "inlineStr":
                value = "".join(text.text or "" for text in cell.iter() if xml_local_name(text.tag) == "t")
            else:
                raw_value = next((child.text or "" for child in cell if xml_local_name(child.tag) == "v"), "")
                if cell_type == "s" and raw_value.strip().isdigit():
                    value = shared_strings[int(raw_value)] if int(raw_value) < len(shared_strings) else ""
                else:
                    value = raw_value
            cells[column_index] = compact_space(value)
        if cells:
            table_rows.append([cells.get(index, "") for index in range(max(cells) + 1)])
    return category_import_rows_from_table(table_rows, "XLSX")


def parse_xml_xls_category_import(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("gb18030")
    root = ElementTree.fromstring(text.lstrip())
    table_rows: list[list[str]] = []
    for row in [element for element in root.iter() if xml_local_name(element.tag).lower() in {"row", "tr"}]:
        values: list[str] = []
        for cell in [element for element in row if xml_local_name(element.tag).lower() in {"cell", "td", "th"}]:
            values.append(compact_space("".join(cell.itertext())))
        if values:
            table_rows.append(values)
    return category_import_rows_from_table(table_rows, "XLS")


def read_u16(data: bytes, offset: int) -> int:
    return int.from_bytes(data[offset : offset + 2], "little")


def read_u32(data: bytes, offset: int) -> int:
    return int.from_bytes(data[offset : offset + 4], "little")


def read_u64(data: bytes, offset: int) -> int:
    return int.from_bytes(data[offset : offset + 8], "little")


def cfb_sector(content: bytes, sector_size: int, sector_id: int) -> bytes:
    start = (sector_id + 1) * sector_size
    return content[start : start + sector_size]


def cfb_read_regular_chain(content: bytes, sector_size: int, fat: list[int], start_sector: int, size: int | None = None) -> bytes:
    end_of_chain = 0xFFFFFFFE
    chunks: list[bytes] = []
    sector_id = start_sector
    seen: set[int] = set()
    while 0 <= sector_id < len(fat) and sector_id not in seen and sector_id != end_of_chain:
        seen.add(sector_id)
        chunks.append(cfb_sector(content, sector_size, sector_id))
        next_sector = fat[sector_id]
        if next_sector >= 0xFFFFFFF8:
            break
        sector_id = next_sector
    data = b"".join(chunks)
    return data[:size] if size is not None else data


def cfb_workbook_stream(content: bytes) -> bytes:
    if content[:8] != b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        raise ValueError("not an OLE2 XLS workbook")
    sector_size = 1 << read_u16(content, 30)
    mini_sector_size = 1 << read_u16(content, 32)
    first_dir_sector = read_u32(content, 48)
    mini_stream_cutoff = read_u32(content, 56)
    first_mini_fat_sector = read_u32(content, 60)
    mini_fat_sector_count = read_u32(content, 64)
    difat = [read_u32(content, offset) for offset in range(76, 512, 4)]
    fat_sectors = [sector_id for sector_id in difat if sector_id < 0xFFFFFFF8]
    fat: list[int] = []
    for sector_id in fat_sectors:
        sector = cfb_sector(content, sector_size, sector_id)
        fat.extend(read_u32(sector, offset) for offset in range(0, len(sector), 4))
    directory = cfb_read_regular_chain(content, sector_size, fat, first_dir_sector)
    entries: dict[str, dict[str, Any]] = {}
    root_entry: dict[str, Any] | None = None
    for offset in range(0, len(directory), 128):
        entry = directory[offset : offset + 128]
        if len(entry) < 128:
            continue
        name_length = read_u16(entry, 64)
        if name_length < 2:
            continue
        name = entry[: name_length - 2].decode("utf-16le", errors="ignore")
        item = {
            "type": entry[66],
            "start": read_u32(entry, 116),
            "size": read_u64(entry, 120),
        }
        entries[name.lower()] = item
        if item["type"] == 5:
            root_entry = item
    workbook = entries.get("workbook") or entries.get("book")
    if not workbook:
        raise ValueError("Workbook stream is missing")
    workbook_size = int(workbook["size"])
    if workbook_size >= mini_stream_cutoff or not root_entry:
        return cfb_read_regular_chain(content, sector_size, fat, int(workbook["start"]), workbook_size)

    mini_fat_stream = cfb_read_regular_chain(
        content,
        sector_size,
        fat,
        first_mini_fat_sector,
        mini_fat_sector_count * sector_size,
    )
    mini_fat = [read_u32(mini_fat_stream, offset) for offset in range(0, len(mini_fat_stream), 4)]
    root_stream = cfb_read_regular_chain(content, sector_size, fat, int(root_entry["start"]), int(root_entry["size"]))
    chunks: list[bytes] = []
    mini_sector_id = int(workbook["start"])
    seen: set[int] = set()
    while 0 <= mini_sector_id < len(mini_fat) and mini_sector_id not in seen:
        seen.add(mini_sector_id)
        start = mini_sector_id * mini_sector_size
        chunks.append(root_stream[start : start + mini_sector_size])
        next_sector = mini_fat[mini_sector_id]
        if next_sector >= 0xFFFFFFF8:
            break
        mini_sector_id = next_sector
    return b"".join(chunks)[:workbook_size]


def parse_biff_unicode_string(data: bytes, offset: int = 0) -> tuple[str, int]:
    if offset + 3 > len(data):
        return "", len(data)
    char_count = read_u16(data, offset)
    flags = data[offset + 2]
    offset += 3
    rich_text_runs = 0
    phonetic_size = 0
    if flags & 0x08 and offset + 2 <= len(data):
        rich_text_runs = read_u16(data, offset)
        offset += 2
    if flags & 0x04 and offset + 4 <= len(data):
        phonetic_size = read_u32(data, offset)
        offset += 4
    is_utf16 = bool(flags & 0x01)
    byte_length = char_count * (2 if is_utf16 else 1)
    raw = data[offset : offset + byte_length]
    offset += byte_length
    text = raw.decode("utf-16le" if is_utf16 else "latin1", errors="ignore")
    offset += rich_text_runs * 4 + phonetic_size
    return text, offset


def parse_biff_xls_category_import(content: bytes) -> list[dict[str, str]]:
    stream = cfb_workbook_stream(content)
    shared_strings: list[str] = []
    cells: dict[tuple[int, int], str] = {}
    in_first_sheet = False
    first_sheet_seen = False
    offset = 0
    while offset + 4 <= len(stream):
        record_id = read_u16(stream, offset)
        record_size = read_u16(stream, offset + 2)
        record = stream[offset + 4 : offset + 4 + record_size]
        offset += 4 + record_size
        if record_id == 0x0809 and len(record) >= 4:
            bof_type = read_u16(record, 2)
            if bof_type == 0x0010 and not first_sheet_seen:
                in_first_sheet = True
                first_sheet_seen = True
            elif bof_type == 0x0010:
                in_first_sheet = False
        elif record_id == 0x000A and in_first_sheet:
            break
        elif record_id == 0x00FC and len(record) >= 8:
            unique_count = read_u32(record, 4)
            position = 8
            strings: list[str] = []
            for _ in range(unique_count):
                if position >= len(record):
                    break
                value, position = parse_biff_unicode_string(record, position)
                strings.append(value)
            shared_strings = strings
        elif in_first_sheet and record_id == 0x00FD and len(record) >= 10:
            row = read_u16(record, 0)
            column = read_u16(record, 2)
            string_index = read_u32(record, 6)
            cells[(row, column)] = shared_strings[string_index] if string_index < len(shared_strings) else ""
        elif in_first_sheet and record_id == 0x0204 and len(record) >= 9:
            row = read_u16(record, 0)
            column = read_u16(record, 2)
            value, _ = parse_biff_unicode_string(record, 6)
            cells[(row, column)] = value
        elif in_first_sheet and record_id == 0x0203 and len(record) >= 14:
            row = read_u16(record, 0)
            column = read_u16(record, 2)
            number = int.from_bytes(record[6:14], "little")
            cells[(row, column)] = str(number)
    if not cells:
        raise ValueError("no readable worksheet cells found")
    max_row = max(row for row, _column in cells)
    max_column = max(column for _row, column in cells)
    table_rows = [
        [compact_space(cells.get((row, column), "")) for column in range(max_column + 1)]
        for row in range(max_row + 1)
    ]
    return category_import_rows_from_table(table_rows, "XLS")


def parse_category_import_file(content: bytes, filename: str, content_type: str) -> list[dict[str, str]]:
    lowered = filename.lower()
    if lowered.endswith(".csv") or "text/csv" in content_type:
        try:
            return parse_category_import_csv(content.decode("utf-8-sig"))
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=422, detail=f"Unable to parse CSV category import file: {exc}") from exc
    if lowered.endswith(".xlsx") or "spreadsheetml.sheet" in content_type or zipfile.is_zipfile(BytesIO(content)):
        try:
            return parse_xlsx_category_import(content)
        except (KeyError, ValueError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
            raise HTTPException(status_code=422, detail=f"Unable to parse XLSX category import file: {exc}") from exc
    if lowered.endswith(".xls") or "application/vnd.ms-excel" in content_type:
        try:
            stripped = content.lstrip()
            if stripped.startswith(b"<"):
                return parse_xml_xls_category_import(content)
            return parse_biff_xls_category_import(content)
        except (UnicodeDecodeError, ValueError, ElementTree.ParseError) as exc:
            raise HTTPException(status_code=422, detail=f"Unable to parse XLS category import file: {exc}") from exc
    raise HTTPException(status_code=422, detail="Unsupported category import file type. Upload CSV, XLSX, or XLS.")


async def category_import_rows_from_request(request: Request) -> list[dict[str, str]]:
    content_type = request.headers.get("content-type", "")
    if "multipart/form-data" in content_type:
        form = await request.form()
        upload = form.get("file") or form.get("csv") or form.get("upload")
        if upload is None or not hasattr(upload, "read"):
            raise HTTPException(status_code=422, detail="Category import file is required")
        content = await upload.read()
        filename = compact_space(str(getattr(upload, "filename", "") or ""))
        upload_content_type = compact_space(str(getattr(upload, "content_type", "") or ""))
        return parse_category_import_file(content, filename, upload_content_type)

    if "text/csv" in content_type:
        content = await request.body()
        try:
            return parse_category_import_csv(content.decode("utf-8-sig"))
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=422, detail=f"Unable to parse CSV category import file: {exc}") from exc

    payload = await request.json()
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        rows = payload.get("rows") or payload.get("items") or payload.get("categories")
        if isinstance(rows, list):
            return rows
    raise HTTPException(status_code=422, detail="Expected JSON array or object with rows")


def find_category_by_parent(db: Session, library_id: int, parent_id: int | None, name: str) -> Category | None:
    query = db.query(Category).filter(
        Category.category_library_id == library_id,
        Category.name == name,
    )
    if parent_id is None:
        query = query.filter(Category.parent_category_id.is_(None))
    else:
        query = query.filter(Category.parent_category_id == parent_id)
    return query.first()


def create_import_category(db: Session, library_id: int, parent_id: int | None, name: str) -> Category:
    category = Category(
        code=next_unique_code(db, Category, "CAT", f"{library_id}:{parent_id or 0}:{name}:{now().isoformat()}"),
        name=name,
        category_library_id=library_id,
        parent_category_id=parent_id,
        description="",
        enabled=True,
    )
    db.add(category)
    db.flush()
    return category


def import_category_levels(db: Session, library_id: int, levels: list[str]) -> dict[str, Any]:
    parent_id: int | None = None
    path: list[str] = []
    created: list[Category] = []
    skipped: list[Category] = []
    for name in [level for level in levels if level]:
        path.append(name)
        existing = find_category_by_parent(db, library_id, parent_id, name)
        if existing:
            skipped.append(existing)
            parent_id = existing.id
            continue
        category = create_import_category(db, library_id, parent_id, name)
        created.append(category)
        parent_id = category.id
    return {"path": path, "created": created, "skipped": skipped}


def split_recognized_category_line(line: str) -> list[str]:
    cleaned = re.sub(r"^\s*[\-\*\d一二三四五六七八九十]+[\.、)]\s*", "", line.strip())
    cleaned = re.sub(r"\s*(?:>|/|\\|,|，|、|;|；|\t)\s*", "/", cleaned)
    parts = [compact_space(part) for part in cleaned.split("/") if compact_space(part)]
    if len(parts) == 1:
        parts = [compact_space(part) for part in re.split(r"\s+", parts[0]) if compact_space(part)]
    return parts[:5]


def confidence_for_category_levels(levels: list[str]) -> float:
    if len(levels) >= 3:
        return 0.92
    if len(levels) == 2:
        return 0.86
    return 0.78


def recognized_category_from_levels(levels: list[str]) -> dict[str, Any]:
    result: dict[str, Any] = {
        "level1": levels[0],
        "confidence": confidence_for_category_levels(levels),
    }
    for index, level in enumerate(levels[1:5], start=2):
        result[f"level{index}"] = level
    return result


def recognize_category_lines(text: str) -> list[dict[str, Any]]:
    categories: list[dict[str, Any]] = []
    seen: set[tuple[str, ...]] = set()
    for raw_line in text.splitlines():
        line = compact_space(raw_line)
        if not line:
            continue
        levels = split_recognized_category_line(line)
        if not levels:
            continue
        key = tuple(levels)
        if key in seen:
            continue
        seen.add(key)
        categories.append(recognized_category_from_levels(levels))
    if not categories:
        levels = split_recognized_category_line(text)
        if levels:
            categories.append(recognized_category_from_levels(levels))
    return categories


CATEGORY_RECOGNITION_JOBS: dict[str, dict[str, Any]] = {}
CATEGORY_RECOGNITION_CAPABILITY = "category_recognition"
CATEGORY_RECOGNITION_DEFAULT_MODEL = "qwen3.6-plus"


class CategoryRecognitionUpstreamError(Exception):
    def __init__(self, message: str, status_code: int = 502, retryable: bool = True):
        super().__init__(message)
        self.status_code = status_code
        self.retryable = retryable


def validate_category_library(db: Session, category_library_id: int | None) -> CategoryLibrary | None:
    if category_library_id is None:
        return None
    library = db.get(CategoryLibrary, category_library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Category library not found")
    return library


def category_path_for(category: Category, by_id: dict[int, Category]) -> list[str]:
    path = [category.name]
    parent_id = category.parent_category_id
    seen = {category.id}
    while parent_id and parent_id not in seen:
        parent = by_id.get(parent_id)
        if not parent:
            break
        path.append(parent.name)
        seen.add(parent.id)
        parent_id = parent.parent_category_id
    return list(reversed(path))[:5]


def category_hierarchy_paths(db: Session, library_id: int | None) -> list[list[str]]:
    query = db.query(Category).filter(Category.enabled.is_(True))
    if library_id is not None:
        query = query.filter(Category.category_library_id == library_id)
    categories = query.order_by(Category.parent_category_id.isnot(None), Category.parent_category_id, Category.id).all()
    by_id = {category.id: category for category in categories}
    paths: list[list[str]] = []
    seen: set[tuple[str, ...]] = set()
    for category in categories:
        path = category_path_for(category, by_id)
        key = tuple(path)
        if key and key not in seen:
            seen.add(key)
            paths.append(path)
    return paths


def category_recognition_messages(text: str, hierarchy_paths: list[list[str]]) -> list[dict[str, str]]:
    hierarchy = "\n".join(f"- {' / '.join(path)}" for path in hierarchy_paths) or "- No configured categories"
    system_prompt = (
        "You are a category recognition agent for an enterprise material master data system. "
        "Use the provided category hierarchy context and return structured JSON output only. "
        "The JSON schema is {\"categories\":[{\"level1\":\"...\",\"level2\":\"...\","
        "\"level3\":\"...\",\"level4\":\"...\",\"level5\":\"...\",\"confidence\":0.0}],\"suggestions\":[\"...\"]}. "
        "Return multiple category candidates when the input is ambiguous. Confidence must be between 0.0 and 1.0.\n\n"
        f"Category hierarchy:\n{hierarchy}\n\n"
        "Examples:\n"
        "- HP LaserJet printer -> 办公设备 / 打印机 / 激光打印机\n"
        "- 防割手套 -> 安全用品 / 防护用品 / 防护手套"
    )
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": f"Recognize the category path for this material description:\n{text}"},
    ]


def clamp_confidence(value: Any) -> float:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        numeric = 0.5
    return round(max(0.0, min(1.0, numeric)), 4)


def extract_json_payload(content: str) -> dict[str, Any]:
    text = content.strip()
    candidates = [text]
    candidates.extend(re.findall(r"```(?:json)?\s*(\{.*?\})\s*```", text, flags=re.IGNORECASE | re.DOTALL))
    if "{" in text and "}" in text:
        candidates.append(text[text.find("{"): text.rfind("}") + 1])
    for candidate in candidates:
        try:
            parsed = json.loads(candidate)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            continue
    raise ValueError("Provider response did not contain a JSON object")


def normalized_category_candidate(raw: Any) -> dict[str, Any] | None:
    if not isinstance(raw, dict):
        return None
    levels: list[str] = []
    if isinstance(raw.get("levels"), list):
        levels = [compact_space(str(item)) for item in raw["levels"] if compact_space(str(item))]
    elif raw.get("path"):
        levels = split_recognized_category_line(str(raw["path"]))
    else:
        levels = [compact_space(str(raw.get(f"level{index}") or "")) for index in range(1, 6)]
        levels = [level for level in levels if level]
    if not levels:
        return None
    candidate: dict[str, Any] = {"level1": levels[0], "confidence": clamp_confidence(raw.get("confidence"))}
    for index, level in enumerate(levels[1:5], start=2):
        candidate[f"level{index}"] = level
    return candidate


def parse_category_recognition_response(content: str) -> dict[str, Any]:
    parsed = extract_json_payload(content)
    raw_categories = parsed.get("categories") or parsed.get("candidates") or parsed.get("results") or []
    categories = [candidate for candidate in (normalized_category_candidate(item) for item in raw_categories) if candidate]
    categories.sort(key=lambda item: item["confidence"], reverse=True)
    suggestions = [compact_space(str(item)) for item in parsed.get("suggestions", []) if compact_space(str(item))]
    if not suggestions:
        suggestions = ["请人工复核候选类目与物料描述是否匹配"]
    if not categories:
        raise ValueError("Provider JSON did not include category candidates")
    return {"categories": categories, "suggestions": suggestions}


def local_category_recognition_response(text: str) -> dict[str, Any]:
    categories = recognize_category_lines(text)
    if not categories:
        raise HTTPException(status_code=422, detail="No recognizable category paths found")
    return {
        "categories": categories,
        "suggestions": [
            "Review recognized names before confirming import",
            "Use one line per category path for best results",
        ],
    }


def mark_root_trace_model(
    collector: SpanCollector,
    provider: Model | ModelConfig | AIAgentConfig,
    model_name: str,
    metadata: dict[str, Any] | None = None,
) -> None:
    for span in collector.spans:
        if span["span_id"] == collector.root_span_id:
            span["provider"] = provider.provider
            span["model"] = model_name
            span["metadata"].update({"provider": provider.provider, "model": model_name, **(metadata or {})})
            return


def call_category_recognition_provider(
    db: Session,
    request: CategoryRecognitionRequest,
    hierarchy_paths: list[list[str]],
) -> dict[str, Any]:
    resolution = model_for_capability(db, CATEGORY_RECOGNITION_CAPABILITY)
    provider = resolution.model
    model_name = compact_space(request.model_override or provider.model_name or CATEGORY_RECOGNITION_DEFAULT_MODEL)
    collector = SpanCollector("category_recognition.recognize", CATEGORY_RECOGNITION_CAPABILITY)
    mark_root_trace_model(collector, provider, model_name, resolution_trace_metadata(resolution))
    try:
        if provider.provider == "mock" or (provider.base_url or "").startswith("local://"):
            result = local_category_recognition_response(request.text)
            collector.finish_span(
                collector.root_span_id,
                "ok",
                metadata={"mode": "local", "model": model_name, **resolution_trace_metadata(resolution)},
            )
            return {
                **result,
                "trace_id": collector.trace_id,
                "provider": provider.provider,
                "model": model_name,
                "resolution_source": "capability_mapping",
            }

        url = model_chat_url(provider)
        if not url:
            raise CategoryRecognitionUpstreamError("Model base URL is not configured", 502, True)
        headers = {"Content-Type": "application/json"}
        api_key = decrypt_api_key(provider.encrypted_api_key)
        if api_key:
            headers["Authorization"] = f"Bearer {api_key}"
        body = {
            "model": model_name,
            "messages": category_recognition_messages(request.text, hierarchy_paths),
            "temperature": getattr(provider, "temperature", 0),
            "response_format": {"type": "json_object"},
        }
        max_tokens = getattr(provider, "max_tokens", None)
        if max_tokens:
            body["max_tokens"] = int(max_tokens)
        timeout_value = getattr(provider, "timeout", getattr(provider, "timeout_seconds", 30))
        timeout_seconds = min(120, max(1, int(timeout_value or 30)))
        last_error = ""
        for attempt in range(2):
            span_id = collector.start_span(
                "category_recognition.llm.chat",
                "llm",
                capability=CATEGORY_RECOGNITION_CAPABILITY,
                parent_span_id=collector.root_span_id,
                provider=provider.provider,
                model=model_name,
                metadata={"attempt": attempt + 1, "url": url, **resolution_trace_metadata(resolution)},
            )
            try:
                response = httpx.post(url, json=body, headers=headers, timeout=timeout_seconds)
            except httpx.TimeoutException as exc:
                last_error = f"Provider timed out after {timeout_seconds}s"
                collector.finish_span(span_id, "error", str(exc))
                raise CategoryRecognitionUpstreamError(last_error, 504, True) from exc
            except httpx.RequestError as exc:
                last_error = f"Provider request failed: {exc}"
                collector.finish_span(span_id, "error", str(exc))
                raise CategoryRecognitionUpstreamError(last_error, 503, True) from exc
            if response.status_code >= 500:
                last_error = f"Provider returned HTTP {response.status_code}"
                collector.finish_span(span_id, "error", last_error)
                if attempt == 0:
                    continue
                raise CategoryRecognitionUpstreamError(last_error, 502, True)
            if response.status_code >= 400:
                last_error = f"Provider rejected request with HTTP {response.status_code}"
                collector.finish_span(span_id, "error", last_error)
                raise CategoryRecognitionUpstreamError(last_error, 502, False)
            body_json = response.json()
            choices = body_json.get("choices") if isinstance(body_json, dict) else None
            message = choices[0].get("message", {}) if choices and isinstance(choices[0], dict) else {}
            content = str(message.get("content") or choices[0].get("text") or "")
            try:
                result = parse_category_recognition_response(content)
            except Exception as exc:
                collector.finish_span(span_id, "error", str(exc))
                raise
            collector.finish_span(span_id, "ok", metadata={"status_code": response.status_code})
            collector.finish_span(
                collector.root_span_id,
                "ok",
                metadata={"model": model_name, "attempts": attempt + 1, **resolution_trace_metadata(resolution)},
            )
            return {
                **result,
                "trace_id": collector.trace_id,
                "provider": provider.provider,
                "model": model_name,
                "resolution_source": "capability_mapping",
            }
        raise CategoryRecognitionUpstreamError(last_error or "Provider request failed", 502, True)
    except CapabilityResolutionError:
        raise
    except CategoryRecognitionUpstreamError as exc:
        try:
            result = local_category_recognition_response(request.text)
        except HTTPException:
            collector.finish_span(collector.root_span_id, "error", str(exc), {"model": model_name, "retryable": exc.retryable})
            raise
        collector.finish_span(
            collector.root_span_id,
            "ok",
            metadata={"model": model_name, "retryable": exc.retryable, "local_fallback_reason": str(exc), **resolution_trace_metadata(resolution)},
        )
        return {
            **result,
            "trace_id": collector.trace_id,
            "provider": provider.provider,
            "model": model_name,
            "resolution_source": "capability_mapping",
        }
    except HTTPException:
        collector.finish_span(collector.root_span_id, "error", "local recognition failed", {"model": model_name})
        raise
    except Exception as exc:
        collector.finish_span(collector.root_span_id, "error", str(exc), {"model": model_name})
        raise CategoryRecognitionUpstreamError(str(exc), 502, True) from exc
    finally:
        collector.flush(db)


def run_category_recognition(db: Session, payload: CategoryRecognitionRequest) -> dict[str, Any]:
    text = str(payload.text or "").strip()
    if not text:
        raise HTTPException(status_code=422, detail="Recognition text is required")
    payload.text = text
    validate_category_library(db, payload.category_library_id)
    hierarchy_paths = category_hierarchy_paths(db, payload.category_library_id)
    try:
        return call_category_recognition_provider(db, payload, hierarchy_paths)
    except CapabilityResolutionError as exc:
        raise capability_resolution_http_error(exc) from exc
    except CategoryRecognitionUpstreamError as exc:
        raise HTTPException(
            status_code=exc.status_code,
            detail={
                "error": str(exc),
                "suggestions": ["请稍后重试或切换可用的类目识别模型", "可手工选择类目后继续业务流程"],
                "retryable": exc.retryable,
            },
        ) from exc


def batch_item_text(item: Any) -> str:
    if isinstance(item, str):
        return compact_space(item)
    if isinstance(item, dict):
        return compact_space(str(item.get("text") or item.get("name") or item.get("description") or ""))
    return compact_space(str(item))


@app.post("/api/v1/ai/category-recognition/recognize", response_model=CategoryRecognitionResponse, response_model_exclude_none=True)
def recognize_categories(
    payload: CategoryRecognitionRequest,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/ai/category-recognition/recognize")),
) -> CategoryRecognitionResponse:
    require_super_admin(auth)
    return CategoryRecognitionResponse(**run_category_recognition(db, payload))


@app.post("/api/v1/ai/category-recognition/recognize-async", response_model=CategoryRecognitionJob)
def recognize_categories_async(
    payload: CategoryRecognitionRequest,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/ai/category-recognition/recognize-async")),
) -> CategoryRecognitionJob:
    require_super_admin(auth)
    job_id = f"catrec-{uuid.uuid4().hex[:20]}"
    CATEGORY_RECOGNITION_JOBS[job_id] = {
        "job_id": job_id,
        "status": "running",
        "text": payload.text,
        "category_library_id": payload.category_library_id,
        "result": None,
        "error": "",
    }
    try:
        result = run_category_recognition(db, payload)
        job_result = CategoryRecognitionJobResult(
            text=payload.text,
            category_library_id=payload.category_library_id,
            categories=result["categories"],
            suggestions=result["suggestions"],
        )
        CATEGORY_RECOGNITION_JOBS[job_id].update({"status": "succeeded", "result": job_result, "error": ""})
    except HTTPException as exc:
        CATEGORY_RECOGNITION_JOBS[job_id].update({"status": "failed", "error": str(exc.detail)})
    return CategoryRecognitionJob(**CATEGORY_RECOGNITION_JOBS[job_id])


@app.get("/api/v1/ai/category-recognition/jobs/{job_id}", response_model=CategoryRecognitionJob)
def get_category_recognition_job(
    job_id: str,
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/ai/category-recognition/jobs/{job_id}")),
) -> CategoryRecognitionJob:
    require_super_admin(auth)
    job = CATEGORY_RECOGNITION_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Category recognition job not found")
    return CategoryRecognitionJob(**job)


@app.post("/api/v1/ai/category-recognition/batch")
def recognize_categories_batch(
    payload: CategoryRecognitionBatchRequest,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/ai/category-recognition/batch")),
) -> dict[str, Any]:
    require_super_admin(auth)
    job_id = f"catrec-batch-{uuid.uuid4().hex[:20]}"
    results: list[dict[str, Any]] = []
    for item in payload.items:
        text = batch_item_text(item)
        if not text:
            raise HTTPException(status_code=422, detail="Batch item text is required")
        recognition = run_category_recognition(
            db,
            CategoryRecognitionRequest(
                text=text,
                category_library_id=payload.category_library_id,
                model_override=payload.model_override,
            ),
        )
        results.append({"text": text, **recognition})
    CATEGORY_RECOGNITION_JOBS[job_id] = {
        "job_id": job_id,
        "status": "succeeded",
        "text": None,
        "category_library_id": payload.category_library_id,
        "result": None,
        "error": "",
        "results": results,
    }
    return {"job_id": job_id, "status": "succeeded", "results": results}


@app.get("/api/v1/categories/template")
def download_category_template(
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/categories/template")),
) -> StreamingResponse:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(CATEGORY_IMPORT_HEADERS)
    writer.writerow(["办公设备", "", "", "", ""])
    writer.writerow(["办公设备", "打印设备", "", "", ""])
    writer.writerow(["办公设备", "打印设备", "激光打印机", "", ""])
    writer.writerow(["办公设备", "打印设备", "激光打印机", "A4纸", "80g"])
    content = output.getvalue()
    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="category-template.csv"'},
    )


@app.post("/api/v1/categories/bulk-import")
async def bulk_import_categories(
    request: Request,
    category_library_id: int = Query(...),
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/categories/bulk-import")),
) -> dict[str, Any]:
    require_super_admin(auth)
    library = db.get(CategoryLibrary, category_library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Category library not found")

    rows = await category_import_rows_from_request(request)
    normalized_rows = [normalize_category_import_row(row, index + 1) for index, row in enumerate(rows)]
    errors = [
        {"row_number": row["row_number"], "errors": row["errors"]}
        for row in normalized_rows
        if row["errors"]
    ]
    success_details: list[dict[str, Any]] = []
    skipped_details: list[dict[str, Any]] = []

    for row in [item for item in normalized_rows if not item["errors"]]:
        result = import_category_levels(db, library.id, row["levels"])
        created = result["created"]
        skipped = result["skipped"]
        path = " / ".join(result["path"])
        if created:
            success_details.extend(
                {
                    "row_number": row["row_number"],
                    "id": category.id,
                    "name": category.name,
                    "parent_category_id": category.parent_category_id,
                    "path": path,
                }
                for category in created
            )
        elif skipped:
            skipped_details.append(
                {
                    "row_number": row["row_number"],
                    "reason": "duplicate",
                    "path": path,
                }
            )
        else:
            skipped_details.append(
                {
                    "row_number": row["row_number"],
                    "reason": "empty",
                    "path": path,
                }
            )

    db.commit()
    if library.qdrant_enabled and success_details:
        reembed_category_library(db, library)
    return {
        "category_library_id": library.id,
        "success_count": len(success_details),
        "skipped_count": len(skipped_details),
        "error_count": len(errors),
        "success": success_details,
        "skipped": skipped_details,
        "errors": errors,
    }


@app.post("/api/v1/categories", response_model=CategoryOut)
def create_category(
    payload: CategoryIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/categories")),
) -> CategoryOut:
    require_button_permission(auth, "button.category_management.create")
    library = db.get(CategoryLibrary, payload.category_library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Category library not found")
    parent = db.get(Category, payload.parent_category_id) if payload.parent_category_id else None
    if payload.parent_category_id and not parent:
        raise HTTPException(status_code=404, detail="Parent category not found")
    if parent and parent.category_library_id != library.id:
        raise HTTPException(status_code=422, detail="Parent category must be in the same library")
    name = compact_space(payload.name)
    if not name:
        raise HTTPException(status_code=422, detail="Category name is required")
    code = compact_space(payload.code).upper() or next_unique_code(db, Category, "CAT", f"{name}:{now().isoformat()}")
    duplicate_name = find_category_by_parent(db, library.id, parent.id if parent else None, name)
    if duplicate_name:
        raise HTTPException(status_code=409, detail="Category name must be unique within the same parent")
    if db.query(Category).filter(Category.code == code).first():
        raise HTTPException(status_code=409, detail="Category code must be unique")
    category = Category(
        code=code,
        name=name,
        category_library_id=library.id,
        parent_category_id=parent.id if parent else None,
        description=payload.description.strip(),
        enabled=payload.enabled,
    )
    db.add(category)
    db.commit()
    db.refresh(category)
    try:
        qdrant_sync_category(db, category, "create")
    except QdrantSyncError as exc:
        raise qdrant_http_exception("category create sync", exc) from exc
    return category_to_out(category)


@app.get("/api/v1/categories/{category_id}/attributes", response_model=list[CategoryAttributeRead])
def list_category_attributes(
    category_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/categories/{category_id}/attributes")),
) -> list[CategoryAttributeRead]:
    del auth
    return compute_category_properties(db, category_id)


@app.get("/api/v1/categories/{category_id}/attributes/own", response_model=list[CategoryAttributeRead])
def list_own_category_attributes(
    category_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/categories/{category_id}/attributes/own")),
) -> list[CategoryAttributeRead]:
    del auth
    ensure_category_attribute_schema()
    category = db.get(Category, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    attributes = (
        db.query(CategoryAttribute)
        .filter(CategoryAttribute.category_id == category.id)
        .order_by(CategoryAttribute.sort_order, CategoryAttribute.id)
        .all()
    )
    return [category_attribute_to_read(attribute, category.id, category) for attribute in attributes]


@app.get("/api/v1/categories/{category_id}/properties", response_model=CategoryPropertyList)
def get_category_properties(
    category_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/categories/{category_id}/properties")),
) -> CategoryPropertyList:
    del auth
    return category_property_list(db, category_id)


@app.post("/api/v1/categories/{category_id}/attributes", response_model=CategoryAttributeRead)
def create_category_attribute(
    category_id: int,
    payload: CategoryAttributeCreate,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/categories/{category_id}/attributes")),
) -> CategoryAttributeRead:
    require_button_permission(auth, "button.category_management.edit")
    ensure_category_attribute_schema()
    category = db.get(Category, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    values = validate_category_attribute_payload(db, category, payload)
    attribute = CategoryAttribute(category_id=category.id, **values)
    db.add(attribute)
    db.flush()
    add_audit_log(
        db,
        auth,
        "category_attribute",
        "create",
        {},
        {"id": attribute.id, "category_id": category.id, **values},
        "human",
    )
    db.commit()
    db.refresh(attribute)
    return category_attribute_to_read(attribute, category.id, category)


@app.post("/api/v1/categories/{category_id}/attributes/batch", response_model=list[CategoryAttributeRead])
def create_category_attributes_batch(
    category_id: int,
    payload: list[CategoryAttributeCreate],
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/categories/{category_id}/attributes/batch")),
) -> list[CategoryAttributeRead]:
    require_button_permission(auth, "button.category_management.edit")
    ensure_category_attribute_schema()
    category = db.get(Category, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    if not payload:
        raise HTTPException(status_code=422, detail="At least one attribute is required")
    created: list[CategoryAttribute] = []
    for item in payload:
        values = validate_category_attribute_payload(db, category, item)
        attribute = CategoryAttribute(category_id=category.id, **values)
        db.add(attribute)
        db.flush()
        created.append(attribute)
    add_audit_log(
        db,
        auth,
        "category_attribute",
        "batch_create",
        {},
        {"category_id": category.id, "attribute_ids": [attribute.id for attribute in created]},
        "human",
    )
    db.commit()
    for attribute in created:
        db.refresh(attribute)
    return [category_attribute_to_read(attribute, category.id, category) for attribute in created]


@app.put("/api/v1/categories/{category_id}/attributes/{attribute_id}", response_model=CategoryAttributeRead)
def update_category_attribute(
    category_id: int,
    attribute_id: int,
    payload: CategoryAttributeUpdate,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/categories/{category_id}/attributes/{attribute_id}")),
) -> CategoryAttributeRead:
    require_button_permission(auth, "button.category_management.edit")
    ensure_category_attribute_schema()
    category = db.get(Category, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    attribute = db.get(CategoryAttribute, attribute_id)
    if not attribute or attribute.category_id != category.id:
        raise HTTPException(status_code=404, detail="Own category attribute not found")
    before = category_attribute_to_read(attribute, category.id, category).model_dump()
    values = validate_category_attribute_payload(db, category, payload, existing=attribute)
    if "attr_type" not in values and "options" in values and attribute.attr_type == "enum" and not category_attribute_options(values["options"]):
        raise HTTPException(status_code=422, detail="Enum attributes require at least one option")
    for field, value in values.items():
        setattr(attribute, field, value)
    attribute.updated_at = now()
    db.flush()
    after = category_attribute_to_read(attribute, category.id, category).model_dump()
    add_audit_log(db, auth, "category_attribute", "update", before, after, "human")
    db.commit()
    db.refresh(attribute)
    return category_attribute_to_read(attribute, category.id, category)


@app.delete("/api/v1/categories/{category_id}/attributes/{attribute_id}")
def delete_category_attribute(
    category_id: int,
    attribute_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.DELETE./api/v1/categories/{category_id}/attributes/{attribute_id}")),
) -> dict[str, Any]:
    require_button_permission(auth, "button.category_management.edit")
    ensure_category_attribute_schema()
    category = db.get(Category, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    attribute = db.get(CategoryAttribute, attribute_id)
    if not attribute or attribute.category_id != category.id:
        raise HTTPException(status_code=404, detail="Own category attribute not found")
    before = category_attribute_to_read(attribute, category.id, category).model_dump()
    db.delete(attribute)
    add_audit_log(db, auth, "category_attribute", "delete", before, {"deleted": True, "id": attribute_id}, "human")
    db.commit()
    return {"deleted": True, "id": attribute_id}


@app.put("/api/v1/categories/{category_id}", response_model=CategoryOut)
def update_category(
    category_id: int,
    payload: CategoryUpdate,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/categories/{category_id}")),
) -> CategoryOut:
    require_button_permission(auth, "button.category_management.edit")
    category = db.get(Category, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    if payload.category_library_id is not None:
        library = db.get(CategoryLibrary, payload.category_library_id)
        if not library:
            raise HTTPException(status_code=404, detail="Category library not found")
        category.category_library_id = library.id
    if payload.parent_category_id is not None:
        if payload.parent_category_id == category.id:
            raise HTTPException(status_code=422, detail="Category cannot be its own parent")
        parent = db.get(Category, payload.parent_category_id)
        if not parent:
            raise HTTPException(status_code=404, detail="Parent category not found")
        if parent.category_library_id != category.category_library_id:
            raise HTTPException(status_code=422, detail="Parent category must be in the same library")
        category.parent_category_id = parent.id
    if payload.name is not None:
        name = compact_space(payload.name)
        if not name:
            raise HTTPException(status_code=422, detail="Category name is required")
        duplicate = (
            db.query(Category)
            .filter(
                Category.category_library_id == category.category_library_id,
                Category.parent_category_id == category.parent_category_id,
                Category.name == name,
                Category.id != category.id,
            )
            .first()
        )
        if duplicate:
            raise HTTPException(status_code=409, detail="Category name must be unique within the same parent")
        category.name = name
    if payload.code is not None:
        code = compact_space(payload.code).upper()
        if not code:
            raise HTTPException(status_code=422, detail="Category code is required")
        duplicate = db.query(Category).filter(Category.code == code, Category.id != category.id).first()
        if duplicate:
            raise HTTPException(status_code=409, detail="Category code must be unique")
        category.code = code
    if payload.description is not None:
        category.description = payload.description.strip()
    if payload.enabled is not None:
        category.enabled = payload.enabled
    category.updated_at = now()
    db.commit()
    db.refresh(category)
    try:
        qdrant_sync_category_subtree(db, category, "update")
    except QdrantSyncError as exc:
        raise qdrant_http_exception("category update sync", exc) from exc
    return category_to_out(category)


@app.delete("/api/v1/categories/{category_id}")
def delete_category(
    category_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.DELETE./api/v1/categories/{category_id}")),
) -> dict[str, Any]:
    require_button_permission(auth, "button.category_management.delete")
    category = db.get(Category, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    if db.query(Material).filter(Material.category_id == category.id).first():
        raise HTTPException(status_code=409, detail="Category cannot be deleted while it contains materials")
    if db.query(Category).filter(Category.parent_category_id == category.id).first():
        raise HTTPException(status_code=409, detail="Category cannot be deleted while it contains child categories")
    library = enabled_library_for_category(db, category)
    deleted_category_id = category.id
    deleted_library_id = library.id if library else None
    db.delete(category)
    db.commit()
    if deleted_library_id is not None:
        try:
            delete_category_point(deleted_library_id, deleted_category_id)
        except QdrantSyncError as exc:
            trace_qdrant_error(db, "qdrant.category.delete", str(exc), {"category_id": deleted_category_id, "library_id": deleted_library_id})
            raise qdrant_http_exception("category delete sync", exc) from exc
    return {"deleted": True, "id": category_id}


@app.get("/api/v1/system/config", response_model=SystemConfigOut)
def get_system_config(
    request: Request,
    db: Session = Depends(get_db),
) -> SystemConfigOut:
    current_auth(request, db)
    return config_to_out(ensure_system_config(db))


@app.put("/api/v1/system/config", response_model=SystemConfigOut)
def update_system_config(
    payload: SystemConfigIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/system/config")),
) -> SystemConfigOut:
    config = ensure_system_config(db)
    before = system_config_payload(config)
    after = dict(before)
    if payload.system_name is not None:
        system_name = compact_space(payload.system_name)
        if not system_name:
            raise HTTPException(status_code=422, detail="system_name is required")
        after["system_name"] = system_name
    if payload.icon is not None:
        after["icon"] = sanitize_icon(payload.icon)
    if payload.stop_purchase_reasons is not None:
        after["stop_purchase_reasons"] = normalize_reason_options(payload.stop_purchase_reasons)
    if payload.stop_use_reasons is not None:
        after["stop_use_reasons"] = normalize_reason_options(payload.stop_use_reasons)
    if payload.approval_mode is not None:
        mode = payload.approval_mode.strip()
        if mode not in APPROVAL_MODES:
            raise HTTPException(status_code=422, detail="approval_mode must be simple or multi_node")
        after["approval_mode"] = mode
    config.value = json.dumps(after, ensure_ascii=False)
    config.updated_by = auth.username
    config.updated_at = now()
    add_audit_log(db, auth, "system_config", "update", before, after)
    db.commit()
    db.refresh(config)
    return config_to_out(config)


@app.post("/api/v1/system/config", response_model=SystemConfigOut)
def save_system_config(
    payload: SystemConfigIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/system/config")),
) -> SystemConfigOut:
    return update_system_config(payload, db, auth)


@app.get("/api/v1/observability/slow-queries", response_model=list[SlowQueryLogOut])
def list_slow_queries(
    limit: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
) -> list[SlowQueryLogOut]:
    ensure_slow_query_schema()
    logs = db.query(SlowQueryLog).order_by(SlowQueryLog.id.desc()).limit(limit).all()
    return [slow_query_to_out(log) for log in logs]


@app.post("/api/v1/telemetry/web-vitals", response_model=WebVitalsTelemetryOut, status_code=201)
def create_web_vitals_telemetry(
    payload: WebVitalsTelemetryIn,
    db: Session = Depends(get_db),
) -> WebVitalsTelemetryOut:
    ensure_web_vitals_schema()
    record = TelemetryWebVital(
        metric=payload.metric,
        value=float(payload.value),
        rating=payload.rating,
        client_metric_id=payload.client_metric_id,
        navigation_type=payload.navigation_type,
        url=payload.url,
        path=payload.path,
        user_agent=payload.user_agent,
        timestamp=payload.timestamp,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return web_vitals_to_out(record)


@app.get("/api/v1/telemetry/web-vitals", response_model=list[WebVitalsTelemetryOut])
def list_web_vitals_telemetry(
    client_metric_id: str = Query(..., min_length=1),
    db: Session = Depends(get_db),
) -> list[WebVitalsTelemetryOut]:
    ensure_web_vitals_schema()
    records = (
        db.query(TelemetryWebVital)
        .filter(TelemetryWebVital.client_metric_id == client_metric_id)
        .order_by(TelemetryWebVital.id.asc())
        .all()
    )
    return [web_vitals_to_out(record) for record in records]


def audit_query(
    db: Session,
    user: str = "",
    resource: str = "",
    action: str = "",
    source: str = "",
    start_time: str = "",
    end_time: str = "",
):
    query = db.query(AuditLog)
    if user:
        query = query.filter(AuditLog.user.like(f"%{user}%"))
    if resource:
        query = query.filter(AuditLog.resource == resource)
    if action:
        query = query.filter(AuditLog.action == action)
    if source:
        query = query.filter(AuditLog.source == source)
    if start_time:
        try:
            query = query.filter(AuditLog.timestamp >= datetime.fromisoformat(start_time.replace("Z", "+00:00")))
        except ValueError as exc:
            raise HTTPException(status_code=422, detail="start_time must be ISO-8601") from exc
    if end_time:
        try:
            query = query.filter(AuditLog.timestamp <= datetime.fromisoformat(end_time.replace("Z", "+00:00")))
        except ValueError as exc:
            raise HTTPException(status_code=422, detail="end_time must be ISO-8601") from exc
    return query


def xlsx_cell(value: Any) -> str:
    text = xml_escape(str(value if value is not None else ""))
    return f'<c t="inlineStr"><is><t>{text}</t></is></c>'


def build_audit_workbook(rows: list[list[Any]]) -> BytesIO:
    sheet_rows = "\n".join(
        f'<row r="{row_index}">' + "".join(xlsx_cell(value) for value in row) + "</row>"
        for row_index, row in enumerate(rows, start=1)
    )
    output = BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "[Content_Types].xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
</Types>""",
        )
        archive.writestr(
            "_rels/.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>""",
        )
        archive.writestr(
            "xl/workbook.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="Audit Logs" sheetId="1" r:id="rId1"/></sheets>
</workbook>""",
        )
        archive.writestr(
            "xl/_rels/workbook.xml.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>""",
        )
        archive.writestr(
            "xl/worksheets/sheet1.xml",
            f"""<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData>{sheet_rows}</sheetData>
</worksheet>""",
        )
    output.seek(0)
    return output


@app.get("/api/v1/audit-logs", response_model=AuditLogListOut)
def list_audit_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user: str = "",
    resource: str = "",
    action: str = "",
    source: str = "",
    start_time: str = "",
    end_time: str = "",
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/audit-logs")),
) -> AuditLogListOut:
    query = audit_query(db, user, resource, action, source, start_time, end_time)
    total = query.count()
    logs = (
        query.order_by(AuditLog.timestamp.desc(), AuditLog.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    pages = max(1, (total + page_size - 1) // page_size)
    return AuditLogListOut(items=[audit_to_out(log) for log in logs], total=total, page=page, page_size=page_size, pages=pages)


@app.get("/api/v1/audit-logs/export")
def export_audit_logs(
    user: str = "",
    resource: str = "",
    action: str = "",
    source: str = "",
    start_time: str = "",
    end_time: str = "",
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/audit-logs/export")),
) -> StreamingResponse:
    logs = audit_query(db, user, resource, action, source, start_time, end_time).order_by(AuditLog.timestamp.desc(), AuditLog.id.desc()).all()
    headers = ["timestamp", "user", "resource", "action", "source", "before value", "after value"]
    rows: list[list[Any]] = [headers]
    for log in logs:
        item = audit_to_out(log)
        rows.append(
            [
                item.timestamp,
                item.user,
                item.resource,
                item.action,
                item.source,
                json.dumps(item.before_value, ensure_ascii=False, sort_keys=True),
                json.dumps(item.after_value, ensure_ascii=False, sort_keys=True),
            ]
        )
    output = build_audit_workbook(rows)
    filename = f"audit-logs-{now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/v1/audit-logs/{log_id}", response_model=AuditLogOut)
def get_audit_log(
    log_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/audit-logs/{log_id}")),
) -> AuditLogOut:
    log = db.get(AuditLog, log_id)
    if not log:
        raise HTTPException(status_code=404, detail="Audit log not found")
    return audit_to_out(log)


@app.get("/api/v1/users", response_model=list[UserOut])
def list_users(
    search: str = "",
    unit: str = "",
    department: str = "",
    team: str = "",
    account_ownership: str = "",
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/users")),
) -> list[UserOut]:
    ensure_hcm_seed_users(db)
    query = db.query(User)
    if search:
        like = f"%{search}%"
        query = query.filter(or_(User.username.like(like), User.display_name.like(like), User.email.like(like)))
    if unit:
        query = query.filter(User.unit == unit)
    if department:
        query = query.filter(User.department == department)
    if team:
        query = query.filter(User.team == team)
    if account_ownership:
        if account_ownership not in ACCOUNT_OWNERSHIPS:
            raise HTTPException(status_code=422, detail="account_ownership must be HCM or local")
        query = query.filter(User.account_ownership == account_ownership)
    return [user_to_out(user) for user in query.order_by(User.account_ownership, User.id).all()]


@app.post("/api/v1/users", response_model=UserOut)
def create_user(
    payload: UserIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/users")),
) -> UserOut:
    require_button_permission(auth, "button.users.create")
    ensure_hcm_seed_users(db)
    username = payload.username.strip()
    display_name = payload.display_name.strip()
    if not username or not display_name:
        raise HTTPException(status_code=422, detail="username and display_name are required")
    if db.query(User).filter(User.username == username).first():
        raise HTTPException(status_code=409, detail="Username must be unique")
    user = User(
        username=username,
        display_name=display_name,
        unit=payload.unit.strip(),
        department=payload.department.strip(),
        team=payload.team.strip(),
        email=payload.email.strip(),
        account_ownership="local",
        status=validate_user_status(payload.status),
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user_to_out(user)


@app.get("/api/v1/users/{user_id}", response_model=UserOut)
def get_user(
    user_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/users/{user_id}")),
) -> UserOut:
    ensure_hcm_seed_users(db)
    return user_to_out(get_user_or_404(db, user_id))


@app.put("/api/v1/users/{user_id}", response_model=UserOut)
def update_user(
    user_id: int,
    payload: UserUpdate,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/users/{user_id}")),
) -> UserOut:
    require_button_permission(auth, "button.users.edit")
    user = get_user_or_404(db, user_id)
    require_local_user(user)
    for field in ["display_name", "unit", "department", "team", "email"]:
        value = getattr(payload, field)
        if value is not None:
            setattr(user, field, value.strip())
    if payload.status is not None:
        user.status = validate_user_status(payload.status)
    if not user.display_name.strip():
        raise HTTPException(status_code=422, detail="display_name is required")
    user.updated_at = now()
    db.commit()
    db.refresh(user)
    return user_to_out(user)


@app.post("/api/v1/users/{user_id}/password-reset", response_model=PasswordResetOut)
def reset_user_password(
    user_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/users/{user_id}/password-reset")),
) -> PasswordResetOut:
    require_button_permission(auth, "button.users.reset_password")
    user = get_user_or_404(db, user_id)
    require_local_user(user)
    token = sha1(f"password-reset:{user.id}:{user.username}:{now().isoformat()}".encode("utf-8")).hexdigest()[:12].upper()
    user.password_reset_token = token
    user.updated_at = now()
    db.commit()
    return PasswordResetOut(
        user_id=user.id,
        username=user.username,
        reset_token=token,
        temporary_password=f"Temp-{token}",
        message="Local user password reset succeeded",
    )


@app.delete("/api/v1/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.DELETE./api/v1/users/{user_id}")),
) -> dict[str, Any]:
    require_button_permission(auth, "button.users.delete")
    user = get_user_or_404(db, user_id)
    require_local_user(user)
    db.delete(user)
    db.commit()
    return {"deleted": True, "id": user_id}


@app.get("/api/v1/permissions/catalog", response_model=list[PermissionEntry])
def get_permission_catalog(
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/permissions/catalog")),
) -> list[PermissionEntry]:
    return permission_catalog_entries(db)


@app.get("/api/v1/roles", response_model=list[RoleOut])
def list_roles(
    search: str = "",
    enabled: bool | None = None,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/roles")),
) -> list[RoleOut]:
    query = db.query(Role)
    if search:
        like = f"%{search}%"
        query = query.filter(or_(Role.name.like(like), Role.code.like(like), Role.description.like(like)))
    if enabled is not None:
        query = query.filter(Role.enabled.is_(enabled))
    return [role_to_out(role) for role in query.order_by(Role.id.desc()).all()]


@app.post("/api/v1/roles", response_model=RoleOut)
def create_role(
    payload: RoleIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/roles")),
) -> RoleOut:
    require_button_permission(auth, "button.roles.create")
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=422, detail="role name is required")
    code = generate_role_code(db)
    validate_role_uniqueness(db, name, code)
    role = Role(name=name, code=code, description=payload.description.strip(), enabled=payload.enabled)
    db.add(role)
    db.flush()
    add_audit_log(
        db,
        auth,
        "role",
        "create",
        {},
        {
            "id": role.id,
            "name": role.name,
            "code": role.code,
            "description": role.description,
            "enabled": role.enabled,
        },
    )
    db.commit()
    db.refresh(role)
    return role_to_out(role)


@app.get("/api/v1/roles/{role_id}", response_model=RoleOut)
def get_role(
    role_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/roles/{role_id}")),
) -> RoleOut:
    return role_to_out(get_role_or_404(db, role_id))


@app.put("/api/v1/roles/{role_id}", response_model=RoleOut)
def update_role(
    role_id: int,
    payload: RoleUpdate,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/roles/{role_id}")),
) -> RoleOut:
    require_button_permission(auth, "button.roles.edit")
    role = get_role_or_404(db, role_id)
    name = payload.name.strip() if payload.name is not None else role.name
    if not name:
        raise HTTPException(status_code=422, detail="role name is required")
    validate_role_uniqueness(db, name, role.code, role.id)
    role.name = name
    if payload.description is not None:
        role.description = payload.description.strip()
    if payload.enabled is not None:
        role.enabled = payload.enabled
    role.updated_at = now()
    db.commit()
    db.refresh(role)
    return role_to_out(role)


@app.patch("/api/v1/roles/{role_id}/enable", response_model=RoleOut)
def enable_role(
    role_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PATCH./api/v1/roles/{role_id}/enable")),
) -> RoleOut:
    require_button_permission(auth, "button.roles.edit")
    role = get_role_or_404(db, role_id)
    role.enabled = True
    role.updated_at = now()
    db.commit()
    db.refresh(role)
    return role_to_out(role)


@app.patch("/api/v1/roles/{role_id}/disable", response_model=RoleOut)
def disable_role(
    role_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PATCH./api/v1/roles/{role_id}/disable")),
) -> RoleOut:
    require_button_permission(auth, "button.roles.edit")
    role = get_role_or_404(db, role_id)
    role.enabled = False
    role.updated_at = now()
    db.commit()
    db.refresh(role)
    return role_to_out(role)


@app.delete("/api/v1/roles/{role_id}")
def delete_role(
    role_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.DELETE./api/v1/roles/{role_id}")),
) -> dict[str, Any]:
    require_button_permission(auth, "button.roles.delete")
    role = get_role_or_404(db, role_id)
    db.delete(role)
    db.commit()
    return {"deleted": True, "id": role_id}


@app.get("/api/v1/roles/{role_id}/users", response_model=list[UserSummaryOut])
def list_role_users(
    role_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/roles/{role_id}/users")),
) -> list[UserSummaryOut]:
    role = get_role_or_404(db, role_id)
    return [user_summary(link.user) for link in sorted(role.user_links, key=lambda link: link.user.username)]


@app.post("/api/v1/roles/{role_id}/users", response_model=RoleOut)
def add_role_user(
    role_id: int,
    payload: RoleUserBindingIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/roles/{role_id}/users")),
) -> RoleOut:
    require_button_permission(auth, "button.roles.bind_users")
    role = get_role_or_404(db, role_id)
    validate_bindable_role(role)
    user = get_user_or_404(db, payload.user_id)
    existing = db.query(RoleUser).filter(RoleUser.role_id == role.id, RoleUser.user_id == user.id).first()
    if not existing:
        db.add(RoleUser(role_id=role.id, user_id=user.id))
    role.updated_at = now()
    db.commit()
    db.refresh(role)
    return role_to_out(role)


@app.put("/api/v1/roles/{role_id}/users", response_model=RoleOut)
def replace_role_users(
    role_id: int,
    payload: RoleUserReplaceIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/roles/{role_id}/users")),
) -> RoleOut:
    require_button_permission(auth, "button.roles.bind_users")
    role = get_role_or_404(db, role_id)
    validate_bindable_role(role)
    user_ids = set(payload.user_ids)
    users = db.query(User).filter(User.id.in_(user_ids)).all() if user_ids else []
    found_ids = {user.id for user in users}
    missing = user_ids - found_ids
    if missing:
        raise HTTPException(status_code=404, detail=f"User not found: {sorted(missing)[0]}")
    db.query(RoleUser).filter(RoleUser.role_id == role.id).delete()
    for user_id in sorted(user_ids):
        db.add(RoleUser(role_id=role.id, user_id=user_id))
    role.updated_at = now()
    db.commit()
    db.refresh(role)
    return role_to_out(role)


@app.delete("/api/v1/roles/{role_id}/users/{user_id}", response_model=RoleOut)
def remove_role_user(
    role_id: int,
    user_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.DELETE./api/v1/roles/{role_id}/users/{user_id}")),
) -> RoleOut:
    require_button_permission(auth, "button.roles.bind_users")
    role = get_role_or_404(db, role_id)
    get_user_or_404(db, user_id)
    link = db.query(RoleUser).filter(RoleUser.role_id == role.id, RoleUser.user_id == user_id).first()
    if link:
        db.delete(link)
    role.updated_at = now()
    db.commit()
    db.refresh(role)
    return role_to_out(role)


@app.get("/api/v1/roles/{role_id}/permissions", response_model=RolePermissionsOut)
def get_role_permissions(
    role_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/roles/{role_id}/permissions")),
) -> RolePermissionsOut:
    role = get_role_or_404(db, role_id)
    return RolePermissionsOut(
        role_id=role.id,
        role_name=role.name,
        permissions=[
            permission_to_entry(permission)
            for permission in sorted(role.permissions, key=lambda item: (item.permission_type, item.permission_key))
            if permission.enabled
        ],
        catalog=permission_catalog_entries(db),
    )


@app.put("/api/v1/roles/{role_id}/permissions", response_model=RolePermissionsOut)
def save_role_permissions(
    role_id: int,
    payload: RolePermissionsIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/roles/{role_id}/permissions")),
) -> RolePermissionsOut:
    require_button_permission(auth, "button.roles.configure_permissions")
    role = get_role_or_404(db, role_id)
    entries = normalize_permission_payload(payload, db)
    db.query(FeaturePermission).filter(FeaturePermission.role_id == role.id).delete()
    for entry in entries:
        db.add(
            FeaturePermission(
                role_id=role.id,
                module=entry.module,
                permission_type=entry.permission_type,
                permission_key=entry.permission_key,
                label=entry.label,
                enabled=True,
            )
        )
    role.updated_at = now()
    db.commit()
    db.refresh(role)
    return get_role_permissions(role.id, db, auth)


@app.get("/api/v1/workflows/applications", response_model=list[WorkflowApplicationOut])
def list_workflow_applications(
    applicant: str = "",
    status: str = "",
    type: str = "",
    material_id: int | None = None,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/workflows/applications")),
) -> list[WorkflowApplicationOut]:
    query = db.query(WorkflowApplication)
    if applicant:
        query = query.filter(WorkflowApplication.applicant == applicant)
    if status:
        query = query.filter(WorkflowApplication.status == status)
    if type:
        query = query.filter(WorkflowApplication.type == type)
    applications = query.order_by(WorkflowApplication.id.desc()).all()
    if material_id is not None:
        applications = [
            application
            for application in applications
            if int(workflow_payload(application.payload).get("material_id") or 0) == material_id
        ]
    return [workflow_to_out(application) for application in applications]


@app.post("/api/v1/workflows/applications", response_model=WorkflowApplicationOut)
def submit_workflow_application(
    payload: WorkflowApplicationIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/workflows/applications")),
) -> WorkflowApplicationOut:
    require_button_permission(auth, "button.workflow.submit")
    if payload.material_library_id:
        require_library_scope(auth, payload.material_library_id)
    if payload.material_id:
        material = db.get(Material, payload.material_id)
        if material:
            require_library_scope(auth, material.material_library_id)
    data = build_workflow_payload(payload, db)
    mode = approval_mode(db)
    status, node = initial_workflow_state(mode)
    seed = f"{payload.type}:{payload.applicant}:{now().isoformat()}:{data}"
    application = WorkflowApplication(
        application_no=application_no(seed),
        type=payload.type,
        status=status,
        applicant=payload.applicant.strip() or "material_manager",
        current_node=node,
        business_reason=payload.business_reason.strip(),
        payload=json.dumps(data, ensure_ascii=False),
    )
    db.add(application)
    db.flush()
    add_workflow_history(
        application,
        "submit",
        application.applicant,
        "applicant",
        "draft",
        status,
        payload.business_reason.strip(),
    )
    db.commit()
    db.refresh(application)
    return workflow_to_out(application)


@app.post("/api/v1/workflows/applications/new-category", response_model=WorkflowApplicationOut)
def submit_new_category_application(
    payload: WorkflowApplicationIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/workflows/applications")),
) -> WorkflowApplicationOut:
    payload.type = "new_category"
    return submit_workflow_application(payload, db, auth)


@app.post("/api/v1/workflows/applications/new-material-code", response_model=WorkflowApplicationOut)
def submit_new_material_code_application(
    payload: WorkflowApplicationIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/workflows/applications")),
) -> WorkflowApplicationOut:
    payload.type = "new_material_code"
    return submit_workflow_application(payload, db, auth)


@app.post("/api/v1/workflows/applications/stop-purchase", response_model=WorkflowApplicationOut)
def submit_stop_purchase_application(
    payload: WorkflowApplicationIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/workflows/applications")),
) -> WorkflowApplicationOut:
    payload.type = "stop_purchase"
    return submit_workflow_application(payload, db, auth)


@app.post("/api/v1/workflows/applications/stop-use", response_model=WorkflowApplicationOut)
def submit_stop_use_application(
    payload: WorkflowApplicationIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/workflows/applications")),
) -> WorkflowApplicationOut:
    payload.type = "stop_use"
    return submit_workflow_application(payload, db, auth)


@app.get("/api/v1/workflows/tasks", response_model=list[WorkflowApplicationOut])
def list_workflow_tasks(
    node: str = "",
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/workflows/tasks")),
) -> list[WorkflowApplicationOut]:
    query = db.query(WorkflowApplication).filter(WorkflowApplication.status.notin_(TERMINAL_WORKFLOW_STATUSES))
    if node:
        query = query.filter(WorkflowApplication.current_node == node)
    applications = query.order_by(WorkflowApplication.id).all()
    return [workflow_to_out(application) for application in applications]


@app.get("/api/v1/workflows/applications/{application_id}", response_model=WorkflowApplicationOut)
def get_workflow_application(
    application_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/workflows/applications/{application_id}")),
) -> WorkflowApplicationOut:
    application = db.get(WorkflowApplication, application_id)
    if not application:
        raise HTTPException(status_code=404, detail="Workflow application not found")
    return workflow_to_out(application)


@app.post("/api/v1/workflows/applications/{application_id}/approve", response_model=WorkflowApplicationOut)
def approve_workflow_application(
    application_id: int,
    payload: WorkflowActionIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/workflows/applications/{application_id}/approve")),
) -> WorkflowApplicationOut:
    require_button_permission(auth, "button.workflow.approve")
    application = db.get(WorkflowApplication, application_id)
    if not application:
        raise HTTPException(status_code=404, detail="Workflow application not found")
    if application.status in TERMINAL_WORKFLOW_STATUSES:
        raise HTTPException(status_code=409, detail="Terminal workflow applications cannot be approved again")
    requested_node = payload.node or application.current_node
    if requested_node != application.current_node:
        raise HTTPException(status_code=409, detail=f"Current workflow node is {application.current_node}, not {requested_node}")
    from_status = application.status
    to_status, next_node = next_approval_state(application.current_node, approval_mode(db))
    application.status = to_status
    application.current_node = next_node
    application.updated_at = now()
    if to_status == "approved":
        complete_workflow_application(application, db)
    add_workflow_history(
        application,
        "approve",
        payload.actor.strip() or requested_node,
        requested_node,
        from_status,
        to_status,
        payload.comment,
    )
    db.commit()
    db.refresh(application)
    return workflow_to_out(application)


@app.post("/api/v1/workflows/applications/{application_id}/reject", response_model=WorkflowApplicationOut)
def reject_workflow_application(
    application_id: int,
    payload: WorkflowActionIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/workflows/applications/{application_id}/reject")),
) -> WorkflowApplicationOut:
    require_button_permission(auth, "button.workflow.reject")
    application = db.get(WorkflowApplication, application_id)
    if not application:
        raise HTTPException(status_code=404, detail="Workflow application not found")
    if application.status in TERMINAL_WORKFLOW_STATUSES:
        raise HTTPException(status_code=409, detail="Terminal workflow applications cannot be rejected again")
    comment = payload.comment.strip()
    if not comment:
        raise HTTPException(status_code=422, detail="A rejection reason is required")
    requested_node = payload.node or application.current_node
    if requested_node != application.current_node:
        raise HTTPException(status_code=409, detail=f"Current workflow node is {application.current_node}, not {requested_node}")
    from_status = application.status
    application.status = "rejected"
    application.current_node = "rejected"
    application.rejection_reason = comment
    application.updated_at = now()
    add_workflow_history(
        application,
        "reject",
        payload.actor.strip() or requested_node,
        requested_node,
        from_status,
        "rejected",
        comment,
    )
    db.commit()
    db.refresh(application)
    return workflow_to_out(application)


@app.get("/api/v1/materials", response_model=list[MaterialOut])
def list_materials(
    search: str = "",
    status: str = "",
    product_name_id: int | None = None,
    material_library_id: int | None = None,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/materials")),
) -> list[MaterialOut]:
    ensure_seed_material_context(db)
    query = db.query(Material).join(ProductName).join(MaterialLibrary).join(Category)
    if material_library_id:
        if not is_library_in_scope(auth, material_library_id):
            return []
        query = query.filter(Material.material_library_id == material_library_id)
    elif not auth.is_super_admin:
        query = query.filter(Material.material_library_id.in_(auth.library_scope_ids or {-1}))
    if product_name_id:
        query = query.filter(Material.product_name_id == product_name_id)
    if status:
        validate_material_status(status)
        query = query.filter(Material.status == status)
    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                Material.name.like(like),
                Material.code.like(like),
                Material.description.like(like),
                Material.attributes.like(like),
                ProductName.name.like(like),
            )
        )
    materials = query.order_by(Material.id.desc()).all()
    return [material_to_out(material) for material in materials]


@app.post("/api/v1/materials", response_model=MaterialOut)
def create_material(
    payload: MaterialIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials")),
) -> MaterialOut:
    require_button_permission(auth, "button.material_archives.create")
    require_library_scope(auth, payload.material_library_id)
    status = validate_material_status(payload.status)
    if status != "normal":
        raise HTTPException(status_code=400, detail="New materials must start in normal status")
    product, library, category = material_context_by_payload(
        db,
        payload.product_name_id,
        payload.material_library_id,
        payload.category_id,
    )
    validate_required_category_properties(db, library, category, payload.attributes)
    existing = db.query(Material).filter(Material.product_name_id == product.id, Material.name == payload.name).first()
    if existing:
        raise HTTPException(status_code=409, detail="Material already exists for this product name")
    brand = db.get(Brand, payload.brand_id) if payload.brand_id else None
    if payload.brand_id and not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    active_rule = active_rule_for_library(db, library) if library.auto_code_enabled else None
    material_code = next_unique_code(db, Material, "MAT", f"{product.id}:{payload.name}:{now().isoformat()}")
    code_status = "manual"
    if library.auto_code_enabled:
        if not active_rule:
            raise HTTPException(status_code=422, detail="Active code rule is required for auto-code material library")
        material_code = generate_material_code(
            db,
            "default",
            library.id,
            {"product": product, "library": library, "category": category, "attributes": payload.attributes},
            active_rule,
        )
        code_status = "active"
    material = Material(
        code=material_code,
        name=payload.name,
        product_name_id=product.id,
        material_library_id=library.id,
        category_id=category.id,
        unit=payload.unit or product.unit,
        brand_id=brand.id if brand else None,
        status=status,
        description=payload.description,
        attributes=json.dumps(payload.attributes, ensure_ascii=False),
        code_rule_version_id=active_rule.id if active_rule else None,
        code_change_count=0,
        code_status=code_status,
        enabled=payload.enabled,
    )
    db.add(material)
    db.commit()
    db.refresh(material)
    return material_to_out(material)


def build_ai_material_preview(payload: AiMaterialAddPreviewIn, db: Session) -> dict[str, Any]:
    text = compact_space(payload.input_text)
    if not text:
        raise HTTPException(status_code=422, detail="input_text is required")
    try:
        gateway_result = invoke_gateway_capability(db, "material_add", text)
        match_gateway = invoke_gateway_capability(db, "material_match", text)
    except CapabilityResolutionError as exc:
        raise capability_resolution_http_error(exc) from exc
    library = db.get(MaterialLibrary, payload.material_library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Material library not found")

    inferred_product, inferred_category, inferred_unit = infer_product_category(text)
    unit = payload.unit or infer_unit(text) or inferred_unit
    product = db.get(ProductName, payload.product_name_id) if payload.product_name_id else None
    category = db.get(Category, payload.category_id) if payload.category_id else None
    if payload.product_name_id and not product:
        raise HTTPException(status_code=404, detail="Product name not found")
    if payload.category_id and not category:
        raise HTTPException(status_code=404, detail="Category not found")
    category = category or get_or_create_category(db, inferred_category)
    product = product or get_or_create_product_name(db, inferred_product, unit, category.name)

    brand = db.get(Brand, payload.brand_id) if payload.brand_id else None
    if payload.brand_id and not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    brand_name = brand.name if brand else infer_brand_name(text)
    proposed_brand = brand or get_or_create_brand(db, brand_name)

    attributes, attribute_sources = extract_material_attributes(text)
    name = infer_material_name(text)
    description = text
    field_sources = {
        "name": "unstructured input phrase after add/new intent" if name else "missing",
        "unit": "explicit unit hint or inferred category default" if unit else "missing",
        "brand": "brand hint, brand field, or known brand dictionary" if brand_name else "missing",
        "product_name": "keyword category inference",
        "category": "keyword category inference",
        "attributes": attribute_sources,
    }
    errors: list[str] = []
    if not name:
        errors.append("Material name is required")
    if not unit:
        errors.append("Unit is missing or ambiguous")
    if len(attributes) < 2:
        errors.append("At least two material attributes should be extracted before confirmation")

    confidence = round(min(0.98, 0.58 + (0.12 if name else 0) + (0.08 if unit else 0) + (0.08 if brand_name else 0) + min(len(attributes), 4) * 0.06), 2)
    query_text = material_search_text(name, brand_name, description, attributes)
    matches = material_matches(db, library.id, query_text, brand_name, attributes, 3)
    top_classification = matches[0]["classification"] if matches else "normal"
    trace_id = gateway_result["trace_id"]
    proposed = {
        "name": name,
        "unit": unit,
        "product_name_id": product.id,
        "product_name": product.name,
        "material_library_id": library.id,
        "material_library": library.name,
        "category_id": category.id,
        "category": category.name,
        "brand_id": proposed_brand.id if proposed_brand else None,
        "brand": proposed_brand.name if proposed_brand else brand_name,
        "description": description,
        "attributes": attributes,
        "status": "normal",
    }
    db.commit()
    return {
        "capability": "material_add",
        "provider": gateway_result["provider"],
        "model": gateway_result["model"],
        "trace_id": trace_id,
        "preview_token": sha1(json.dumps(proposed, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest(),
        "confidence": confidence,
        "validation_errors": errors,
        "field_sources": field_sources,
        "name": proposed["name"],
        "unit": proposed["unit"],
        "product_name_id": proposed["product_name_id"],
        "product_name": proposed["product_name"],
        "material_library_id": proposed["material_library_id"],
        "material_library": proposed["material_library"],
        "category_id": proposed["category_id"],
        "category": proposed["category"],
        "brand_id": proposed["brand_id"],
        "brand": proposed["brand"],
        "description": proposed["description"],
        "attributes": proposed["attributes"],
        "proposed_material": proposed,
        "duplicate_check": {
            "capability": "material_match",
            "provider": match_gateway["provider"],
            "model": match_gateway["model"],
            "trace_id": match_gateway["trace_id"],
            "engine": "qdrant_hybrid_with_local_fallback",
            "classification": top_classification,
            "top_matches": matches,
        },
    }


@app.post(
    "/api/v1/materials/ai-add/preview",
    description="AI natural language material addition preview. capability: material_add",
)
def preview_ai_material_add(
    payload: AiMaterialAddPreviewIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials/ai-add/preview")),
) -> dict[str, Any]:
    require_library_scope(auth, payload.material_library_id)
    return build_ai_material_preview(payload, db)


@app.post(
    "/api/v1/materials/ai-add/confirm",
    description="AI natural language material addition confirmation. capability: material_add",
)
def confirm_ai_material_add(
    payload: AiMaterialAddConfirmIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials/ai-add/confirm")),
) -> dict[str, Any]:
    require_button_permission(auth, "button.material_archives.create")
    preview = payload.preview
    try:
        resolution = model_for_capability(db, "material_add")
    except CapabilityResolutionError as exc:
        raise capability_resolution_http_error(exc) from exc
    errors = preview.get("validation_errors") or []
    if errors:
        raise HTTPException(status_code=422, detail={"validation_errors": errors})
    proposed = preview.get("proposed_material") or preview
    duplicate_check = preview.get("duplicate_check") or {}
    if not payload.allow_duplicate and duplicate_check.get("classification") == "highly_duplicate":
        raise HTTPException(status_code=409, detail="Highly duplicate material requires explicit duplicate override")
    name = str(proposed.get("name", "")).strip()
    if not name:
        raise HTTPException(status_code=422, detail="Material name is required")
    product, library, category = material_context_by_payload(
        db,
        int(proposed.get("product_name_id") or 0),
        int(proposed.get("material_library_id") or 0),
        int(proposed.get("category_id") or 0),
    )
    require_library_scope(auth, library.id)
    brand_id = proposed.get("brand_id")
    brand = db.get(Brand, int(brand_id)) if brand_id else get_or_create_brand(db, str(proposed.get("brand", "")).strip())
    existing = db.query(Material).filter(Material.product_name_id == product.id, Material.name == name).first()
    if existing:
        raise HTTPException(status_code=409, detail="Material already exists for this product name")
    material = Material(
        code=next_unique_code(db, Material, "MAT", f"{product.id}:{name}:{now().isoformat()}"),
        name=name,
        product_name_id=product.id,
        material_library_id=library.id,
        category_id=category.id,
        unit=str(proposed.get("unit") or product.unit),
        brand_id=brand.id if brand else None,
        status="normal",
        description=str(proposed.get("description", "")),
        attributes=json.dumps(material_attributes(proposed.get("attributes")), ensure_ascii=False),
        enabled=True,
    )
    db.add(material)
    db.commit()
    db.refresh(material)
    return {
        "capability": "material_add",
        "provider": resolution.model.provider,
        "model": resolution.model.model_name,
        "resolution_source": "capability_mapping",
        "trace_id": preview.get("trace_id") or f"trace-{sha1(material.code.encode('utf-8')).hexdigest()[:16]}",
        "material": material_to_out(material).model_dump(),
    }


@app.post(
    "/api/v1/materials/match",
    description="AI vector similarity material matching. capability: material_match; hybrid semantic + BM25 evidence with Qdrant-compatible local fallback.",
)
def match_materials(
    payload: MaterialMatchIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials/match")),
) -> dict[str, Any]:
    library = db.get(MaterialLibrary, payload.material_library_id)
    if not library:
        raise HTTPException(status_code=404, detail="Material library not found")
    require_library_scope(auth, library.id)
    brand = db.get(Brand, payload.brand_id).name if payload.brand_id and db.get(Brand, payload.brand_id) else (payload.brand or "")
    query_text = payload.query or material_search_text(payload.name or "", brand, payload.description, payload.attributes)
    try:
        gateway_result = invoke_gateway_capability(db, "material_match", query_text)
    except CapabilityResolutionError as exc:
        raise capability_resolution_http_error(exc) from exc
    matches = material_matches(db, library.id, query_text, brand, payload.attributes, payload.top_k)
    return {
        "capability": "material_match",
        "provider": gateway_result["provider"],
        "model": gateway_result["model"],
        "trace_id": gateway_result["trace_id"],
        "fallback_used": gateway_result["fallback_used"],
        "embedding_provider": gateway_result["provider"],
        "engine": "qdrant_hybrid_with_local_fallback",
        "query": query_text,
        "matches": matches,
    }


@app.post("/api/v1/ai/material-add/preview")
def preview_ai_material_add_alias(
    payload: AiMaterialAddPreviewIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials/ai-add/preview")),
) -> dict[str, Any]:
    return preview_ai_material_add(payload, db, auth)


@app.post("/api/v1/ai/material-add/confirm")
def confirm_ai_material_add_alias(
    payload: AiMaterialAddConfirmIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials/ai-add/confirm")),
) -> dict[str, Any]:
    return confirm_ai_material_add(payload, db, auth)


@app.post("/api/v1/ai/material-match")
def match_materials_alias(
    payload: MaterialMatchIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials/match")),
) -> dict[str, Any]:
    return match_materials(payload, db, auth)


@app.post("/api/v1/ai/material-category-match", response_model=MaterialCategoryMatchOut)
def match_material_category(
    payload: MaterialCategoryMatchIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/ai/material-category-match")),
) -> MaterialCategoryMatchOut:
    require_super_admin(auth)
    material_name = compact_space(payload.material_name)
    if not material_name:
        raise HTTPException(status_code=422, detail="material_name is required")
    library_ids = unique_int_ids(payload.category_library_ids)
    if not library_ids:
        return MaterialCategoryMatchOut(matches=[], results=[], message="No category libraries selected", resolution_source="capability_mapping")

    libraries = db.query(CategoryLibrary).filter(CategoryLibrary.id.in_(library_ids)).order_by(CategoryLibrary.id).all()
    found_ids = {library.id for library in libraries}
    missing_ids = [library_id for library_id in library_ids if library_id not in found_ids]
    if missing_ids:
        raise HTTPException(status_code=404, detail=f"Category library not found: {missing_ids[0]}")

    enabled_libraries = [library for library in libraries if library.qdrant_enabled]
    if not enabled_libraries:
        return MaterialCategoryMatchOut(matches=[], results=[], message="No Qdrant-enabled category libraries selected", resolution_source="capability_mapping")

    query_text = " ".join(
        part
        for part in [
            material_name,
            compact_space(payload.brand),
            compact_space(payload.description),
        ]
        if part
    )
    collector = SpanCollector("material_match.category_match", "material_match")
    try:
        resolution = model_for_capability(db, "material_match")
    except CapabilityResolutionError as exc:
        collector.finish_span(collector.root_span_id, "error", str(exc), {"capability": exc.capability})
        collector.flush(db)
        raise capability_resolution_http_error(exc) from exc
    mark_root_trace_model(collector, resolution.model, resolution.model.model_name, resolution_trace_metadata(resolution))
    collector.finish_span(collector.root_span_id, "ok", metadata={"query": query_text, **resolution_trace_metadata(resolution)})
    collector.flush(db)
    vector = category_embedding(query_text)
    candidates: list[dict[str, Any]] = []
    for library in enabled_libraries:
        try:
            for raw in search_category_collection(library.id, vector, 3):
                item = category_match_item(raw)
                if item:
                    candidates.append(item)
        except QdrantSyncError as exc:
            trace_qdrant_error(db, "qdrant.category.search", str(exc), {"library_id": library.id})

    deduped: dict[int, dict[str, Any]] = {}
    for item in candidates:
        existing = deduped.get(item["category_id"])
        if existing is None or item["score"] > existing["score"]:
            deduped[item["category_id"]] = item
    matches = sorted(deduped.values(), key=lambda item: item["score"], reverse=True)[:3]
    message = "" if matches else "No matching categories found"
    return MaterialCategoryMatchOut(
        matches=matches,
        results=matches,
        message=message,
        trace_id=collector.trace_id,
        provider=resolution.model.provider,
        model=resolution.model.model_name,
        resolution_source="capability_mapping",
    )


def capability_resolution_http_error(exc: CapabilityResolutionError) -> HTTPException:
    return HTTPException(
        status_code=exc.status_code,
        detail={
            "error": str(exc),
            "capability": exc.capability,
            "suggestion": exc.suggestion,
        },
    )


@app.get("/api/v1/ai/resolve-model")
def resolve_ai_model(
    capability: str,
    prefer_fallback: bool = False,
    include_metrics: bool = False,
    request: Request = None,
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    current_auth(request, db)
    try:
        return timed_resolution_response_payload(db, capability, prefer_fallback)
    except CapabilityResolutionError as exc:
        raise capability_resolution_http_error(exc) from exc


@app.post("/api/v1/ai/resolve-model/batch")
def resolve_ai_model_batch(
    payload: dict[str, Any] | list[str],
    request: Request,
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    current_auth(request, db)
    if isinstance(payload, list):
        capabilities = payload
    else:
        raw = payload.get("capabilities") or payload.get("items") or []
        capabilities = raw if isinstance(raw, list) else []
    started = time.perf_counter()
    results: list[dict[str, Any]] = []
    for capability in capabilities:
        try:
            results.append(timed_resolution_response_payload(db, str(capability)))
        except CapabilityResolutionError as exc:
            results.append(
                {
                    "capability": exc.capability,
                    "error": str(exc),
                    "suggestion": exc.suggestion,
                }
            )
    return {
        "batch_lookup_ms": max(0, (time.perf_counter() - started) * 1000),
        "results": results,
    }


@app.post("/api/v1/model-gateway-migration/run")
def run_model_gateway_migration(
    request: Request,
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    auth = current_auth(request, db)
    require_super_admin(auth)
    return run_sprint55_migration(db)


@app.get("/api/v1/models", response_model=list[ModelRead])
def list_models(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    provider: str = "",
    enabled: bool | None = None,
    request: Request = None,
    db: Session = Depends(get_db),
) -> list[ModelRead]:
    current_auth(request, db)
    ensure_model_gateway_schema(db)
    query = db.query(Model)
    if provider:
        query = query.filter(Model.provider == compact_space(provider).lower())
    if enabled is not None:
        query = query.filter(Model.enabled.is_(enabled))
    models = query.order_by(Model.enabled.desc(), Model.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return [model_to_read(model) for model in models]


@app.post("/api/v1/models", response_model=ModelRead)
def create_model(
    payload: ModelCreate,
    request: Request,
    db: Session = Depends(get_db),
) -> ModelRead:
    auth = current_auth(request, db)
    require_super_admin(auth)
    ensure_model_gateway_schema(db)
    values = payload.model_dump(exclude={"api_key"})
    model = Model(connection_status="untested", migration_data_version="migrated")
    apply_gateway_model_values(model, values, payload.api_key)
    db.add(model)
    try:
        db.flush()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="provider and model_name pair must be unique") from exc
    add_audit_log(db, auth, "model", "create", {}, model_snapshot(model), "human")
    db.commit()
    db.refresh(model)
    return model_to_read(model)


@app.get("/api/v1/models/{model_id}/test", response_model=ModelTestResult)
def test_saved_model(
    model_id: int,
    request: Request,
    db: Session = Depends(get_db),
) -> ModelTestResult:
    auth = current_auth(request, db)
    require_super_admin(auth)
    ensure_model_gateway_schema(db)
    model = db.get(Model, model_id)
    if not model:
        raise HTTPException(status_code=404, detail="Model not found")
    started = time.perf_counter()
    result = test_model_connection(model)
    latency_ms = max(0, int((time.perf_counter() - started) * 1000))
    before = model_snapshot(model)
    model.connection_status = "ok" if result["ok"] else "error"
    model.last_tested_at = now()
    model.updated_at = now()
    add_audit_log(db, auth, "model", "test", before, model_snapshot(model), "human")
    db.commit()
    return ModelTestResult(
        ok=bool(result["ok"]),
        status=model.connection_status,
        message=str(result["message"]),
        latency_ms=latency_ms,
        provider=model.provider,
        model_name=model.model_name,
        tested_at=model.last_tested_at.isoformat(),
        last_tested_at=model.last_tested_at.isoformat(),
    )


@app.get("/api/v1/models/{model_id}", response_model=ModelRead)
def get_model(
    model_id: int,
    request: Request,
    db: Session = Depends(get_db),
) -> ModelRead:
    current_auth(request, db)
    ensure_model_gateway_schema(db)
    model = db.get(Model, model_id)
    if not model:
        raise HTTPException(status_code=404, detail="Model not found")
    return model_to_read(model)


@app.put("/api/v1/models/{model_id}", response_model=ModelRead)
def update_model(
    model_id: int,
    payload: ModelUpdate,
    request: Request,
    db: Session = Depends(get_db),
) -> ModelRead:
    auth = current_auth(request, db)
    require_super_admin(auth)
    ensure_model_gateway_schema(db)
    model = db.get(Model, model_id)
    if not model:
        raise HTTPException(status_code=404, detail="Model not found")
    before = model_snapshot(model)
    values = payload.model_dump(exclude_unset=True, exclude={"api_key"})
    apply_gateway_model_values(model, values, payload.api_key)
    try:
        db.flush()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="provider and model_name pair must be unique") from exc
    add_audit_log(db, auth, "model", "update", before, model_snapshot(model), "human")
    db.commit()
    db.refresh(model)
    return model_to_read(model)


@app.patch("/api/v1/models/{model_id}/toggle", response_model=ModelRead)
def toggle_model(
    model_id: int,
    request: Request,
    db: Session = Depends(get_db),
) -> ModelRead:
    auth = current_auth(request, db)
    require_super_admin(auth)
    ensure_model_gateway_schema(db)
    model = db.get(Model, model_id)
    if not model:
        raise HTTPException(status_code=404, detail="Model not found")
    before = model_snapshot(model)
    model.enabled = not model.enabled
    model.updated_at = now()
    add_audit_log(db, auth, "model", "toggle", before, model_snapshot(model), "human")
    db.commit()
    db.refresh(model)
    return model_to_read(model)


@app.delete("/api/v1/models/{model_id}")
def delete_model(
    model_id: int,
    request: Request,
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    auth = current_auth(request, db)
    require_super_admin(auth)
    ensure_model_gateway_schema(db)
    model = db.get(Model, model_id)
    if not model:
        raise HTTPException(status_code=404, detail="Model not found")
    if db.query(CapabilityMapping).filter(
        or_(
            CapabilityMapping.primary_model_id == model.id,
            CapabilityMapping.fallback_model_id == model.id,
        )
    ).first():
        raise HTTPException(status_code=409, detail="Model is referenced by a capability mapping; remove the mapping before deleting")
    before = model_snapshot(model)
    db.delete(model)
    add_audit_log(db, auth, "model", "delete", before, {}, "human")
    db.commit()
    return {"deleted": True, "id": model_id}


@app.get("/api/v1/capability-mappings", response_model=list[CapabilityMappingRead])
def list_gateway_capability_mappings(
    request: Request,
    db: Session = Depends(get_db),
) -> list[CapabilityMappingRead]:
    current_auth(request, db)
    ensure_model_gateway_schema(db)
    mappings = db.query(CapabilityMapping).order_by(CapabilityMapping.capability).all()
    return [capability_mapping_to_read(mapping) for mapping in mappings]


@app.post("/api/v1/capability-mappings", response_model=CapabilityMappingRead)
def create_gateway_capability_mapping(
    payload: CapabilityMappingCreate,
    request: Request,
    db: Session = Depends(get_db),
) -> CapabilityMappingRead:
    auth = current_auth(request, db)
    require_super_admin(auth)
    ensure_model_gateway_schema(db)
    capability = compact_space(payload.capability)
    if not capability:
        raise HTTPException(status_code=422, detail="capability is required")
    ensure_distinct_mapping_models(payload.primary_model_id, payload.fallback_model_id)
    ensure_gateway_model_reference(db, payload.primary_model_id, "Primary")
    ensure_gateway_model_reference(db, payload.fallback_model_id, "Fallback")
    mapping = CapabilityMapping(
        capability=capability,
        primary_model_id=payload.primary_model_id,
        fallback_model_id=payload.fallback_model_id,
        enabled=payload.enabled,
        migration_data_version="migrated",
    )
    db.add(mapping)
    try:
        db.flush()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="capability must be unique") from exc
    add_audit_log(db, auth, "capability_mapping", "create", {}, capability_mapping_snapshot(mapping), "human")
    db.commit()
    db.refresh(mapping)
    return capability_mapping_to_read(mapping)


@app.get("/api/v1/capability-mappings/{mapping_id}", response_model=CapabilityMappingRead)
def get_gateway_capability_mapping(
    mapping_id: int,
    request: Request,
    db: Session = Depends(get_db),
) -> CapabilityMappingRead:
    current_auth(request, db)
    ensure_model_gateway_schema(db)
    mapping = db.get(CapabilityMapping, mapping_id)
    if not mapping:
        raise HTTPException(status_code=404, detail="Capability mapping not found")
    return capability_mapping_to_read(mapping)


@app.put("/api/v1/capability-mappings/{mapping_id}", response_model=CapabilityMappingRead)
def update_gateway_capability_mapping(
    mapping_id: int,
    payload: CapabilityMappingUpdate,
    request: Request,
    db: Session = Depends(get_db),
) -> CapabilityMappingRead:
    auth = current_auth(request, db)
    require_super_admin(auth)
    ensure_model_gateway_schema(db)
    mapping = db.get(CapabilityMapping, mapping_id)
    if not mapping:
        raise HTTPException(status_code=404, detail="Capability mapping not found")
    primary_model_id = payload.primary_model_id if "primary_model_id" in payload.model_fields_set else mapping.primary_model_id
    fallback_model_id = payload.fallback_model_id if "fallback_model_id" in payload.model_fields_set else mapping.fallback_model_id
    ensure_distinct_mapping_models(primary_model_id, fallback_model_id)
    ensure_gateway_model_reference(db, primary_model_id, "Primary")
    ensure_gateway_model_reference(db, fallback_model_id, "Fallback")
    before = capability_mapping_snapshot(mapping)
    if "primary_model_id" in payload.model_fields_set:
        mapping.primary_model_id = payload.primary_model_id
    if "fallback_model_id" in payload.model_fields_set:
        mapping.fallback_model_id = payload.fallback_model_id
    if payload.enabled is not None:
        mapping.enabled = payload.enabled
    mapping.updated_at = now()
    add_audit_log(db, auth, "capability_mapping", "update", before, capability_mapping_snapshot(mapping), "human")
    db.commit()
    db.refresh(mapping)
    return capability_mapping_to_read(mapping)


@app.delete("/api/v1/capability-mappings/{mapping_id}")
def delete_gateway_capability_mapping(
    mapping_id: int,
    request: Request,
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    auth = current_auth(request, db)
    require_super_admin(auth)
    ensure_model_gateway_schema(db)
    mapping = db.get(CapabilityMapping, mapping_id)
    if not mapping:
        raise HTTPException(status_code=404, detail="Capability mapping not found")
    before = capability_mapping_snapshot(mapping)
    db.delete(mapping)
    add_audit_log(db, auth, "capability_mapping", "delete", before, {}, "human")
    db.commit()
    return {"deleted": True, "id": mapping_id}


@app.post("/api/v1/ai/capabilities/{capability}/invoke")
def invoke_ai_capability(
    capability: str,
    payload: GatewayInvokeIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials/ai-add/preview")),
) -> dict[str, Any]:
    if capability not in AI_CAPABILITIES:
        raise HTTPException(status_code=422, detail="Unsupported AI capability")
    try:
        return invoke_gateway_capability(db, capability, payload.prompt, payload.messages)
    except CapabilityResolutionError as exc:
        raise capability_resolution_http_error(exc) from exc


@app.get("/api/v1/debug/trace", response_model=list[TraceSummaryOut])
def list_traces(
    status: str = "",
    operation: str = "",
    capability: str = "",
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/system/config")),
) -> list[TraceSummaryOut]:
    if not ai_debug_enabled():
        raise HTTPException(status_code=403, detail="AI debug trace UI is disabled")
    query = db.query(TracerSpan)
    if status:
        query = query.filter(TracerSpan.status == status)
    if capability:
        query = query.filter(TracerSpan.capability == capability)
    if operation:
        query = query.filter(TracerSpan.operation_name.like(f"%{operation}%"))
    spans = query.order_by(TracerSpan.start_time.desc()).limit(500).all()
    grouped: dict[str, list[TracerSpan]] = {}
    for span in spans:
        grouped.setdefault(span.trace_id, []).append(span)
    summaries: list[TraceSummaryOut] = []
    for trace_id, trace_spans in grouped.items():
        root = next((span for span in trace_spans if not span.parent_span_id), trace_spans[0])
        summaries.append(
            TraceSummaryOut(
                trace_id=trace_id,
                operation_name=root.operation_name,
                capability=root.capability,
                provider=root.provider,
                model=root.model,
                status="error" if any(span.status == "error" for span in trace_spans) else root.status,
                start_time=root.start_time.isoformat(),
                duration_ms=sum(span.duration_ms for span in trace_spans if not span.parent_span_id) or root.duration_ms,
                span_count=len(trace_spans),
            )
        )
    return sorted(summaries, key=lambda item: item.start_time, reverse=True)


@app.get("/api/v1/debug/trace/{trace_id}", response_model=TraceDetailOut)
def get_trace_detail(
    trace_id: str,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/system/config")),
) -> TraceDetailOut:
    if not ai_debug_enabled():
        raise HTTPException(status_code=403, detail="AI debug trace UI is disabled")
    spans = db.query(TracerSpan).filter(TracerSpan.trace_id == trace_id).order_by(TracerSpan.start_time).all()
    if not spans:
        raise HTTPException(status_code=404, detail="Trace not found")
    return TraceDetailOut(
        trace_id=trace_id,
        spans=[
            {
                "span_id": span.span_id,
                "parent_span_id": span.parent_span_id,
                "operation_name": span.operation_name,
                "span_type": span.span_type,
                "capability": span.capability,
                "provider": span.provider,
                "model": span.model,
                "status": span.status,
                "start_time": span.start_time.isoformat(),
                "duration_ms": span.duration_ms,
                "metadata": json.loads(span.metadata_json or "{}"),
                "error": span.error,
            }
            for span in spans
        ],
    )


@app.put("/api/v1/materials/{material_id}", response_model=MaterialOut)
def update_material(
    material_id: int,
    payload: MaterialUpdate,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/materials/{material_id}")),
) -> MaterialOut:
    require_button_permission(auth, "button.material_archives.edit")
    material = db.get(Material, material_id)
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    require_library_scope(auth, material.material_library_id)
    if payload.material_library_id:
        require_library_scope(auth, payload.material_library_id)
    if payload.product_name_id or payload.material_library_id or payload.category_id:
        product, library, category = material_context_by_payload(
            db,
            payload.product_name_id or material.product_name_id,
            payload.material_library_id or material.material_library_id,
            payload.category_id or material.category_id,
        )
        material.product_name_id = product.id
        material.material_library_id = library.id
        material.category_id = category.id
    if payload.name is not None and payload.name != material.name:
        exists = (
            db.query(Material)
            .filter(Material.product_name_id == material.product_name_id, Material.name == payload.name, Material.id != material.id)
            .first()
        )
        if exists:
            raise HTTPException(status_code=409, detail="Material already exists for this product name")
        material.name = payload.name
    for field in ["unit", "description", "enabled"]:
        value = getattr(payload, field)
        if value is not None:
            setattr(material, field, value)
    if payload.brand_id is not None:
        brand = db.get(Brand, payload.brand_id) if payload.brand_id else None
        if payload.brand_id and not brand:
            raise HTTPException(status_code=404, detail="Brand not found")
        material.brand_id = brand.id if brand else None
    if payload.attributes is not None:
        existing_history = material_attributes(material.attributes).get("_lifecycle_history")
        attributes = dict(payload.attributes)
        if existing_history and "_lifecycle_history" not in attributes:
            attributes["_lifecycle_history"] = existing_history
        material.attributes = json.dumps(attributes, ensure_ascii=False)
    if payload.status is not None:
        enforce_material_transition(material.status, payload.status, payload.transition_reason)
        if material.status != payload.status:
            record_material_lifecycle(
                material,
                material.status,
                payload.status,
                payload.transition_reason or "",
                "manual",
                "super_admin",
            )
        material.status = payload.status
    material.updated_at = now()
    db.commit()
    db.refresh(material)
    return material_to_out(material)


@app.get("/api/v1/materials/{material_id}", response_model=MaterialOut)
def get_material(
    material_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/materials/{material_id}")),
) -> MaterialOut:
    material = db.get(Material, material_id)
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    require_library_scope(auth, material.material_library_id)
    return material_to_out(material)


@app.patch("/api/v1/materials/{material_id}/stop-purchase", response_model=MaterialOut)
def admin_stop_purchase_material(
    material_id: int,
    payload: ManualStopPurchaseIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PATCH./api/v1/materials/{material_id}/stop-purchase")),
) -> MaterialOut:
    require_button_permission(auth, "button.material_archives.approval")
    material = db.get(Material, material_id)
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    require_library_scope(auth, material.material_library_id)
    reason = payload.reason.strip()
    if not reason:
        raise HTTPException(status_code=422, detail="An exemption reason is required for manual stop purchase")
    if material.status != "normal":
        raise HTTPException(status_code=409, detail="Manual stop purchase requires a material in normal status")
    actor = payload.actor.strip() or "super_admin"
    application = WorkflowApplication(
        application_no=application_no(f"manual-stop-purchase:{material.id}:{actor}:{now().isoformat()}"),
        type="stop_purchase",
        status="approved",
        applicant=actor,
        current_node="approved",
        business_reason=reason,
        payload=json.dumps(
            {
                **material_summary(material),
                "reason_code": reason,
                "reason": reason,
                "business_reason": reason,
                "from_status": "normal",
                "target_status": "stop_purchase",
                "source": "admin_manual",
                "exemption_reason": reason,
            },
            ensure_ascii=False,
        ),
        created_resource_type="material",
        created_resource_id=material.id,
    )
    db.add(application)
    db.flush()
    add_workflow_history(application, "manual_stop_purchase", actor, "super_admin", "normal", "stop_purchase", reason)
    material.status = "stop_purchase"
    material.updated_at = now()
    record_material_lifecycle(material, "normal", "stop_purchase", reason, "admin_manual", actor, application.application_no)
    data = workflow_payload(application.payload)
    data["current_material_status"] = "stop_purchase"
    data["approved_at"] = material.updated_at.isoformat()
    application.payload = json.dumps(data, ensure_ascii=False)
    db.commit()
    db.refresh(material)
    return material_to_out(material)


@app.post("/api/v1/materials/{material_id}/transition", response_model=MaterialOut)
def transition_material(
    material_id: int,
    payload: MaterialTransitionIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials/{material_id}/transition")),
) -> MaterialOut:
    require_button_permission(auth, "button.material_archives.edit")
    material = db.get(Material, material_id)
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    require_library_scope(auth, material.material_library_id)
    enforce_material_transition(material.status, payload.target_status, payload.reason)
    if material.status != payload.target_status:
        record_material_lifecycle(material, material.status, payload.target_status, payload.reason, "manual", "super_admin")
    material.status = payload.target_status
    material.updated_at = now()
    db.commit()
    db.refresh(material)
    return material_to_out(material)


@app.delete("/api/v1/materials/{material_id}")
def delete_material(
    material_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.DELETE./api/v1/materials/{material_id}")),
) -> dict[str, Any]:
    require_button_permission(auth, "button.material_archives.delete")
    material = db.get(Material, material_id)
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    require_library_scope(auth, material.material_library_id)
    db.delete(material)
    db.commit()
    return {"deleted": True, "id": material_id}


@app.post(
    "/api/v1/materials/governance/preview",
    description="AI material governance preview. capability: material_governance",
)
def preview_material_governance(
    payload: MaterialGovernancePreviewIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials/governance/preview")),
) -> dict[str, Any]:
    if payload.material_library_id:
        require_library_scope(auth, payload.material_library_id)
    items = material_governance_items(payload, db)
    return {"capability": "material_governance", "items": items, "count": len(items)}


@app.post(
    "/api/v1/materials/governance/import",
    response_model=list[MaterialOut],
    description="AI material governance batch confirmation import. capability: material_governance",
)
def import_material_governance(
    payload: MaterialGovernanceImportIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials/governance/import")),
) -> list[MaterialOut]:
    require_button_permission(auth, "button.material_archives.import")
    if payload.material_library_id:
        require_library_scope(auth, payload.material_library_id)
    product = product_by_payload(db, payload.product_name_id, payload.product_name) if payload.product_name_id or payload.product_name else ensure_seed_product(db)
    default_library, default_category = ensure_seed_material_context(db)
    imported: list[Material] = []
    for item in payload.items:
        if item.get("validation_status") == "invalid" or item.get("errors"):
            raise HTTPException(status_code=422, detail="Invalid preview rows cannot be imported")
        name = str(item.get("name", "")).strip()
        if not name:
            raise HTTPException(status_code=422, detail="Material name is required")
        item_product = db.get(ProductName, int(item.get("product_name_id") or product.id)) or product
        library = db.get(MaterialLibrary, int(item.get("material_library_id") or payload.material_library_id or default_library.id)) or default_library
        require_library_scope(auth, library.id)
        category = db.get(Category, int(item.get("category_id") or payload.category_id or default_category.id)) or default_category
        brand_id = item.get("brand_id")
        brand_name = str(item.get("brand_name", "")).strip()
        brand = db.get(Brand, int(brand_id)) if brand_id else None
        if not brand and brand_name:
            brand = db.query(Brand).filter(Brand.name == brand_name).first()
            if not brand:
                brand = Brand(
                    code=next_unique_code(db, Brand, "BRAND", brand_name),
                    name=brand_name,
                    description="Created during material governance import",
                    enabled=True,
                )
                db.add(brand)
                db.flush()
        existing = db.query(Material).filter(Material.product_name_id == item_product.id, Material.name == name).first()
        if existing:
            imported.append(existing)
            continue
        material = Material(
            code=next_unique_code(db, Material, "MAT", f"{item_product.id}:{name}:{item.get('source_row')}:{now().isoformat()}"),
            name=name,
            product_name_id=item_product.id,
            material_library_id=library.id,
            category_id=category.id,
            unit=str(item.get("unit") or item_product.unit),
            brand_id=brand.id if brand else None,
            status="normal",
            description=str(item.get("description", "")),
            attributes=json.dumps(material_attributes(item.get("attributes")), ensure_ascii=False),
            enabled=True,
        )
        db.add(material)
        db.flush()
        imported.append(material)
    db.commit()
    for material in imported:
        db.refresh(material)
    return [material_to_out(material) for material in imported]


@app.post("/api/v1/ai/material-governance/preview")
def preview_ai_material_governance(
    payload: MaterialGovernancePreviewIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials/governance/preview")),
) -> dict[str, Any]:
    return preview_material_governance(payload, db, auth)


@app.post("/api/v1/ai/material-governance/import", response_model=list[MaterialOut])
def import_ai_material_governance(
    payload: MaterialGovernanceImportIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/materials/governance/import")),
) -> list[MaterialOut]:
    return import_material_governance(payload, db, auth)


@app.get("/api/v1/attributes/changes", response_model=list[ChangeOut])
def list_attribute_changes(
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/attributes/changes")),
) -> list[ChangeOut]:
    changes = db.query(AttributeChange).order_by(AttributeChange.id.desc()).all()
    return [change_to_out(change) for change in changes]


@app.post("/api/v1/attributes/governance/preview")
def preview_attribute_governance(
    payload: GovernancePreviewIn,
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/attributes/governance/preview")),
) -> dict[str, Any]:
    items = governance_items(payload.rows)
    return {"items": items, "count": len(items)}


@app.post("/api/v1/ai/attribute-governance/preview")
def preview_ai_attribute_governance(
    payload: GovernancePreviewIn,
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/attributes/governance/preview")),
) -> dict[str, Any]:
    return preview_attribute_governance(payload, auth)


@app.post("/api/v1/attributes/governance/import", response_model=list[AttributeOut])
def import_attribute_governance(
    payload: GovernanceImportIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/attributes/governance/import")),
) -> list[AttributeOut]:
    require_button_permission(auth, "button.attribute_management.import")
    product = product_by_payload(db, payload.product_name_id, payload.product_name)
    imported: list[Attribute] = []
    for item in payload.items:
        existing = (
            db.query(Attribute)
            .filter(Attribute.product_name_id == product.id, Attribute.name == item.get("name", ""))
            .first()
        )
        if existing:
            imported.append(existing)
            continue
        attribute = Attribute(
            code=next_unique_code(db, Attribute, "ATTR", f"{product.id}:{item.get('name')}:{item.get('source_row')}"),
            product_name_id=product.id,
            name=str(item.get("name", "")),
            data_type=str(item.get("data_type", "text")),
            unit=str(item.get("unit", "")),
            required=bool(item.get("required", False)),
            default_value=str(item.get("default_value", "")),
            options=",".join(normalize_options(item.get("options", []))),
            description=str(item.get("description", "")),
            source=str(item.get("source", "AI governance import")),
        )
        db.add(attribute)
        db.flush()
        add_change(attribute, ["created"], {}, snapshot(attribute), "AI governance")
        imported.append(attribute)
    db.commit()
    for attribute in imported:
        db.refresh(attribute)
    return [attribute_to_out(attribute) for attribute in imported]


@app.post("/api/v1/ai/attribute-governance/import", response_model=list[AttributeOut])
def import_ai_attribute_governance(
    payload: GovernanceImportIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/attributes/governance/import")),
) -> list[AttributeOut]:
    return import_attribute_governance(payload, db, auth)


@app.get("/api/v1/attributes", response_model=list[AttributeOut])
def list_attributes(
    product_name_id: int | None = None,
    search: str = "",
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/attributes")),
) -> list[AttributeOut]:
    query = db.query(Attribute).join(ProductName)
    if product_name_id:
        query = query.filter(Attribute.product_name_id == product_name_id)
    if search:
        like = f"%{search}%"
        query = query.filter(or_(Attribute.name.like(like), Attribute.code.like(like), Attribute.options.like(like)))
    attributes = query.order_by(Attribute.id.desc()).all()
    return [attribute_to_out(attribute) for attribute in attributes]


@app.post("/api/v1/attributes", response_model=AttributeOut)
def create_attribute(
    payload: AttributeIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/attributes")),
) -> AttributeOut:
    require_button_permission(auth, "button.attribute_management.create")
    product = product_by_payload(db, payload.product_name_id, payload.product_name)
    attribute = Attribute(
        code=next_unique_code(db, Attribute, "ATTR", f"{product.id}:{payload.name}:{func.now()}"),
        product_name_id=product.id,
        name=payload.name,
        data_type=payload.data_type,
        unit=payload.unit,
        required=payload.required,
        default_value=payload.default_value,
        options=",".join(normalize_options(payload.options)),
        description=payload.description,
        source=payload.source,
    )
    db.add(attribute)
    db.flush()
    add_change(attribute, ["created"], {}, snapshot(attribute), "super_admin")
    db.commit()
    db.refresh(attribute)
    return attribute_to_out(attribute)


@app.put("/api/v1/attributes/{attribute_id}", response_model=AttributeOut)
def update_attribute(
    attribute_id: int,
    payload: AttributeUpdate,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/attributes/{attribute_id}")),
) -> AttributeOut:
    require_button_permission(auth, "button.attribute_management.edit")
    attribute = db.get(Attribute, attribute_id)
    if not attribute:
        raise HTTPException(status_code=404, detail="Attribute not found")
    before = snapshot(attribute)
    changed: list[str] = []
    for field in ["name", "data_type", "unit", "required", "default_value", "description", "source", "enabled"]:
        value = getattr(payload, field)
        if value is not None and getattr(attribute, field) != value:
            setattr(attribute, field, value)
            changed.append(field)
    if payload.options is not None:
        options = ",".join(normalize_options(payload.options))
        if attribute.options != options:
            attribute.options = options
            changed.append("options")
    if changed:
        attribute.version += 1
        attribute.updated_at = now()
        add_change(attribute, changed, {key: before[key] for key in changed}, {key: snapshot(attribute)[key] for key in changed}, "super_admin")
    db.commit()
    db.refresh(attribute)
    return attribute_to_out(attribute)


@app.delete("/api/v1/attributes/{attribute_id}")
def delete_attribute(
    attribute_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.DELETE./api/v1/attributes/{attribute_id}")),
) -> dict[str, Any]:
    require_button_permission(auth, "button.attribute_management.delete")
    attribute = db.get(Attribute, attribute_id)
    if not attribute:
        raise HTTPException(status_code=404, detail="Attribute not found")
    db.delete(attribute)
    db.commit()
    return {"deleted": True, "id": attribute_id}


@app.get("/api/v1/attributes/{attribute_id}/changes", response_model=list[ChangeOut])
def get_attribute_changes(
    attribute_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/attributes/{attribute_id}/changes")),
) -> list[ChangeOut]:
    changes = (
        db.query(AttributeChange)
        .filter(AttributeChange.attribute_id == attribute_id)
        .order_by(AttributeChange.id.desc())
        .all()
    )
    return [change_to_out(change) for change in changes]


@app.post("/api/v1/ai/attribute-recommend")
def recommend_attributes(
    payload: RecommendIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/ai/attribute-recommend")),
) -> dict[str, Any]:
    product = product_by_payload(db, payload.product_name_id, payload.product_name)
    recommendations = [
        {
            "name": "打印速度",
            "data_type": "number",
            "unit": "页/分钟",
            "required": True,
            "default_value": "30",
            "options": [],
            "confidence": 0.95,
            "source": "category common attributes",
            "reason": f"{product.category}常用性能指标",
        },
        {
            "name": "颜色模式",
            "data_type": "enum",
            "unit": "",
            "required": True,
            "default_value": "彩色",
            "options": ["黑白", "彩色"],
            "confidence": 0.91,
            "source": "historical data",
            "reason": "同类物料历史属性高频出现",
        },
        {
            "name": "纸张尺寸",
            "data_type": "enum",
            "unit": "",
            "required": False,
            "default_value": "A4",
            "options": ["A4", "A5"],
            "confidence": 0.88,
            "source": "standard references",
            "reason": "办公打印设备标准属性",
        },
    ]
    return {"capability": "attr_recommend", "product_name": product.name, "recommendations": recommendations}


@app.get("/api/v1/brands", response_model=list[BrandOut])
def list_brands(
    search: str = "",
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.GET./api/v1/brands")),
) -> list[BrandOut]:
    query = db.query(Brand)
    if search:
        like = f"%{search}%"
        query = query.filter(or_(Brand.name.like(like), Brand.code.like(like), Brand.description.like(like)))
    brands = query.order_by(Brand.id.desc()).all()
    return [brand_to_out(brand) for brand in brands]


@app.post("/api/v1/brands", response_model=BrandOut)
def create_brand(
    payload: BrandIn,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.POST./api/v1/brands")),
) -> BrandOut:
    existing = db.query(Brand).filter(Brand.name == payload.name).first()
    if existing:
        raise HTTPException(status_code=409, detail="Brand already exists")
    brand = Brand(
        code=next_unique_code(db, Brand, "BRAND", payload.name),
        name=payload.name,
        description=payload.description,
        enabled=payload.enabled,
    )
    logo_to_model(brand, payload.logo)
    db.add(brand)
    db.commit()
    db.refresh(brand)
    return brand_to_out(brand)


@app.put("/api/v1/brands/{brand_id}", response_model=BrandOut)
def update_brand(
    brand_id: int,
    payload: BrandUpdate,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.PUT./api/v1/brands/{brand_id}")),
) -> BrandOut:
    brand = db.get(Brand, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    for field in ["name", "description", "enabled"]:
        value = getattr(payload, field)
        if value is not None:
            setattr(brand, field, value)
    if payload.logo is not None:
        logo_to_model(brand, payload.logo)
    brand.updated_at = now()
    db.commit()
    db.refresh(brand)
    return brand_to_out(brand)


@app.delete("/api/v1/brands/{brand_id}")
def delete_brand(
    brand_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_api_permission("api.DELETE./api/v1/brands/{brand_id}")),
) -> dict[str, Any]:
    brand = db.get(Brand, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    db.delete(brand)
    db.commit()
    return {"deleted": True, "id": brand_id}


def snapshot(attribute: Attribute) -> dict[str, Any]:
    return {
        "name": attribute.name,
        "data_type": attribute.data_type,
        "unit": attribute.unit,
        "required": attribute.required,
        "default_value": attribute.default_value,
        "options": normalize_options(attribute.options),
        "description": attribute.description,
        "source": attribute.source,
        "enabled": attribute.enabled,
    }


def add_change(
    attribute: Attribute,
    fields: list[str],
    before_values: dict[str, Any],
    after_values: dict[str, Any],
    operator: str,
) -> None:
    attribute.changes.append(
        AttributeChange(
            attribute_code=attribute.code,
            attribute_name=attribute.name,
            version=attribute.version,
            operator=operator,
            changed_fields=json.dumps(fields, ensure_ascii=False),
            before_values=json.dumps(before_values, ensure_ascii=False),
            after_values=json.dumps(after_values, ensure_ascii=False),
        )
    )
